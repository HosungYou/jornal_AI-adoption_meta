#!/usr/bin/env python3
"""
Prepare coder workbooks for MASEM coding:
1. Assign study_ids (S001-S225) to confirmed includes
2. Pre-populate STUDY_METADATA in master workbook
3. Select 10 calibration studies (Phase 0) with stratified diversity
4. Assign 100 Phase 1 studies to pairs (50 each)
5. Assign ~125 Phase 2 studies to individual coders (~31 each)
6. Create 4 coder-specific workbooks (R1-R4)
7. Save study_id mapping CSV for reference
"""

import csv
import random
import shutil
from pathlib import Path
from datetime import date

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Paths ──
BASE = Path("/Volumes/External SSD/Projects/Meta-Analysis/jornal_AI-adoption_meta")
CONFIRMED = BASE / "data/02_screening/confirmed_includes.csv"
TEMPLATE = BASE / "data/04_extraction/AI_Adoption_MASEM_Coding_v2.xlsx"
OUTPUT_DIR = BASE / "data/04_extraction/coder_workbooks"
MAPPING_CSV = BASE / "data/04_extraction/study_id_mapping.csv"

random.seed(2026)  # Reproducible


def load_confirmed():
    """Load confirmed includes and assign study_ids."""
    with open(CONFIRMED) as f:
        rows = list(csv.DictReader(f))
    # Sort by year then record_id for stable ordering
    rows.sort(key=lambda r: (r["year"], r["record_id"]))
    for i, row in enumerate(rows, 1):
        row["study_id"] = f"S{i:03d}"
    return rows


def extract_doi(doi_url):
    """Extract DOI from URL."""
    if doi_url and "doi.org/" in doi_url:
        return doi_url.split("doi.org/", 1)[1]
    return doi_url or ""


def select_calibration(studies, n=10):
    """
    Select 10 calibration studies with stratified diversity:
    - Year coverage (2022-2026)
    - Title keyword diversity (generative, chatbot, UTAUT, TAM, etc.)
    """
    # Stratify by year
    by_year = {}
    for s in studies:
        y = s["year"]
        by_year.setdefault(y, []).append(s)

    selected = []

    # Take 1 from 2022 (rare), 1 from 2023, 3 from 2024, 4 from 2025, 1 from 2026
    allocation = {"2022": 1, "2023": 1, "2024": 3, "2025": 4, "2026": 1}

    for year, count in allocation.items():
        pool = by_year.get(year, [])
        if len(pool) >= count:
            # Prefer keyword diversity within each year
            chosen = random.sample(pool, min(count, len(pool)))
            selected.extend(chosen)
        elif pool:
            selected.extend(pool)

    # If not enough, top up from 2025 (largest pool)
    while len(selected) < n:
        remaining = [s for s in by_year.get("2025", []) if s not in selected]
        if remaining:
            selected.append(random.choice(remaining))
        else:
            break

    return selected[:n]


def assign_phases(studies, calibration_ids):
    """
    Assign studies to phases:
    - Phase 0: calibration (10 studies, all 4 coders)
    - Phase 1: dual coding (100 studies, 50 per pair)
    - Phase 2: single coding (remaining ~115, ~29 per coder)
    """
    remaining = [s for s in studies if s["study_id"] not in calibration_ids]
    random.shuffle(remaining)

    phase1 = remaining[:100]
    phase2 = remaining[100:]

    # Phase 1: split into Pair A (50) and Pair B (50)
    pair_a = phase1[:50]
    pair_b = phase1[50:100]

    # Phase 2: split into 4 roughly equal groups
    n = len(phase2)
    q = n // 4
    r = n % 4
    p2_r1 = phase2[0:q + (1 if r > 0 else 0)]
    offset = len(p2_r1)
    p2_r2 = phase2[offset:offset + q + (1 if r > 1 else 0)]
    offset += len(p2_r2)
    p2_r3 = phase2[offset:offset + q + (1 if r > 2 else 0)]
    offset += len(p2_r3)
    p2_r4 = phase2[offset:]

    return {
        "pair_a": pair_a,
        "pair_b": pair_b,
        "p2_r1": p2_r1,
        "p2_r2": p2_r2,
        "p2_r3": p2_r3,
        "p2_r4": p2_r4,
    }


def populate_metadata_sheet(ws, studies_for_coder, coder_label, phase_info):
    """Fill STUDY_METADATA sheet with pre-populated data."""
    # Header row is already row 1
    for idx, study in enumerate(studies_for_coder, 2):
        ws.cell(row=idx, column=1, value=study["study_id"])           # study_id
        ws.cell(row=idx, column=2, value="")                          # first_author (to be coded)
        ws.cell(row=idx, column=3, value=int(study["year"]))          # year
        ws.cell(row=idx, column=4, value=study["title"])              # title
        ws.cell(row=idx, column=5, value=extract_doi(study["doi_url"]))  # doi
        ws.cell(row=idx, column=21, value=coder_label)                # human_coder
        ws.cell(row=idx, column=25, value=phase_info.get(study["study_id"], ""))  # coding_phase

        # Apply light formatting
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=idx, column=col)
            cell.font = Font(name="Calibri", size=10)
            cell.alignment = Alignment(vertical="center", wrap_text=True)

    # Auto-width for key columns
    for col_idx in [1, 3, 4, 5]:
        max_len = 0
        col_letter = get_column_letter(col_idx)
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=col_idx, max_col=col_idx):
            for cell in row:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
        if col_idx == 4:  # title column - cap width
            ws.column_dimensions[col_letter].width = min(60, max_len + 2)
        else:
            ws.column_dimensions[col_letter].width = min(30, max_len + 2)


def style_header_row(ws):
    """Apply consistent header styling."""
    header_fill = PatternFill(start_color="1B3A5C", end_color="1B3A5C", fill_type="solid")
    header_font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
    thin_border = Border(
        bottom=Side(style="thin", color="B4C7DC"),
    )
    for col in range(1, ws.max_column + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border


def add_assignment_sheet(wb, calibration, phases, coder_label):
    """Add an ASSIGNMENT sheet showing which studies to code and in what order."""
    ws = wb.create_sheet("ASSIGNMENT", 0)  # First sheet

    # Title
    ws.merge_cells("A1:E1")
    cell = ws.cell(row=1, column=1, value=f"Coding Assignment — {coder_label}")
    cell.font = Font(name="Calibri", size=14, bold=True, color="1B3A5C")
    cell.alignment = Alignment(horizontal="center")

    ws.cell(row=2, column=1, value=f"Generated: {date.today().isoformat()}")
    ws.cell(row=2, column=1).font = Font(name="Calibri", size=9, color="888888")

    # Headers
    row = 4
    headers = ["Study ID", "Title", "Year", "Phase", "Status"]
    header_fill = PatternFill(start_color="1B3A5C", end_color="1B3A5C", fill_type="solid")
    for i, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=i, value=h)
        cell.fill = header_fill
        cell.font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center")

    row = 5

    # Phase 0 - Calibration (highlighted)
    cal_fill = PatternFill(start_color="FFF3E0", end_color="FFF3E0", fill_type="solid")
    for s in calibration:
        ws.cell(row=row, column=1, value=s["study_id"])
        ws.cell(row=row, column=2, value=s["title"][:80])
        ws.cell(row=row, column=3, value=int(s["year"]))
        ws.cell(row=row, column=4, value="Phase 0: Calibration")
        ws.cell(row=row, column=5, value="")
        for col in range(1, 6):
            ws.cell(row=row, column=col).fill = cal_fill
            ws.cell(row=row, column=col).font = Font(name="Calibri", size=10)
        row += 1

    # Phase 1 - Dual coding (light blue)
    p1_fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
    phase1_studies = phases.get("phase1", [])
    for s in phase1_studies:
        ws.cell(row=row, column=1, value=s["study_id"])
        ws.cell(row=row, column=2, value=s["title"][:80])
        ws.cell(row=row, column=3, value=int(s["year"]))
        pair_label = phases.get("pair_label", "Phase 1: Dual")
        ws.cell(row=row, column=4, value=pair_label)
        ws.cell(row=row, column=5, value="")
        for col in range(1, 6):
            ws.cell(row=row, column=col).fill = p1_fill
            ws.cell(row=row, column=col).font = Font(name="Calibri", size=10)
        row += 1

    # Phase 2 - Single coding (white/alternating)
    phase2_studies = phases.get("phase2", [])
    alt_fill = PatternFill(start_color="F2F7FB", end_color="F2F7FB", fill_type="solid")
    for i, s in enumerate(phase2_studies):
        ws.cell(row=row, column=1, value=s["study_id"])
        ws.cell(row=row, column=2, value=s["title"][:80])
        ws.cell(row=row, column=3, value=int(s["year"]))
        ws.cell(row=row, column=4, value="Phase 2: Single")
        ws.cell(row=row, column=5, value="")
        for col in range(1, 6):
            ws.cell(row=row, column=col).font = Font(name="Calibri", size=10)
            if i % 2 == 1:
                ws.cell(row=row, column=col).fill = alt_fill
        row += 1

    # Summary
    row += 1
    ws.cell(row=row, column=1, value="Summary:")
    ws.cell(row=row, column=1).font = Font(name="Calibri", size=10, bold=True)
    row += 1
    ws.cell(row=row, column=1, value=f"  Phase 0 (Calibration): {len(calibration)} studies")
    row += 1
    ws.cell(row=row, column=1, value=f"  Phase 1 (Dual Coding): {len(phase1_studies)} studies")
    row += 1
    ws.cell(row=row, column=1, value=f"  Phase 2 (Single Coding): {len(phase2_studies)} studies")
    row += 1
    total = len(calibration) + len(phase1_studies) + len(phase2_studies)
    ws.cell(row=row, column=1, value=f"  Total: {total} studies")
    ws.cell(row=row, column=1).font = Font(name="Calibri", size=10, bold=True)

    # Column widths
    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 65
    ws.column_dimensions["C"].width = 8
    ws.column_dimensions["D"].width = 22
    ws.column_dimensions["E"].width = 12


def create_coder_workbook(template_path, output_path, coder_label, calibration,
                          phase1_studies, phase2_studies, pair_label, all_studies, phase_map):
    """Create a coder-specific workbook."""
    shutil.copy2(template_path, output_path)
    wb = openpyxl.load_workbook(output_path)

    # Combine all studies for this coder
    coder_studies = calibration + phase1_studies + phase2_studies

    # Populate STUDY_METADATA
    ws = wb["STUDY_METADATA"]
    style_header_row(ws)
    populate_metadata_sheet(ws, coder_studies, coder_label, phase_map)

    # Style other sheet headers too
    for sheet_name in wb.sheetnames:
        if sheet_name != "CODEBOOK":
            style_header_row(wb[sheet_name])

    # Add ASSIGNMENT sheet
    add_assignment_sheet(wb, calibration, {
        "phase1": phase1_studies,
        "pair_label": pair_label,
        "phase2": phase2_studies,
    }, coder_label)

    wb.save(output_path)
    total = len(coder_studies)
    print(f"  {coder_label}: {total} studies "
          f"(Cal:{len(calibration)} + P1:{len(phase1_studies)} + P2:{len(phase2_studies)})"
          f" -> {output_path.name}")


def main():
    # Load data
    studies = load_confirmed()
    print(f"Loaded {len(studies)} confirmed includes (S001-S{len(studies):03d})")

    # Select calibration studies
    calibration = select_calibration(studies, n=10)
    cal_ids = {s["study_id"] for s in calibration}
    print(f"\nPhase 0 Calibration ({len(calibration)} studies):")
    for s in calibration:
        print(f"  {s['study_id']}  [{s['year']}]  {s['title'][:70]}...")

    # Assign phases
    phases = assign_phases(studies, cal_ids)
    print(f"\nPhase 1 Dual Coding: {len(phases['pair_a']) + len(phases['pair_b'])} studies")
    print(f"  Pair A (R1+R2): {len(phases['pair_a'])} studies")
    print(f"  Pair B (R3+R4): {len(phases['pair_b'])} studies")
    print(f"\nPhase 2 Single Coding: {len(phases['p2_r1']) + len(phases['p2_r2']) + len(phases['p2_r3']) + len(phases['p2_r4'])} studies")
    print(f"  R1: {len(phases['p2_r1'])} | R2: {len(phases['p2_r2'])} | R3: {len(phases['p2_r3'])} | R4: {len(phases['p2_r4'])}")

    # Build phase map (study_id -> phase label)
    phase_map = {}
    for s in calibration:
        phase_map[s["study_id"]] = "Phase 0: Calibration"
    for s in phases["pair_a"]:
        phase_map[s["study_id"]] = "Phase 1: Pair A"
    for s in phases["pair_b"]:
        phase_map[s["study_id"]] = "Phase 1: Pair B"
    for s in phases["p2_r1"] + phases["p2_r2"] + phases["p2_r3"] + phases["p2_r4"]:
        phase_map[s["study_id"]] = "Phase 2: Single"

    # Create output directory
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    print(f"\nCreating coder workbooks in: {OUTPUT_DIR}")

    # R1: Calibration (10) + Pair A (50) + Phase 2 (~29)
    create_coder_workbook(
        TEMPLATE, OUTPUT_DIR / "AI_Adoption_MASEM_Coding_v2_R1.xlsx",
        "R1", calibration, phases["pair_a"], phases["p2_r1"],
        "Phase 1: Pair A (R1+R2)", studies, phase_map,
    )

    # R2: Calibration (10) + Pair A (50) + Phase 2 (~29)
    create_coder_workbook(
        TEMPLATE, OUTPUT_DIR / "AI_Adoption_MASEM_Coding_v2_R2.xlsx",
        "R2", calibration, phases["pair_a"], phases["p2_r2"],
        "Phase 1: Pair A (R1+R2)", studies, phase_map,
    )

    # R3: Calibration (10) + Pair B (50) + Phase 2 (~29)
    create_coder_workbook(
        TEMPLATE, OUTPUT_DIR / "AI_Adoption_MASEM_Coding_v2_R3.xlsx",
        "R3", calibration, phases["pair_b"], phases["p2_r3"],
        "Phase 1: Pair B (R3+R4)", studies, phase_map,
    )

    # R4: Calibration (10) + Pair B (50) + Phase 2 (~28)
    create_coder_workbook(
        TEMPLATE, OUTPUT_DIR / "AI_Adoption_MASEM_Coding_v2_R4.xlsx",
        "R4", calibration, phases["pair_b"], phases["p2_r4"],
        "Phase 1: Pair B (R3+R4)", studies, phase_map,
    )

    # Save study_id mapping CSV
    with open(MAPPING_CSV, "w", newline="") as f:
        writer = csv.writer(f)
        writer.writerow(["study_id", "record_id", "title", "year", "doi", "phase",
                         "pair_a_coders", "pair_b_coders", "single_coder"])
        for s in studies:
            sid = s["study_id"]
            phase = phase_map.get(sid, "")
            pair_a = "R1, R2" if sid in {x["study_id"] for x in phases["pair_a"]} else ""
            pair_b = "R3, R4" if sid in {x["study_id"] for x in phases["pair_b"]} else ""
            single = ""
            if sid in {x["study_id"] for x in phases["p2_r1"]}:
                single = "R1"
            elif sid in {x["study_id"] for x in phases["p2_r2"]}:
                single = "R2"
            elif sid in {x["study_id"] for x in phases["p2_r3"]}:
                single = "R3"
            elif sid in {x["study_id"] for x in phases["p2_r4"]}:
                single = "R4"
            cal = "R1, R2, R3, R4" if sid in cal_ids else ""

            writer.writerow([
                sid, s["record_id"], s["title"], s["year"],
                extract_doi(s["doi_url"]), phase,
                pair_a or cal, pair_b or cal, single,
            ])

    print(f"\nStudy ID mapping saved: {MAPPING_CSV.name}")
    print(f"\nDone! Files created:")
    for f in sorted(OUTPUT_DIR.glob("*.xlsx")):
        print(f"  {f.name}")


if __name__ == "__main__":
    main()
