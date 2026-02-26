#!/usr/bin/env python3
"""
human_review_sheet_v5.xlsx → CSV 변환 스크립트

인간 리뷰어가 v5 Excel에 코딩을 완료한 후 실행하면:
1. screening_ai_dual.csv의 human 컬럼 업데이트
2. human_review_queue.csv의 human 컬럼 업데이트
3. IRR 통계 산출 (Cohen's κ)
4. PRISMA 숫자 업데이트

사용법:
    python3 convert_screening.py [--dry-run] [--irr-only]
"""

import argparse
import csv
import os
import sys
from collections import Counter
from openpyxl import load_workbook

# ── 경로 설정 ──────────────────────────────────────────
REPO_ROOT = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
EXCEL_PATH = os.path.join(REPO_ROOT, "data", "templates", "human_review_sheet_v5.xlsx")
DUAL_CSV = os.path.join(REPO_ROOT, "data", "03_screening", "screening_ai_dual.csv")
QUEUE_CSV = os.path.join(REPO_ROOT, "data", "03_screening", "human_review_queue.csv")
OUTPUT_DIR = os.path.join(REPO_ROOT, "paper_a", "data", "02_screening", "human_verification")

# v5 컬럼 매핑 (1-indexed)
V5_COLS = {
    "id": 1,
    "title": 2,
    "year": 3,
    "r1_decision": None,  # 동적 계산
    "r1_code": None,
    "r1_memo": None,
    "r2_decision": None,
    "r2_code": None,
    "r2_memo": None,
    "final_decision": None,
    "final_code": None,
    "final_memo": None,
}

# 판단값 → CSV 변환
DECISION_TO_CSV = {
    "O": "include",
    "X": "exclude",
    "?": "uncertain",
    "Include": "include",
    "Exclude": "exclude",
    "Uncertain": "uncertain",
    "": "",
}


def detect_columns(ws):
    """v5 시트에서 컬럼 위치 자동 감지"""
    headers = {}
    for col in range(1, ws.max_column + 1):
        val = ws.cell(row=1, column=col).value or ""
        val_clean = val.replace("\n", "").strip()
        headers[val_clean] = col
    return headers


def read_v5_sheet(ws):
    """v5 시트에서 리뷰어 데이터 읽기"""
    cols = detect_columns(ws)
    records = []

    # 컬럼 매핑
    id_col = cols.get("ID", 1)
    r1d = cols.get("리뷰어1판단", None)
    r1c = cols.get("리뷰어1제외코드", None)
    r1m = cols.get("리뷰어1메모", None)
    r2d = cols.get("리뷰어2판단", None)
    r2c = cols.get("리뷰어2제외코드", None)
    r2m = cols.get("리뷰어2메모", None)
    fd = cols.get("최종판단", None)
    fc = cols.get("최종코드", None)
    fm = cols.get("메모", None)

    for row in range(2, ws.max_row + 1):
        rec_id = ws.cell(row=row, column=id_col).value
        if not rec_id or not str(rec_id).startswith("REC_"):
            continue

        record = {
            "id": str(rec_id),
            "r1_decision": str(ws.cell(row=row, column=r1d).value or "") if r1d else "",
            "r1_code": str(ws.cell(row=row, column=r1c).value or "") if r1c else "",
            "r1_memo": str(ws.cell(row=row, column=r1m).value or "") if r1m else "",
            "r2_decision": str(ws.cell(row=row, column=r2d).value or "") if r2d else "",
            "r2_code": str(ws.cell(row=row, column=r2c).value or "") if r2c else "",
            "r2_memo": str(ws.cell(row=row, column=r2m).value or "") if r2m else "",
            "final_decision": str(ws.cell(row=row, column=fd).value or "") if fd else "",
            "final_code": str(ws.cell(row=row, column=fc).value or "") if fc else "",
        }
        records.append(record)

    return records


def compute_irr(records):
    """Cohen's κ 산출 (리뷰어1 vs 리뷰어2)"""
    # 양쪽 모두 판단이 있는 레코드만
    paired = []
    for r in records:
        d1 = DECISION_TO_CSV.get(r["r1_decision"], "")
        d2 = DECISION_TO_CSV.get(r["r2_decision"], "")
        if d1 and d2:
            paired.append((d1, d2))

    if not paired:
        return None, 0, {}

    n = len(paired)
    categories = sorted(set(d for pair in paired for d in pair))

    # 일치 행렬
    matrix = Counter(paired)
    agreement = sum(matrix[(c, c)] for c in categories)
    p_o = agreement / n  # 관찰된 일치율

    # 기대 일치율
    p_e = 0
    for c in categories:
        p1 = sum(1 for d1, _ in paired if d1 == c) / n
        p2 = sum(1 for _, d2 in paired if d2 == c) / n
        p_e += p1 * p2

    # Cohen's κ
    if p_e == 1:
        kappa = 1.0
    else:
        kappa = (p_o - p_e) / (1 - p_e)

    stats = {
        "n_paired": n,
        "agreement_pct": round(p_o * 100, 1),
        "kappa": round(kappa, 3),
        "matrix": dict(matrix),
        "categories": categories,
    }

    return kappa, n, stats


def update_csv(csv_path, review_data, dry_run=False):
    """CSV 파일의 human 컬럼 업데이트"""
    if not os.path.exists(csv_path):
        print(f"  CSV not found: {csv_path}")
        return 0

    # 리뷰 데이터를 ID로 인덱싱
    by_id = {r["id"]: r for r in review_data}

    # CSV 읽기
    with open(csv_path, "r", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        fieldnames = reader.fieldnames
        rows = list(reader)

    updated = 0
    for row in rows:
        rid = row.get("record_id", "")
        if rid in by_id:
            rev = by_id[rid]
            d1 = DECISION_TO_CSV.get(rev["r1_decision"], "")
            d2 = DECISION_TO_CSV.get(rev["r2_decision"], "")
            fd = DECISION_TO_CSV.get(rev["final_decision"], "")

            if d1 or d2 or fd:
                row["human1_decision"] = d1
                row["human2_decision"] = d2
                row["adjudicated_final_decision"] = fd
                row["exclude_code"] = rev["final_code"] or rev["r1_code"] or ""
                row["decision_rationale"] = rev.get("r1_memo", "") or rev.get("r2_memo", "")
                row["adjudicator_id"] = "PI" if fd else ""
                updated += 1

    if not dry_run and updated > 0:
        with open(csv_path, "w", encoding="utf-8", newline="") as f:
            writer = csv.DictWriter(f, fieldnames=fieldnames)
            writer.writeheader()
            writer.writerows(rows)

    return updated


def export_irr_report(all_stats, output_dir):
    """IRR 보고서 생성"""
    os.makedirs(output_dir, exist_ok=True)
    report_path = os.path.join(output_dir, "irr_report.md")

    lines = [
        "# Inter-Rater Reliability (IRR) Report",
        "",
        f"Generated from: `human_review_sheet_v5.xlsx`",
        "",
        "## Summary",
        "",
        "| Sheet | Paired (n) | Agreement (%) | Cohen's κ | Interpretation |",
        "|-------|-----------|---------------|-----------|----------------|",
    ]

    for sheet_name, stats in all_stats.items():
        if stats is None:
            lines.append(f"| {sheet_name} | 0 | - | - | No paired data |")
            continue
        k = stats["kappa"]
        interp = (
            "Almost perfect" if k >= 0.81 else
            "Substantial" if k >= 0.61 else
            "Moderate" if k >= 0.41 else
            "Fair" if k >= 0.21 else
            "Slight" if k >= 0.01 else
            "Poor"
        )
        lines.append(
            f"| {sheet_name} | {stats['n_paired']} | {stats['agreement_pct']}% | {k} | {interp} |"
        )

    lines += [
        "",
        "## Interpretation Guide (Landis & Koch, 1977)",
        "",
        "| κ Range | Interpretation |",
        "|---------|----------------|",
        "| 0.81-1.00 | Almost perfect |",
        "| 0.61-0.80 | Substantial |",
        "| 0.41-0.60 | Moderate |",
        "| 0.21-0.40 | Fair |",
        "| 0.01-0.20 | Slight |",
        "| < 0.00 | Poor |",
    ]

    with open(report_path, "w") as f:
        f.write("\n".join(lines))

    print(f"  IRR report saved: {report_path}")


def main():
    parser = argparse.ArgumentParser(description="v5 Excel → CSV 변환")
    parser.add_argument("--dry-run", action="store_true", help="변환 미리보기 (파일 수정 안 함)")
    parser.add_argument("--irr-only", action="store_true", help="IRR 통계만 산출")
    args = parser.parse_args()

    if not os.path.exists(EXCEL_PATH):
        print(f"Error: Excel not found at {EXCEL_PATH}")
        sys.exit(1)

    print(f"Loading: {EXCEL_PATH}")
    wb = load_workbook(EXCEL_PATH, read_only=True, data_only=True)

    # 모든 시트에서 데이터 읽기
    all_records = []
    all_irr_stats = {}

    for sheet_name in wb.sheetnames:
        if sheet_name in ("AI_참조", "코딩가이드"):
            continue

        ws = wb[sheet_name]
        records = read_v5_sheet(ws)
        all_records.extend(records)

        # IRR 산출
        kappa, n_paired, stats = compute_irr(records)
        if n_paired > 0:
            all_irr_stats[sheet_name] = stats
            print(f"  {sheet_name}: {len(records)} records, {n_paired} paired → κ={stats['kappa']}")
        else:
            all_irr_stats[sheet_name] = None
            print(f"  {sheet_name}: {len(records)} records, no paired data for IRR")

    # IRR 전체 합산
    kappa_all, n_all, stats_all = compute_irr(all_records)
    if n_all > 0:
        all_irr_stats["TOTAL"] = stats_all
        print(f"\n  TOTAL: {len(all_records)} records, {n_all} paired → κ={stats_all['kappa']}")

    # IRR 보고서 생성
    export_irr_report(all_irr_stats, OUTPUT_DIR)

    if args.irr_only:
        print("\n[IRR only mode] Done.")
        return

    # CSV 업데이트
    print(f"\n{'[DRY RUN] ' if args.dry_run else ''}Updating CSVs...")

    n1 = update_csv(DUAL_CSV, all_records, dry_run=args.dry_run)
    print(f"  screening_ai_dual.csv: {n1} records {'would be ' if args.dry_run else ''}updated")

    n2 = update_csv(QUEUE_CSV, all_records, dry_run=args.dry_run)
    print(f"  human_review_queue.csv: {n2} records {'would be ' if args.dry_run else ''}updated")

    # 요약 CSV 내보내기
    summary_path = os.path.join(OUTPUT_DIR, "screening_human_decisions.csv")
    if not args.dry_run:
        os.makedirs(OUTPUT_DIR, exist_ok=True)
        with open(summary_path, "w", encoding="utf-8", newline="") as f:
            writer = csv.DictWriter(f, fieldnames=[
                "record_id", "r1_decision", "r1_code", "r1_memo",
                "r2_decision", "r2_code", "r2_memo",
                "final_decision", "final_code"
            ])
            writer.writeheader()
            for r in all_records:
                if r["r1_decision"] or r["r2_decision"] or r["final_decision"]:
                    writer.writerow({
                        "record_id": r["id"],
                        "r1_decision": DECISION_TO_CSV.get(r["r1_decision"], ""),
                        "r1_code": r["r1_code"],
                        "r1_memo": r["r1_memo"],
                        "r2_decision": DECISION_TO_CSV.get(r["r2_decision"], ""),
                        "r2_code": r["r2_code"],
                        "r2_memo": r["r2_memo"],
                        "final_decision": DECISION_TO_CSV.get(r["final_decision"], ""),
                        "final_code": r["final_code"],
                    })
        print(f"  Human decisions summary: {summary_path}")

    # 진행률 표시
    filled = sum(1 for r in all_records if r["r1_decision"] or r["r2_decision"])
    total = len(all_records)
    final_done = sum(1 for r in all_records if r["final_decision"])
    print(f"\n── Progress ──")
    print(f"  Reviewed (at least 1 reviewer): {filled}/{total} ({filled/total*100:.1f}%)")
    print(f"  Final decision entered: {final_done}/{total} ({final_done/total*100:.1f}%)")

    wb.close()
    print("\nDone.")


if __name__ == "__main__":
    main()
