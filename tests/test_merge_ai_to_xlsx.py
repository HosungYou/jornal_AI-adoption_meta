"""
Tests for scripts/screening/merge_ai_to_xlsx.py

Tests exercise the merge logic by constructing minimal XLSX workbooks
and CSV files, then calling the core merge logic directly.
"""

import sys
import pytest
import pandas as pd
import openpyxl
from pathlib import Path

sys.path.insert(0, str(Path(__file__).parent.parent / "scripts" / "screening"))

from merge_ai_to_xlsx import AI_COLUMNS, main as merge_main

# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

AI_COLS_SUBSET = [
    "screen_decision_codex",
    "screen_confidence_codex",
    "rationale_codex",
    "screen_decision_gemini",
    "screen_confidence_gemini",
    "rationale_gemini",
    "screen_consensus",
    "oauth_auth_method_codex",
    "oauth_auth_method_gemini",
    "exclude_code_codex",
    "exclude_code_gemini",
]

HUMAN_COLS = ["human1_decision", "human2_decision", "adjudicated_final_decision"]


def _make_xlsx(tmp_path, n=5, prefill_ai=False, prefill_human=True):
    """Create a minimal SCREENING workbook for testing."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "SCREENING"

    all_cols = ["record_id", "title"] + AI_COLS_SUBSET + HUMAN_COLS
    ws.append(all_cols)

    for i in range(1, n + 1):
        row = [i, f"Study {i}"]
        for col in AI_COLS_SUBSET:
            row.append("existing_ai_value" if prefill_ai else None)
        for col in HUMAN_COLS:
            row.append("human_value" if prefill_human else None)
        ws.append(row)

    path = tmp_path / "workbook.xlsx"
    wb.save(path)
    return path


def _make_ai_csv(tmp_path, n=5, record_ids=None):
    """Create a minimal AI screening CSV."""
    if record_ids is None:
        record_ids = list(range(1, n + 1))
    records = []
    for i in record_ids:
        records.append({
            "record_id": i,
            "title": f"Study {i}",
            "screen_decision_codex": "include",
            "screen_confidence_codex": 0.90,
            "exclude_code_codex": "NA",
            "rationale_codex": f"Codex rationale {i}",
            "screen_decision_gemini": "include",
            "screen_confidence_gemini": 0.88,
            "exclude_code_gemini": "NA",
            "rationale_gemini": f"Gemini rationale {i}",
            "screen_consensus": "include",
            "oauth_auth_method_codex": "oauth",
            "oauth_auth_method_gemini": "oauth",
        })
    df = pd.DataFrame(records)
    path = tmp_path / "ai_results.csv"
    df.to_csv(path, index=False)
    return path


def _run_merge(csv_path, xlsx_path):
    """Invoke merge_main via sys.argv patching."""
    import sys
    old_argv = sys.argv
    try:
        sys.argv = [
            "merge_ai_to_xlsx.py",
            "--screening-csv", str(csv_path),
            "--xlsx", str(xlsx_path),
        ]
        merge_main()
    finally:
        sys.argv = old_argv


# ---------------------------------------------------------------------------
# Tests
# ---------------------------------------------------------------------------

def test_merge_updates_ai_columns(tmp_path):
    """AI columns should be populated in XLSX after merge."""
    xlsx_path = _make_xlsx(tmp_path, n=5, prefill_ai=False, prefill_human=False)
    csv_path = _make_ai_csv(tmp_path, n=5)

    _run_merge(csv_path, xlsx_path)

    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb["SCREENING"]
    header = {cell.value: idx + 1 for idx, cell in enumerate(ws[1])}

    col_idx = header.get("screen_decision_codex")
    assert col_idx is not None
    # All 5 data rows should now have a value
    for row_num in range(2, 7):
        val = ws.cell(row=row_num, column=col_idx).value
        assert val is not None and str(val).strip() != "", (
            f"Row {row_num}: screen_decision_codex not updated"
        )


def test_merge_preserves_human_columns(tmp_path):
    """Human coder columns must not be touched."""
    xlsx_path = _make_xlsx(tmp_path, n=5, prefill_ai=False, prefill_human=True)
    csv_path = _make_ai_csv(tmp_path, n=5)

    _run_merge(csv_path, xlsx_path)

    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb["SCREENING"]
    header = {cell.value: idx + 1 for idx, cell in enumerate(ws[1])}

    for hcol in HUMAN_COLS:
        col_idx = header.get(hcol)
        if col_idx is None:
            continue  # column not present in this minimal workbook - skip
        for row_num in range(2, 7):
            val = ws.cell(row=row_num, column=col_idx).value
            assert val == "human_value", (
                f"Row {row_num}, column {hcol}: human value was overwritten (got {val!r})"
            )


def test_merge_no_overwrite_existing(tmp_path):
    """Cells that already have AI values must not be overwritten."""
    xlsx_path = _make_xlsx(tmp_path, n=5, prefill_ai=True, prefill_human=False)
    csv_path = _make_ai_csv(tmp_path, n=5)

    _run_merge(csv_path, xlsx_path)

    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb["SCREENING"]
    header = {cell.value: idx + 1 for idx, cell in enumerate(ws[1])}

    col_idx = header.get("screen_decision_codex")
    assert col_idx is not None
    for row_num in range(2, 7):
        val = ws.cell(row=row_num, column=col_idx).value
        assert val == "existing_ai_value", (
            f"Row {row_num}: existing AI value was overwritten (got {val!r})"
        )


def test_merge_missing_record_id(tmp_path):
    """CSV record_ids not present in XLSX should be silently skipped."""
    xlsx_path = _make_xlsx(tmp_path, n=5, prefill_ai=False, prefill_human=False)
    # CSV has record_ids 6-10 which don't exist in the XLSX (which has 1-5)
    csv_path = _make_ai_csv(tmp_path, record_ids=[6, 7, 8, 9, 10])

    # Should complete without error
    _run_merge(csv_path, xlsx_path)

    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb["SCREENING"]
    header = {cell.value: idx + 1 for idx, cell in enumerate(ws[1])}

    col_idx = header.get("screen_decision_codex")
    assert col_idx is not None
    # All cells in XLSX should remain empty (nothing matched)
    for row_num in range(2, 7):
        val = ws.cell(row=row_num, column=col_idx).value
        assert val is None or str(val).strip() == "", (
            f"Row {row_num}: unexpected value written for unmatched record"
        )


def test_merge_empty_csv(tmp_path):
    """An empty AI results CSV should leave the XLSX unchanged."""
    xlsx_path = _make_xlsx(tmp_path, n=5, prefill_ai=False, prefill_human=False)
    # CSV with header only, no data rows
    df = pd.DataFrame(columns=["record_id"] + AI_COLS_SUBSET)
    csv_path = tmp_path / "empty_ai.csv"
    df.to_csv(csv_path, index=False)

    _run_merge(csv_path, xlsx_path)

    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb["SCREENING"]
    header = {cell.value: idx + 1 for idx, cell in enumerate(ws[1])}

    col_idx = header.get("screen_decision_codex")
    assert col_idx is not None
    for row_num in range(2, 7):
        val = ws.cell(row=row_num, column=col_idx).value
        assert val is None or str(val).strip() == "", (
            f"Row {row_num}: cell should be empty after empty-CSV merge"
        )
