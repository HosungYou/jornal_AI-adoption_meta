#!/usr/bin/env python3
"""Prepare Phase 2 returned human-coder workbooks for source adjudication."""

from __future__ import annotations

import csv
import hashlib
import shutil
from collections import Counter, defaultdict
from dataclasses import dataclass
from datetime import date
from pathlib import Path

from openpyxl import load_workbook


REPO = Path(__file__).resolve().parents[4]
TODAY = date(2026, 5, 25)
STAMP = "20260525"

PHASE2_BASE = REPO / "data/04_extraction/01_raw_human_coder_data_freeze/phase2"
RAW_BASE = PHASE2_BASE / "returned_raw"
FREEZE_BASE = PHASE2_BASE / "freeze_candidates"
ADJ_BASE = REPO / "data/04_extraction/03_source_document_adjudication/phase2"
DIFF_BASE = REPO / "data/04_extraction/02_pre_adjudication_disagreement/phase2"


@dataclass(frozen=True)
class WorkbookSpec:
    coder: str
    label: str
    source: Path
    raw_name: str
    freeze_name: str | None = None
    use_for_freeze: bool = True


WORKBOOKS = [
    WorkbookSpec(
        "R1",
        "received_user_candidate",
        Path("/Users/hosung/Downloads/AI_Adoption_MASEM_Coding_v3_R1_Phase0_1_2_20260425 (2).xlsx"),
        f"AI_Adoption_MASEM_Coding_v3_R1_Phase0_1_2_20260425_received_{STAMP}.xlsx",
        None,
        False,
    ),
    WorkbookSpec(
        "R1",
        "completed_rows2_118",
        Path("/Users/hosung/Downloads/AI_Adoption_MASEM_Coding_v3_R1_Phase0_1_2_completed_rows2_118_20260523.xlsx"),
        f"AI_Adoption_MASEM_Coding_v3_R1_Phase0_1_2_completed_rows2_118_20260523_raw_{STAMP}.xlsx",
        f"AI_Adoption_MASEM_Coding_v3_R1_Phase0_1_2_freeze_candidate_{STAMP}.xlsx",
        True,
    ),
    WorkbookSpec(
        "R2",
        "received",
        Path("/Users/hosung/Downloads/AI_Adoption_MASEM_Coding_v3_R2_Phase0_1_2_20260425_R2.xlsx"),
        f"AI_Adoption_MASEM_Coding_v3_R2_Phase0_1_2_20260425_received_{STAMP}.xlsx",
        f"AI_Adoption_MASEM_Coding_v3_R2_Phase0_1_2_freeze_candidate_{STAMP}.xlsx",
    ),
    WorkbookSpec(
        "R3",
        "latest_consensus",
        Path("/Users/hosung/Downloads/AI_Adoption_MASEM_Coding_v3_R3_Phase0_1_2_20260425 (1).xlsx"),
        f"AI_Adoption_MASEM_Coding_v3_R3_Phase0_1_2_20260425_latest_received_{STAMP}.xlsx",
        f"AI_Adoption_MASEM_Coding_v3_R3_Phase0_1_2_latest_consensus_freeze_candidate_{STAMP}.xlsx",
    ),
    WorkbookSpec(
        "R4",
        "received",
        Path("/Users/hosung/Downloads/AI_Adoption_MASEM_Coding_v3_R4_Phase0_1_2_20260425_v2.xlsx"),
        f"AI_Adoption_MASEM_Coding_v3_R4_Phase0_1_2_20260425_v2_received_{STAMP}.xlsx",
        f"AI_Adoption_MASEM_Coding_v3_R4_Phase0_1_2_freeze_candidate_{STAMP}.xlsx",
    ),
]


SOURCE_CHECKS = {
    "S014": {
        "decision": "include_candidate",
        "recommended_action": "adjudicate_not_exclude",
        "exclusion_code": "",
        "rationale": (
            "Reports PLS-SEM path analysis for adoption of AI-based data-analysis tools "
            "among academic researchers. R1 coded complete-sample indirect paths, while "
            "R4 has no coded values; adjudication is needed for indirect-effect handling, "
            "construct mapping, and higher-education population eligibility."
        ),
        "source_location": "Abstract; Figure 1; Results 4.3; Table 4",
        "source_excerpt": "Table 4 Path analysis",
        "affected_coders": "R1 coded values; R4 no coded values",
    },
    "S021": {
        "decision": "include_candidate",
        "recommended_action": "adjudicate_not_exclude",
        "exclusion_code": "",
        "rationale": (
            "Reports PLS-SEM paths for genAI acceptance among higher-education staff "
            "before/after a training intervention; source has potentially usable path "
            "coefficients, but the pre/post design and T1/T2 construct separation need adjudication."
        ),
        "source_location": "Abstract; Methods 3.4; Results 4.2-4.3; Figures 1-2; Supplementary Table S4",
        "source_excerpt": "pre- and post-course genAI acceptance",
        "affected_coders": "R2 excluded; R3 no coded values",
    },
    "S039": {
        "decision": "exclude_confirmed",
        "recommended_action": "exclude_study",
        "exclusion_code": "E-FT3",
        "rationale": (
            "Dental-patient acceptance of AI-powered diagnosis in a clinical dental setting; "
            "the focal use case is healthcare diagnosis rather than educational AI adoption."
        ),
        "source_location": "Abstract; Methods: study design and participants",
        "source_excerpt": "AI-powered diagnosis",
        "affected_coders": "R2 excluded; R3 no coded values",
    },
    "S092": {
        "decision": "include_candidate",
        "recommended_action": "adjudicate_not_exclude",
        "exclusion_code": "",
        "rationale": (
            "Reports SEM model fit and standardized path estimates for ChatGPT adoption "
            "among ESP/business-communication students; source appears to contain usable "
            "target path coefficients."
        ),
        "source_location": "Abstract; Results: Structural Equation Modeling; Table 3",
        "source_excerpt": "Structural equation modeling (SEM)",
        "affected_coders": "R2 excluded",
    },
    "S101": {
        "decision": "exclude_confirmed",
        "recommended_action": "exclude_study",
        "exclusion_code": "E-FT1",
        "rationale": (
            "Uses ANN predictive-accuracy models and factor-loading summaries rather than "
            "a usable inter-construct correlation matrix or standardized target path coefficients."
        ),
        "source_location": "Methods: Artificial Neural Networks model; Results Tables 5-6",
        "source_excerpt": "Artificial Neural Networks (ANN)",
        "affected_coders": "R2 excluded; R3 no coded values",
    },
    "S108": {
        "decision": "exclude_confirmed",
        "recommended_action": "exclude_study",
        "exclusion_code": "E-FT1",
        "rationale": (
            "Reports TAM/RIMMS group mean comparisons and t-tests for maker-education "
            "motivation/acceptance; no usable target construct-pair r or standardized path matrix."
        ),
        "source_location": "Methods 2.2; Results Table 2",
        "source_excerpt": "comparison was made between the perceptions",
        "affected_coders": "R1 excluded; R4 coded an effect-size row",
    },
    "S118": {
        "decision": "exclude_confirmed",
        "recommended_action": "exclude_study",
        "exclusion_code": "E-FT1",
        "rationale": (
            "Reports descriptive acceptance results and Spearman correlations between GenAI use "
            "frequency and acceptance dimensions, not an adoption-model inter-construct matrix or SEM path model."
        ),
        "source_location": "Results Tables 1-3",
        "source_excerpt": "GenAI acceptance is positively moderate correlated",
        "affected_coders": "R2 excluded; R3 no coded values",
    },
    "S056": {
        "decision": "include_candidate",
        "recommended_action": "adjudicate_not_exclude",
        "exclusion_code": "",
        "rationale": (
            "Reports TAM/PLS-SEM standardized path coefficients for ChatGPT acceptance "
            "among Chinese-as-a-foreign-language learners. R2 coded the TAM paths, while "
            "R3 has no coded values; adjudication is needed to recover the missing coder-side rows."
        ),
        "source_location": "Abstract; Methods 3.4; Results 4.1.2; Table 3",
        "source_excerpt": "Table 3. Path significance and coefficients.",
        "affected_coders": "R2 coded values; R3 no coded values",
    },
    "S121": {
        "decision": "include_candidate",
        "recommended_action": "adjudicate_not_exclude",
        "exclusion_code": "",
        "rationale": (
            "Reports UTAUT/PLS-SEM paths for generative AI use and intention among students "
            "and teachers; source appears to contain usable target path coefficients."
        ),
        "source_location": "Abstract; Methods; SEM/PLS-SEM results",
        "source_excerpt": "partial least squares SEM (PLS-SEM)",
        "affected_coders": "R3 no coded values",
    },
    "S132": {
        "decision": "exclude_confirmed",
        "recommended_action": "exclude_study",
        "exclusion_code": "E-FT1",
        "rationale": (
            "Mentorship perception study reports identification/evaluation regressions and correlations "
            "for helpfulness, caring, and likelihood ratings; no usable target AI-adoption construct-pair "
            "r or SEM path coefficients."
        ),
        "source_location": "Abstract; Research questions; Results",
        "source_excerpt": "helpfulness, caring, and likelihood",
        "affected_coders": "R1 excluded; R4 no coded values",
    },
    "S195": {
        "decision": "exclude_confirmed",
        "recommended_action": "exclude_study",
        "exclusion_code": "E-FT1",
        "rationale": (
            "Same DOI/PDF as S206. The source uses PLSR component loadings and an image-only "
            "item-level correlation matrix, not a usable construct-level inter-construct "
            "correlation matrix or standardized SEM/path table for the project model."
        ),
        "source_location": "Figure 3; Table 3; Table 4",
        "source_excerpt": "PLSR component loading factors",
        "affected_coders": "R1 coded item-level values; R4 no coded values; duplicate-source issue with S206",
    },
    "S202": {
        "decision": "include_candidate",
        "recommended_action": "adjudicate_not_exclude",
        "exclusion_code": "",
        "rationale": (
            "Reports SEM path coefficients and a Fornell-Larcker-style construct correlation table "
            "for AI-driven LMS automation and student readiness; inclusion depends on construct "
            "mapping and focal-technology adjudication."
        ),
        "source_location": "Research model; Table 4; Table 5",
        "source_excerpt": "structural equation modelling (SEM)",
        "affected_coders": "R2 excluded; R3 no coded values",
    },
    "S206": {
        "decision": "exclude_confirmed",
        "recommended_action": "exclude_study",
        "exclusion_code": "E-FT1",
        "rationale": (
            "Same DOI/PDF as S195. Uses PLSR component loadings and an image-only item-level "
            "correlation matrix for undergraduate generative-AI adoption; it is not a usable "
            "construct-level inter-construct correlation matrix or standardized SEM/path table."
        ),
        "source_location": "Figure 3; Table 3; Table 4",
        "source_excerpt": "PLSR component loading factors",
        "affected_coders": "R2 excluded; R3 no coded values; duplicate-source issue with S195",
    },
    "S224": {
        "decision": "exclude_confirmed",
        "recommended_action": "exclude_study",
        "exclusion_code": "E-FT3",
        "rationale": (
            "Full text focuses on virtual learning/Google Classroom adoption during COVID-19. "
            "AI and machine-learning features are discussed only as possible enhancements, so "
            "the focal technology is not educational AI adoption."
        ),
        "source_location": "Abstract; Introduction; Conclusion",
        "source_excerpt": "Adoption, use and enhancement of virtual learning",
        "affected_coders": "R1 excluded; R4 coded virtual-learning UTAUT paths",
    },
}


EXCLUSION_ENTRIES = {
    sid: {
        "study_id": sid,
        "first_author": {
            "S039": "Naik",
            "S101": "Darmono",
            "S108": "Quintana-Ordorika",
            "S118": "Nevarez Montes",
            "S132": "Lee",
            "S195": "Patterson",
            "S206": "Patterson",
            "S224": "Zhou",
        }.get(sid, ""),
        "year": {
            "S039": 2025,
            "S101": 2025,
            "S108": 2025,
            "S118": 2025,
            "S132": 2025,
            "S195": 2024,
            "S206": 2024,
            "S224": 2022,
        }.get(sid, ""),
        "title": {
            "S039": "Patients' Acceptance and Intentions on Using Artificial Intelligence in Dental Diagnosis",
            "S101": "Determining Factors Influencing Indonesian Higher Education Students' Intention to Adopt Artificial Intelligence Tools for Self-Directed Learning Management",
            "S108": "The impact of artificial intelligence on maker education: Motivation and technology acceptance in teacher training",
            "S118": "Faculty acceptance and use of generative artificial intelligence in their practice",
            "S132": "ChatGPT or Human Mentors? Student Perceptions of Technology Acceptance and Use and the Future of Mentorship in Higher Education",
            "S195": "Examining generative artificial intelligence adoption in academia: a UTAUT perspective",
            "S206": "Examining generative artificial intelligence adoption in academia: a UTAUT perspective",
            "S224": "Adoption, use and enhancement of virtual learning during COVID-19",
        }.get(sid, ""),
        "exclusion_stage": "data_extraction" if sid != "S039" else "full_text",
        "exclusion_code": SOURCE_CHECKS[sid]["exclusion_code"],
        "detailed_rationale": SOURCE_CHECKS[sid]["rationale"],
        "flag": "review",
        "notes": f"Source check {TODAY.isoformat()}.",
    }
    for sid in ["S039", "S101", "S108", "S118", "S132", "S195", "S206", "S224"]
}


TRUE_EXCLUDES_BY_CODER = {
    "R1": ["S108", "S132", "S195", "S224"],
    "R2": ["S039", "S101", "S118", "S206"],
    "R3": ["S039", "S101", "S118", "S206"],
    "R4": ["S108", "S132", "S195", "S224"],
}

REVIEW_BY_CODER = {
    "R2": ["S021", "S092", "S202"],
    "R3": ["S021", "S056", "S121", "S202"],
    "R4": ["S014"],
}


def sha256(path: Path) -> str:
    h = hashlib.sha256()
    with path.open("rb") as f:
        for chunk in iter(lambda: f.read(1024 * 1024), b""):
            h.update(chunk)
    return h.hexdigest()


def ensure_dirs() -> None:
    for base in [RAW_BASE, FREEZE_BASE, ADJ_BASE, DIFF_BASE / "derived"]:
        base.mkdir(parents=True, exist_ok=True)
    for coder in ["R1", "R2", "R3", "R4"]:
        (RAW_BASE / coder).mkdir(parents=True, exist_ok=True)
        (FREEZE_BASE / coder).mkdir(parents=True, exist_ok=True)


def nonempty_correlation_studies(wb) -> set[str]:
    ws = wb["CORRELATIONS"]
    out = set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        sid = row[0]
        if not sid:
            continue
        if row[3] is not None or row[5] is not None:
            out.add(str(sid))
    return out


def phase2_assignment_rows(wb) -> dict[str, int]:
    ws = wb["ASSIGNMENT"]
    rows = {}
    for r in range(2, ws.max_row + 1):
        sid = ws.cell(r, 1).value
        phase = ws.cell(r, 5).value
        if sid and phase and "Phase 2" in str(phase):
            rows[str(sid)] = r
    return rows


def update_assignment_status(wb, coder: str) -> None:
    ws = wb["ASSIGNMENT"]
    phase2_rows = phase2_assignment_rows(wb)
    nonempty = nonempty_correlation_studies(wb)
    true_excludes = set(TRUE_EXCLUDES_BY_CODER.get(coder, []))
    review = set(REVIEW_BY_CODER.get(coder, []))

    for sid, row_idx in phase2_rows.items():
        current = ws.cell(row_idx, 7).value
        if sid in true_excludes:
            ws.cell(row_idx, 7).value = "excluded"
            ws.cell(row_idx, 8).value = "review"
            ws.cell(row_idx, 9).value = f"Source check {TODAY.isoformat()}: excluded pending adjudication log."
        elif sid in review:
            ws.cell(row_idx, 7).value = "review_source"
            ws.cell(row_idx, 8).value = "review"
            ws.cell(row_idx, 9).value = (
                f"Source check {TODAY.isoformat()}: coder return conflicts with source evidence; adjudicate."
            )
        elif (not current or str(current).strip() == "") and sid in nonempty:
            ws.cell(row_idx, 7).value = "done"


def fill_exclusion_log(wb, study_ids: list[str]) -> None:
    ws = wb["EXCLUSION_LOG"]
    headers = [ws.cell(1, c).value for c in range(1, 10)]
    expected = [
        "study_id",
        "first_author",
        "year",
        "title",
        "exclusion_stage",
        "exclusion_code",
        "detailed_rationale",
        "flag",
        "notes",
    ]
    if headers[:9] != expected:
        raise ValueError(f"Unexpected EXCLUSION_LOG headers: {headers[:9]}")

    existing = {}
    first_empty = None
    for r in range(2, max(ws.max_row, 2) + 1):
        sid = ws.cell(r, 1).value
        if sid:
            existing[str(sid)] = r
        elif first_empty is None:
            first_empty = r

    for sid in study_ids:
        entry = EXCLUSION_ENTRIES[sid]
        row_idx = existing.get(sid)
        if row_idx is None:
            row_idx = first_empty or ws.max_row + 1
            first_empty = None
        values = [
            entry["study_id"],
            entry["first_author"],
            entry["year"],
            entry["title"],
            entry["exclusion_stage"],
            entry["exclusion_code"],
            entry["detailed_rationale"],
            entry["flag"],
            entry["notes"],
        ]
        for col, value in enumerate(values, start=1):
            ws.cell(row_idx, col).value = value
        existing[sid] = row_idx


def promote_r3_latest_sheet(wb) -> list[str]:
    repairs = []
    if "(합의) CORRELATIONS" in wb.sheetnames:
        if "CORRELATIONS" in wb.sheetnames:
            wb["CORRELATIONS"].title = "CORRELATIONS_original_return"
            repairs.append("renamed original CORRELATIONS to CORRELATIONS_original_return")
        wb["(합의) CORRELATIONS"].title = "CORRELATIONS"
        repairs.append("promoted latest consensus sheet to CORRELATIONS")
    ws = wb["CORRELATIONS"]
    if ws.cell(1, 6).value == "original_beta" and ws.cell(1, 7).value == "original_beta":
        ws.cell(1, 7).value = "p_value"
        repairs.append("restored duplicate original_beta header in column G to p_value")
    return repairs


def workbook_summary(path: Path, coder: str) -> dict[str, object]:
    wb = load_workbook(path, read_only=True, data_only=True)
    phase2_rows = {}
    status_counts = Counter()
    for row in wb["ASSIGNMENT"].iter_rows(min_row=2, values_only=True):
        sid, phase, status = row[0], row[4], row[6]
        if sid and phase and "Phase 2" in str(phase):
            phase2_rows[str(sid)] = status
            status_counts[str(status).strip() if status not in (None, "") else "[blank]"] += 1

    corr_sheet = "CORRELATIONS"
    corr_rows = 0
    nonempty_by_study = Counter()
    if corr_sheet in wb.sheetnames:
        for row in wb[corr_sheet].iter_rows(min_row=2, values_only=True):
            sid = row[0]
            if not sid:
                continue
            corr_rows += 1
            if row[3] is not None or row[5] is not None:
                nonempty_by_study[str(sid)] += 1

    exclusion_ids = []
    if "EXCLUSION_LOG" in wb.sheetnames:
        for row in wb["EXCLUSION_LOG"].iter_rows(min_row=2, values_only=True):
            if row[0]:
                exclusion_ids.append(str(row[0]))

    wb.close()
    return {
        "coder": coder,
        "phase2_assigned": len(phase2_rows),
        "phase2_status_counts": dict(status_counts),
        "correlation_rows": corr_rows,
        "phase2_nonempty_studies": sum(1 for sid in phase2_rows if sid in nonempty_by_study),
        "phase2_nonempty_rows": sum(nonempty_by_study[sid] for sid in phase2_rows),
        "phase2_zero_studies": sorted(sid for sid in phase2_rows if sid not in nonempty_by_study),
        "exclusion_ids": sorted(exclusion_ids),
    }


def collect_pairwise_differences(freeze_paths: dict[str, Path]) -> list[dict[str, object]]:
    def load_rows(path: Path) -> dict[tuple[str, str, str], dict[str, object]]:
        wb = load_workbook(path, read_only=True, data_only=True)
        assigned = set(phase2_assignment_rows(wb).keys())
        ws = wb["CORRELATIONS"]
        out = {}
        for row in ws.iter_rows(min_row=2, values_only=True):
            sid, c1, c2 = row[0], row[1], row[2]
            if sid not in assigned or not c1 or not c2:
                continue
            if row[3] is None and row[5] is None:
                continue
            key = (str(sid), str(c1), str(c2))
            out[key] = {
                "value": row[3] if row[3] is not None else row[5],
                "r_value": row[3],
                "r_source": row[4],
                "original_beta": row[5],
                "p_value": row[6],
                "significance": row[7],
                "source_location": row[10],
                "flag": row[11],
                "notes": row[12],
            }
        wb.close()
        return out

    rows = []
    pairs = [("Pair C", "R1", "R4"), ("Pair D", "R2", "R3")]
    for pair_name, a, b in pairs:
        left = load_rows(freeze_paths[a])
        right = load_rows(freeze_paths[b])
        for key in sorted(set(left) | set(right)):
            l = left.get(key)
            r = right.get(key)
            sid, c1, c2 = key
            if l is None or r is None:
                mismatch = "one_coder_only"
                diff = ""
            else:
                lv, rv = l["value"], r["value"]
                try:
                    diff = abs(float(lv) - float(rv))
                except Exception:
                    diff = ""
                mismatch = "numeric_or_source_diff" if lv != rv or l["r_source"] != r["r_source"] else "same"
            if mismatch == "same":
                continue
            rows.append(
                {
                    "pair": pair_name,
                    "study_id": sid,
                    "construct_1": c1,
                    "construct_2": c2,
                    "coder_a": a,
                    "coder_a_value": "" if l is None else l["value"],
                    "coder_a_source": "" if l is None else l["r_source"],
                    "coder_a_notes": "" if l is None else l["notes"],
                    "coder_b": b,
                    "coder_b_value": "" if r is None else r["value"],
                    "coder_b_source": "" if r is None else r["r_source"],
                    "coder_b_notes": "" if r is None else r["notes"],
                    "abs_difference": diff,
                    "mismatch_type": mismatch,
                }
            )
    return rows


def write_source_check_report() -> None:
    ADJ_BASE.mkdir(parents=True, exist_ok=True)
    csv_path = ADJ_BASE / f"phase2_source_check_candidates_{STAMP}.csv"
    with csv_path.open("w", newline="", encoding="utf-8") as f:
        fieldnames = [
            "study_id",
            "decision",
            "recommended_action",
            "exclusion_code",
            "rationale",
            "source_location",
            "source_excerpt",
            "affected_coders",
        ]
        writer = csv.DictWriter(f, fieldnames=fieldnames, lineterminator="\n")
        writer.writeheader()
        for sid in sorted(SOURCE_CHECKS):
            row = {"study_id": sid, **SOURCE_CHECKS[sid]}
            writer.writerow(row)

    md_path = ADJ_BASE / f"phase2_exclusion_source_check_{STAMP}.md"
    lines = [
        "# Phase 2 Exclusion Source Check",
        "",
        f"Date: {TODAY.isoformat()}",
        "",
        "This source check distinguishes confirmed exclusions from coder-return statuses that need adjudication before the source-anchored adjudicated human reference standard is frozen.",
        "",
        "## Confirmed Exclusions",
        "",
    ]
    for sid, item in SOURCE_CHECKS.items():
        if item["decision"] == "exclude_confirmed":
            lines.extend(
                [
                    f"### {sid}",
                    "",
                    f"- Code: `{item['exclusion_code']}`",
                    f"- Rationale: {item['rationale']}",
                    f"- Source location: {item['source_location']}",
                    f"- Short source excerpt: \"{item['source_excerpt']}\"",
                    f"- Affected coder returns: {item['affected_coders']}",
                    "",
                ]
            )
    lines.extend(["## Needs Adjudication Before Exclusion", ""])
    for sid, item in SOURCE_CHECKS.items():
        if item["decision"] != "exclude_confirmed":
            lines.extend(
                [
                    f"### {sid}",
                    "",
                    f"- Source-check status: `{item['decision']}`",
                    f"- Recommended action: `{item['recommended_action']}`",
                    f"- Rationale: {item['rationale']}",
                    f"- Source location: {item['source_location']}",
                    f"- Short source excerpt: \"{item['source_excerpt']}\"",
                    f"- Affected coder returns: {item['affected_coders']}",
                    "",
                ]
            )
    md_path.write_text("\n".join(lines), encoding="utf-8")


def write_manifest(rows: list[dict[str, object]], summaries: list[dict[str, object]]) -> None:
    manifest_csv = PHASE2_BASE / f"phase2_return_manifest_{STAMP}.csv"
    with manifest_csv.open("w", newline="", encoding="utf-8") as f:
        fieldnames = [
            "coder",
            "label",
            "source_path",
            "raw_path",
            "freeze_path",
            "source_sha256",
            "raw_sha256",
            "freeze_sha256",
            "repairs",
        ]
        writer = csv.DictWriter(f, fieldnames=fieldnames, lineterminator="\n")
        writer.writeheader()
        for row in rows:
            writer.writerow(row)

    md = PHASE2_BASE / f"RETURN_MANIFEST_{STAMP}.md"
    lines = [
        "# Phase 2 Return Manifest",
        "",
        f"Date: {TODAY.isoformat()}",
        "",
        "Raw returned workbooks are preserved separately from freeze candidates. Freeze candidates contain only structural repairs and source-check status updates needed for adjudication workflow; raw coder values are not overwritten.",
        "",
        "## Freeze Candidate Summary",
        "",
        "| Coder | Phase 2 assigned | Status counts | Nonempty Phase 2 studies | Nonempty rows | Zero-value Phase 2 studies |",
        "|---|---:|---|---:|---:|---|",
    ]
    for s in summaries:
        lines.append(
            "| {coder} | {phase2_assigned} | {counts} | {nonempty_studies} | {nonempty_rows} | {zero} |".format(
                coder=s["coder"],
                phase2_assigned=s["phase2_assigned"],
                counts=s["phase2_status_counts"],
                nonempty_studies=s["phase2_nonempty_studies"],
                nonempty_rows=s["phase2_nonempty_rows"],
                zero=", ".join(s["phase2_zero_studies"]) or "None",
            )
        )
    lines.extend(["", "## Repairs Applied", ""])
    for row in rows:
        if row["freeze_path"]:
            lines.append(f"- {row['coder']} `{Path(str(row['freeze_path'])).name}`: {row['repairs'] or 'none'}")
    lines.extend(
        [
            "",
            "## Source-Check Boundary",
            "",
            "Confirmed exclusions and review-required candidates are documented in `data/04_extraction/03_source_document_adjudication/phase2/phase2_exclusion_source_check_20260525.md`.",
        ]
    )
    md.write_text("\n".join(lines), encoding="utf-8")


def write_pairwise_outputs(freeze_paths: dict[str, Path]) -> None:
    rows = collect_pairwise_differences(freeze_paths)
    out = DIFF_BASE / "derived" / f"phase2_pairwise_disagreement_long_{STAMP}.csv"
    out.parent.mkdir(parents=True, exist_ok=True)
    with out.open("w", newline="", encoding="utf-8") as f:
        fieldnames = [
            "pair",
            "study_id",
            "construct_1",
            "construct_2",
            "coder_a",
            "coder_a_value",
            "coder_a_source",
            "coder_a_notes",
            "coder_b",
            "coder_b_value",
            "coder_b_source",
            "coder_b_notes",
            "abs_difference",
            "mismatch_type",
        ]
        writer = csv.DictWriter(f, fieldnames=fieldnames, lineterminator="\n")
        writer.writeheader()
        writer.writerows(rows)

    summary = Counter((row["pair"], row["mismatch_type"]) for row in rows)
    summary_path = DIFF_BASE / "derived" / f"phase2_pairwise_disagreement_summary_{STAMP}.csv"
    with summary_path.open("w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=["pair", "mismatch_type", "n"], lineterminator="\n")
        writer.writeheader()
        for (pair, mismatch_type), n in sorted(summary.items()):
            writer.writerow({"pair": pair, "mismatch_type": mismatch_type, "n": n})

    write_study_review_queue(freeze_paths, rows)

    readme = DIFF_BASE / "README.md"
    readme.write_text(
        "\n".join(
            [
                "# Phase 2 Pre-Adjudication Disagreement",
                "",
                "Derived Phase 2 pairwise disagreement files are generated from freeze-candidate workbooks, not from overwritten raw returns.",
                "",
                f"- Long differences: `derived/phase2_pairwise_disagreement_long_{STAMP}.csv`",
                f"- Summary counts: `derived/phase2_pairwise_disagreement_summary_{STAMP}.csv`",
                f"- Study-level meeting queue: `derived/phase2_study_review_queue_{STAMP}.csv`",
                "",
                "These files are a pre-adjudication queue. They do not represent the source-anchored adjudicated human reference standard.",
            ]
        ),
        encoding="utf-8",
    )


def write_study_review_queue(freeze_paths: dict[str, Path], rows: list[dict[str, object]]) -> None:
    pairs = [("Pair C", "R1", "R4"), ("Pair D", "R2", "R3")]
    assignments: dict[tuple[str, str], dict[str, object]] = {}

    for pair_name, coder_a, coder_b in pairs:
        for coder in [coder_a, coder_b]:
            wb = load_workbook(freeze_paths[coder], read_only=True, data_only=True)
            for row in wb["ASSIGNMENT"].iter_rows(min_row=2, values_only=True):
                sid, title, year, doi, phase, pdf, status, flag, notes = row[:9]
                if not sid or not phase or "Phase 2" not in str(phase):
                    continue
                key = (pair_name, str(sid))
                item = assignments.setdefault(
                    key,
                    {
                        "pair": pair_name,
                        "study_id": str(sid),
                        "title": title or "",
                        "year": year or "",
                        "doi": doi or "",
                        "pdf": pdf or "",
                        "coder_a": coder_a,
                        "coder_a_status": "",
                        "coder_a_notes": "",
                        "coder_b": coder_b,
                        "coder_b_status": "",
                        "coder_b_notes": "",
                    },
                )
                suffix = "a" if coder == coder_a else "b"
                item[f"coder_{suffix}_status"] = status or ""
                item[f"coder_{suffix}_notes"] = notes or ""
            wb.close()

    counts: dict[tuple[str, str], Counter] = defaultdict(Counter)
    for row in rows:
        key = (str(row["pair"]), str(row["study_id"]))
        counts[key]["n_difference_rows"] += 1
        counts[key][f"n_{row['mismatch_type']}"] += 1

    queue_rows = []
    for key, item in assignments.items():
        counter = counts.get(key, Counter())
        statuses = {str(item["coder_a_status"] or ""), str(item["coder_b_status"] or "")}
        include = counter["n_difference_rows"] > 0 or statuses != {"done"}
        if not include:
            continue
        note_bits = []
        if "review_source" in statuses:
            note_bits.append("source review")
        if "excluded" in statuses:
            note_bits.append("exclusion/source-check")
        if counter["n_one_coder_only"]:
            note_bits.append("one coder only values")
        if counter["n_numeric_or_source_diff"]:
            note_bits.append("numeric/source differences")
        queue_rows.append(
            {
                **item,
                "n_difference_rows": counter["n_difference_rows"],
                "n_one_coder_only": counter["n_one_coder_only"],
                "n_numeric_or_source_diff": counter["n_numeric_or_source_diff"],
                "review_note": "; ".join(note_bits),
            }
        )

    queue_rows.sort(key=lambda r: (r["pair"], r["study_id"]))
    queue_path = DIFF_BASE / "derived" / f"phase2_study_review_queue_{STAMP}.csv"
    with queue_path.open("w", newline="", encoding="utf-8") as f:
        fieldnames = [
            "pair",
            "study_id",
            "title",
            "year",
            "doi",
            "pdf",
            "coder_a",
            "coder_a_status",
            "coder_a_notes",
            "coder_b",
            "coder_b_status",
            "coder_b_notes",
            "n_difference_rows",
            "n_one_coder_only",
            "n_numeric_or_source_diff",
            "review_note",
        ]
        writer = csv.DictWriter(f, fieldnames=fieldnames, lineterminator="\n")
        writer.writeheader()
        writer.writerows(queue_rows)


def main() -> None:
    ensure_dirs()
    manifest_rows: list[dict[str, object]] = []
    freeze_paths: dict[str, Path] = {}

    for spec in WORKBOOKS:
        if not spec.source.exists():
            raise FileNotFoundError(spec.source)
        raw_path = RAW_BASE / spec.coder / spec.raw_name
        shutil.copy2(spec.source, raw_path)
        freeze_path = ""
        freeze_hash = ""
        repairs: list[str] = []

        if spec.use_for_freeze and spec.freeze_name:
            freeze = FREEZE_BASE / spec.coder / spec.freeze_name
            shutil.copy2(spec.source, freeze)
            wb = load_workbook(freeze)
            if spec.coder == "R3":
                repairs.extend(promote_r3_latest_sheet(wb))
            fill_exclusion_log(wb, TRUE_EXCLUDES_BY_CODER.get(spec.coder, []))
            if TRUE_EXCLUDES_BY_CODER.get(spec.coder):
                repairs.append("filled source-checked EXCLUSION_LOG rows")
            update_assignment_status(wb, spec.coder)
            repairs.append("normalized Phase 2 ASSIGNMENT statuses for freeze candidate")
            wb.save(freeze)
            wb.close()
            freeze_path = str(freeze.relative_to(REPO))
            freeze_hash = sha256(freeze)
            freeze_paths[spec.coder] = freeze

        manifest_rows.append(
            {
                "coder": spec.coder,
                "label": spec.label,
                "source_path": str(spec.source),
                "raw_path": str(raw_path.relative_to(REPO)),
                "freeze_path": freeze_path,
                "source_sha256": sha256(spec.source),
                "raw_sha256": sha256(raw_path),
                "freeze_sha256": freeze_hash,
                "repairs": "; ".join(dict.fromkeys(repairs)),
            }
        )

    summaries = [workbook_summary(freeze_paths[coder], coder) for coder in ["R1", "R2", "R3", "R4"]]
    write_manifest(manifest_rows, summaries)
    write_source_check_report()
    write_pairwise_outputs(freeze_paths)

    print("Prepared Phase 2 returned raw workbooks, freeze candidates, source-check report, and pairwise outputs.")
    for s in summaries:
        print(s)


if __name__ == "__main__":
    main()
