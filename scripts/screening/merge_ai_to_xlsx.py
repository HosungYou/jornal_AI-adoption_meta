"""
merge_ai_to_xlsx.py

Merges AI screening results from a CSV into an existing Screening Workbook XLSX.
Updates only AI columns; does not touch human coder columns or overwrite existing values.
"""

import argparse
import sys
import pandas as pd
import openpyxl


AI_COLUMNS = [
    "screen_decision_codex",
    "screen_confidence_codex",
    "exclude_code_codex",
    "rationale_codex",
    "screen_decision_gemini",
    "screen_confidence_gemini",
    "exclude_code_gemini",
    "rationale_gemini",
    "screen_consensus",
    "oauth_auth_method_codex",
    "oauth_auth_method_gemini",
]


def parse_args():
    parser = argparse.ArgumentParser(
        description="Merge AI screening results CSV into existing Screening Workbook XLSX."
    )
    parser.add_argument(
        "--screening-csv",
        required=True,
        help="Path to AI screening CSV output (from ai_screening.py)",
    )
    parser.add_argument(
        "--xlsx",
        required=True,
        help="Path to existing Screening Workbook XLSX to update",
    )
    return parser.parse_args()


def main():
    args = parse_args()

    # Load AI screening CSV
    print(f"Loading AI screening CSV: {args.screening_csv}")
    try:
        df = pd.read_csv(args.screening_csv, dtype={"record_id": str})
    except FileNotFoundError:
        print(f"ERROR: CSV file not found: {args.screening_csv}", file=sys.stderr)
        sys.exit(1)

    if "record_id" not in df.columns:
        print("ERROR: CSV does not have a 'record_id' column.", file=sys.stderr)
        sys.exit(1)

    # Build a dict: record_id -> row (as dict of col->value)
    csv_records = {}
    for _, row in df.iterrows():
        rid = str(row["record_id"]).strip()
        csv_records[rid] = row

    print(f"  Loaded {len(csv_records)} records from CSV.")

    # Open existing XLSX
    print(f"Opening XLSX: {args.xlsx}")
    try:
        wb = openpyxl.load_workbook(args.xlsx)
    except FileNotFoundError:
        print(f"ERROR: XLSX file not found: {args.xlsx}", file=sys.stderr)
        sys.exit(1)

    # Find SCREENING sheet (case-insensitive)
    sheet_name = None
    for name in wb.sheetnames:
        if name.strip().upper() == "SCREENING":
            sheet_name = name
            break

    if sheet_name is None:
        print(
            f"ERROR: No 'SCREENING' sheet found. Available sheets: {wb.sheetnames}",
            file=sys.stderr,
        )
        sys.exit(1)

    ws = wb[sheet_name]
    print(f"  Using sheet: '{sheet_name}'")

    # Read header row (row 1) to build column name -> col index mapping
    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    col_index = {}  # col_name -> 1-based column number
    record_id_col = None

    for idx, cell_val in enumerate(header_row, start=1):
        if cell_val is None:
            continue
        name = str(cell_val).strip()
        col_index[name] = idx
        if name == "record_id":
            record_id_col = idx

    if record_id_col is None:
        print(
            "ERROR: 'record_id' column not found in SCREENING sheet header row.",
            file=sys.stderr,
        )
        sys.exit(1)

    # Check which AI columns are present in the XLSX
    missing_xlsx_cols = [c for c in AI_COLUMNS if c not in col_index]
    if missing_xlsx_cols:
        print(
            f"WARNING: The following AI columns are not in the XLSX header and will be skipped: {missing_xlsx_cols}"
        )

    present_ai_cols = [c for c in AI_COLUMNS if c in col_index]

    # Build mapping: record_id -> xlsx row number
    print("  Building record_id -> row mapping from XLSX...")
    xlsx_row_map = {}  # record_id str -> row number (1-based, excluding header)
    for row in ws.iter_rows(min_row=2):
        rid_cell = row[record_id_col - 1]
        if rid_cell.value is not None:
            rid = str(rid_cell.value).strip()
            xlsx_row_map[rid] = rid_cell.row

    print(f"  Found {len(xlsx_row_map)} records in XLSX SCREENING sheet.")

    # Merge
    matched = 0
    cells_updated = 0
    cells_skipped = 0
    unmatched_ids = []

    for rid, csv_row in csv_records.items():
        if rid not in xlsx_row_map:
            unmatched_ids.append(rid)
            continue

        matched += 1
        xlsx_row_num = xlsx_row_map[rid]

        for col_name in present_ai_cols:
            if col_name not in csv_row.index:
                continue

            csv_val = csv_row[col_name]
            # Treat NaN as empty
            if pd.isna(csv_val):
                csv_val = None

            col_num = col_index[col_name]
            cell = ws.cell(row=xlsx_row_num, column=col_num)

            existing = cell.value
            # Skip if cell already has a value (resume-safe)
            if existing is not None and str(existing).strip() != "":
                cells_skipped += 1
                continue

            if csv_val is not None:
                cell.value = csv_val
                cells_updated += 1

    # Save
    print(f"Saving XLSX: {args.xlsx}")
    wb.save(args.xlsx)

    # Summary
    print("\n--- Merge Summary ---")
    print(f"  Records in CSV       : {len(csv_records)}")
    print(f"  Records matched      : {matched}")
    print(f"  Records unmatched    : {len(unmatched_ids)}")
    print(f"  Cells updated        : {cells_updated}")
    print(f"  Cells skipped (had values): {cells_skipped}")

    if unmatched_ids:
        print(f"\n  WARNING: {len(unmatched_ids)} record_id(s) from CSV not found in XLSX:")
        for uid in unmatched_ids[:20]:
            print(f"    - {uid}")
        if len(unmatched_ids) > 20:
            print(f"    ... and {len(unmatched_ids) - 20} more.")

    print("\nDone.")


if __name__ == "__main__":
    main()
