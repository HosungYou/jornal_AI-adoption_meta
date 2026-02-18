#!/usr/bin/env python3
"""
Export included studies from a completed Screening Workbook into a new
Coding Workbook, reusing the create_masem_template.py infrastructure.

Usage:
    python export_included_to_coding.py \
        --screening-xlsx data/screening/AI_Adoption_Screening.xlsx \
        --output data/templates/AI_Adoption_MASEM_Coding_v1.xlsx
"""

import argparse
import sys
import os
import re

import openpyxl

# ---------------------------------------------------------------------------
# Path setup — import from data/templates/create_masem_template.py
# ---------------------------------------------------------------------------
_SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
_PROJECT_ROOT = os.path.abspath(os.path.join(_SCRIPT_DIR, "..", ".."))
_TEMPLATE_DIR = os.path.join(_PROJECT_ROOT, "data", "templates")
if _TEMPLATE_DIR not in sys.path:
    sys.path.insert(0, _TEMPLATE_DIR)

from create_masem_template import (  # noqa: E402
    ALL_SHEETS,
    SHEET_TAB_COLORS,
    STUDY_METADATA_COLS,
    apply_header_format,
    add_data_validations,
    create_data_sheet,
    create_codebook_sheet,
    build_codebook_rows,
)


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def extract_first_author(authors_raw: str) -> str:
    """
    Return the last name (or first token) of the first author.

    Handles common formats:
        "Smith, J.; Jones, A."   -> "Smith"
        "Smith J, Jones A"       -> "Smith"
        "John Smith and Jane Doe"-> "Smith"  (last word of first name-unit)
        ""                       -> ""
    """
    if not authors_raw or str(authors_raw).strip() in ("", "nan", "None"):
        return ""

    raw = str(authors_raw).strip()

    # Split on semicolon or " and " to get first author segment
    first_segment = re.split(r";| and ", raw, maxsplit=1)[0].strip()

    # Further split on comma to isolate the very first author unit.
    # "Smith, J."      -> first unit = "Smith"  (single token, no space -> last name)
    # "Smith J, Jones" -> first unit = "Smith J" (has space)
    # "Jones A"        -> no comma, first unit = "Jones A"
    first_unit = first_segment.split(",")[0].strip() if "," in first_segment else first_segment

    tokens = first_unit.split()
    if not tokens:
        return first_unit

    # If the last token looks like an initial (1-2 chars, possibly with a period),
    # the first token is the last name (e.g. "Smith J" or "Smith J.").
    # Otherwise for "John Smith" style, the last token is the last name.
    last_tok = tokens[-1].rstrip(".")
    if len(last_tok) <= 2 and len(tokens) > 1:
        # "LastName Initial" order -> first token is the last name
        return tokens[0]
    else:
        # "FirstName LastName" or bare last name -> last token is the last name
        return tokens[-1]


def read_included_studies(screening_xlsx: str) -> list[dict]:
    """
    Open the screening workbook, read the SCREENING sheet, and return
    a list of dicts for rows where adjudicated_final_decision == 'include'.
    """
    wb = openpyxl.load_workbook(screening_xlsx, read_only=True, data_only=True)

    # Try to find SCREENING sheet (case-insensitive)
    sheet_name = None
    for name in wb.sheetnames:
        if name.upper() == "SCREENING":
            sheet_name = name
            break
    if sheet_name is None:
        raise ValueError(
            f"No SCREENING sheet found in {screening_xlsx}. "
            f"Available sheets: {wb.sheetnames}"
        )

    ws = wb[sheet_name]

    # Read header row
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise ValueError("SCREENING sheet is empty.")

    headers = [str(h).strip() if h is not None else "" for h in rows[0]]

    def col(name: str) -> int | None:
        """Return 0-based column index for a header name, or None."""
        try:
            return headers.index(name)
        except ValueError:
            return None

    # Locate required columns
    decision_col = col("adjudicated_final_decision")
    authors_col = col("authors")
    year_col = col("year")
    title_col = col("title")
    doi_col = col("doi")
    source_col = col("source_database")

    if decision_col is None:
        raise ValueError(
            "Column 'adjudicated_final_decision' not found in SCREENING sheet. "
            f"Available columns: {headers}"
        )

    included = []
    study_counter = 1

    for row in rows[1:]:
        if not any(cell is not None for cell in row):
            continue  # skip blank rows

        decision = str(row[decision_col]).strip().lower() if row[decision_col] is not None else ""
        if decision != "include":
            continue

        authors_raw = str(row[authors_col]).strip() if authors_col is not None and row[authors_col] is not None else ""
        year_raw = row[year_col] if year_col is not None else None
        title_raw = str(row[title_col]).strip() if title_col is not None and row[title_col] is not None else ""
        doi_raw = str(row[doi_col]).strip() if doi_col is not None and row[doi_col] is not None else ""
        source_raw = str(row[source_col]).strip() if source_col is not None and row[source_col] is not None else ""

        included.append({
            "study_id": study_counter,
            "first_author": extract_first_author(authors_raw),
            "year": int(year_raw) if year_raw not in (None, "", "nan") else None,
            "title": title_raw,
            "doi": doi_raw,
            "search_source": source_raw,
        })
        study_counter += 1

    wb.close()
    return included


def prefill_study_metadata(ws, included_studies: list[dict]) -> None:
    """
    Write the bibliographic data for included studies into the
    STUDY_METADATA sheet starting at row 2.

    Maps study dict keys to STUDY_METADATA_COLS column positions.
    """
    # Build column-name -> 1-based column index map from the sheet's header row
    header_to_col = {}
    for col_idx, col_def in enumerate(STUDY_METADATA_COLS, start=1):
        col_name = col_def[0]
        header_to_col[col_name] = col_idx

    for row_offset, study in enumerate(included_studies):
        row_idx = row_offset + 2  # row 1 is header
        for field, value in study.items():
            col_idx = header_to_col.get(field)
            if col_idx is not None and value not in (None, "", "nan", "None"):
                ws.cell(row=row_idx, column=col_idx, value=value)


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    parser = argparse.ArgumentParser(
        description="Export included studies from a Screening Workbook into a new Coding Workbook."
    )
    parser.add_argument(
        "--screening-xlsx",
        required=True,
        help="Path to the completed Screening Workbook (1st XLSX).",
    )
    parser.add_argument(
        "--output",
        default=os.path.join(
            _PROJECT_ROOT, "data", "templates", "AI_Adoption_MASEM_Coding_v1.xlsx"
        ),
        help="Path for the output Coding Workbook (2nd XLSX). "
             "Default: data/templates/AI_Adoption_MASEM_Coding_v1.xlsx",
    )
    args = parser.parse_args()

    screening_path = os.path.abspath(args.screening_xlsx)
    output_path = os.path.abspath(args.output)

    if not os.path.isfile(screening_path):
        sys.exit(f"ERROR: Screening file not found: {screening_path}")

    # ------------------------------------------------------------------
    # 1. Read included studies from screening workbook
    # ------------------------------------------------------------------
    print(f"Reading screening workbook: {screening_path}")
    included_studies = read_included_studies(screening_path)
    n_included = len(included_studies)
    print(f"  Found {n_included} included studies.")

    if n_included == 0:
        print("WARNING: No studies with adjudicated_final_decision == 'include' found.")
        print("         The coding workbook will be created with an empty STUDY_METADATA sheet.")

    # ------------------------------------------------------------------
    # 2. Build the coding workbook using template infrastructure
    # ------------------------------------------------------------------
    # Set DATA_ROWS dynamically: at least the number of included studies,
    # with a minimum of 50 rows for usability.
    data_rows = max(n_included, 50)

    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # remove default blank sheet

    # CODEBOOK (first sheet)
    print("Creating CODEBOOK sheet...")
    create_codebook_sheet(wb)

    # Data sheets
    for sheet_name, col_defs in ALL_SHEETS.items():
        print(f"Creating {sheet_name} sheet...")
        create_data_sheet(wb, sheet_name, col_defs, data_rows=data_rows)

    # ------------------------------------------------------------------
    # 3. Pre-fill STUDY_METADATA with included studies' bibliographic data
    # ------------------------------------------------------------------
    if n_included > 0:
        print("Pre-filling STUDY_METADATA with included studies...")
        ws_metadata = wb["STUDY_METADATA"]
        prefill_study_metadata(ws_metadata, included_studies)

    # ------------------------------------------------------------------
    # 4. Save
    # ------------------------------------------------------------------
    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    wb.save(output_path)

    # ------------------------------------------------------------------
    # 5. Summary
    # ------------------------------------------------------------------
    print(f"\n--- Summary ---")
    print(f"  Included studies exported : {n_included}")
    print(f"  Data rows pre-allocated   : {data_rows} per sheet")
    print(f"  Output workbook           : {output_path}")
    print(f"  Sheets created            : {', '.join(ws.title for ws in wb.worksheets)}")


if __name__ == "__main__":
    main()
