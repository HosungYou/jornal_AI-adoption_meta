#!/usr/bin/env python3
"""
human_review_sheet_v4 → v5 변환 스크립트

변경 사항:
1. 중복 시트 제거: Sheet2(제외권고 72건) + Sheet3(HIGH+MEDIUM 356건) → Sheet1에 통합
2. AI 분석 컬럼(MASEM점수, 감지구인 등) → 별도 'AI_참조' 시트로 분리 (리뷰어 편향 방지)
3. 리뷰어 입력 컬럼 표준화: Include(O) / Exclude(X) / Uncertain(?) + 제외코드
4. 데이터 검증(validation) 추가
5. 기존 리뷰어 데이터 보존 + 표준화 변환
"""

import json
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# ── 설정 ──────────────────────────────────────────────
REPO_ROOT = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
PARSED_DATA = "/tmp/v4_parsed.json"
OUTPUT_PATH = os.path.join(REPO_ROOT, "data", "templates", "human_review_sheet_v5.xlsx")

# 제외코드 정의
EXCLUSION_CODES = {
    "E-FT1": "비실증 연구 (리뷰, 이론, 메타분석)",
    "E-FT2": "TAM/UTAUT 구인 없음 (채택/수용 측정 안 함)",
    "E-FT3": "AI가 초점 기술이 아님 (일반 ICT, 로봇공학 등)",
    "E-FT4": "교육 맥락 아님",
    "E-FT5": "상관/경로계수 추출 불가 (질적, 실험 비교만)",
    "E-FT6": "중복 데이터 (동일 표본의 다른 논문)",
    "E-FT7": "전문 접근 불가",
    "E-FT8": "영어 아님",
}

# 스타일 정의
HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
HEADER_FONT = Font(name="맑은 고딕", bold=True, color="FFFFFF", size=10)
HUMAN_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
HUMAN_FONT = Font(name="맑은 고딕", bold=True, color="BF8F00", size=10)
FINAL_FILL = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
FINAL_FONT = Font(name="맑은 고딕", bold=True, color="375623", size=10)
DATA_FONT = Font(name="맑은 고딕", size=9)
THIN_BORDER = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)

# v4 → v5 판단값 표준화 매핑
DECISION_MAP = {
    "x": "X",       # exclude
    "X": "X",       # exclude
    "o": "O",       # include
    "O": "O",       # include
    "?": "?",       # uncertain
    "": "",
}


def standardize_decision(val):
    """기존 리뷰어 판단값을 O/X/?로 표준화"""
    val = val.strip()
    if val in DECISION_MAP:
        return DECISION_MAP[val]
    # 긴 텍스트(메모 등)는 무시
    if len(val) > 5:
        return ""
    return val


def apply_header_style(ws, row, col, text, fill=HEADER_FILL, font=HEADER_FONT):
    cell = ws.cell(row=row, column=col, value=text)
    cell.fill = fill
    cell.font = font
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = THIN_BORDER
    return cell


def apply_data_style(cell):
    cell.font = DATA_FONT
    cell.border = THIN_BORDER
    cell.alignment = Alignment(vertical="top", wrap_text=True)
    return cell


def create_screening_sheet(wb, title, records, sheet_type="include"):
    """스크리닝 시트 생성"""
    ws = wb.create_sheet(title=title)

    # ── 헤더 정의 ──
    # 기본 정보 (파란색)
    base_headers = ["ID", "제목", "연도", "초록"]
    # AI 참조 (파란색, 읽기 전용)
    ai_headers = ["Codex", "Gemini", "AI합의"]
    if sheet_type == "conflict":
        ai_headers.append("Claude추천")
    # 리뷰어 입력 (노란색)
    human_headers = ["리뷰어1\n판단", "리뷰어1\n제외코드", "리뷰어1\n메모",
                     "리뷰어2\n판단", "리뷰어2\n제외코드", "리뷰어2\n메모"]
    # 최종 (초록색)
    final_headers = ["최종판단", "최종코드", "메모"]

    all_headers = base_headers + ai_headers + human_headers + final_headers
    n_base = len(base_headers)
    n_ai = len(ai_headers)
    n_human = len(human_headers)

    # ── 헤더 쓰기 ──
    for i, h in enumerate(all_headers, 1):
        if i <= n_base + n_ai:
            apply_header_style(ws, 1, i, h)
        elif i <= n_base + n_ai + n_human:
            apply_header_style(ws, 1, i, h, fill=HUMAN_FILL, font=HUMAN_FONT)
        else:
            apply_header_style(ws, 1, i, h, fill=FINAL_FILL, font=FINAL_FONT)

    # ── 데이터 쓰기 ──
    col_map_base = {"A": "ID", "D": "제목", "G": "연도", "F": "초록"}
    col_map_ai = {"H": "Codex", "I": "Gemini", "J": "AI합의"}
    if sheet_type == "conflict":
        col_map_ai["Q"] = "Claude추천"

    for row_idx, rec in enumerate(records, 2):
        col = 1
        # 기본 정보
        for v4_col in ["A", "D", "G", "F"]:
            c = ws.cell(row=row_idx, column=col, value=rec.get(v4_col, ""))
            apply_data_style(c)
            col += 1
        # AI 참조
        ai_cols = ["H", "I", "J"]
        if sheet_type == "conflict":
            ai_cols.append("Q")
        for v4_col in ai_cols:
            c = ws.cell(row=row_idx, column=col, value=rec.get(v4_col, ""))
            apply_data_style(c)
            c.font = Font(name="맑은 고딕", size=9, color="808080")
            col += 1
        # 리뷰어 데이터 (기존 값 표준화하여 보존)
        r1_decision = standardize_decision(rec.get("S", ""))
        r1_code = rec.get("T", "")
        r1_memo = rec.get("U", "")
        r2_decision = standardize_decision(rec.get("V", ""))
        r2_code = rec.get("W", "")
        r2_memo = rec.get("X", "")

        for val in [r1_decision, r1_code, r1_memo, r2_decision, r2_code, r2_memo]:
            c = ws.cell(row=row_idx, column=col, value=val)
            apply_data_style(c)
            col += 1
        # 최종 판단
        for v4_col in ["Y", "Z"]:
            c = ws.cell(row=row_idx, column=col, value=rec.get(v4_col, ""))
            apply_data_style(c)
            col += 1
        # 최종 메모
        c = ws.cell(row=row_idx, column=col, value="")
        apply_data_style(c)

    # ── 데이터 검증 (드롭다운) ──
    n_rows = len(records) + 1

    # 판단 컬럼: O/X/?
    decision_dv = DataValidation(
        type="list", formula1='"O,X,?"',
        allow_blank=True, showErrorMessage=True,
        errorTitle="입력 오류", error="O(포함), X(제외), ?(불확실) 중 선택하세요."
    )
    decision_dv.prompt = "O=포함, X=제외, ?=불확실"
    decision_dv.promptTitle = "판단 입력"
    ws.add_data_validation(decision_dv)

    # 제외코드 컬럼
    codes_str = ",".join(EXCLUSION_CODES.keys())
    code_dv = DataValidation(
        type="list", formula1=f'"{codes_str}"',
        allow_blank=True, showErrorMessage=True,
        errorTitle="코드 오류", error="유효한 제외코드를 선택하세요."
    )
    code_dv.prompt = "제외 시 코드 선택"
    code_dv.promptTitle = "제외코드"
    ws.add_data_validation(code_dv)

    # 최종판단 컬럼
    final_dv = DataValidation(
        type="list", formula1='"Include,Exclude,Uncertain"',
        allow_blank=True
    )
    ws.add_data_validation(final_dv)

    # 판단 드롭다운 적용 (리뷰어1, 리뷰어2)
    r1_decision_col = n_base + n_ai + 1
    r2_decision_col = n_base + n_ai + 4
    r1_code_col = n_base + n_ai + 2
    r2_code_col = n_base + n_ai + 5
    final_col = n_base + n_ai + n_human + 1

    r1d_letter = get_column_letter(r1_decision_col)
    r2d_letter = get_column_letter(r2_decision_col)
    r1c_letter = get_column_letter(r1_code_col)
    r2c_letter = get_column_letter(r2_code_col)
    f_letter = get_column_letter(final_col)

    decision_dv.add(f"{r1d_letter}2:{r1d_letter}{n_rows}")
    decision_dv.add(f"{r2d_letter}2:{r2d_letter}{n_rows}")
    code_dv.add(f"{r1c_letter}2:{r1c_letter}{n_rows}")
    code_dv.add(f"{r2c_letter}2:{r2c_letter}{n_rows}")
    final_dv.add(f"{f_letter}2:{f_letter}{n_rows}")

    # ── 열 너비 ──
    ws.column_dimensions["A"].width = 12   # ID
    ws.column_dimensions["B"].width = 50   # 제목
    ws.column_dimensions["C"].width = 6    # 연도
    ws.column_dimensions["D"].width = 15   # 초록 (좁게, 확장 가능)
    for i in range(n_base + 1, n_base + n_ai + 1):
        ws.column_dimensions[get_column_letter(i)].width = 10
    for i in range(n_base + n_ai + 1, n_base + n_ai + n_human + 1):
        ws.column_dimensions[get_column_letter(i)].width = 12
    for i in range(n_base + n_ai + n_human + 1, len(all_headers) + 1):
        ws.column_dimensions[get_column_letter(i)].width = 12

    # 행 높이
    ws.row_dimensions[1].height = 30
    for r in range(2, n_rows + 1):
        ws.row_dimensions[r].height = 20

    # 고정 틀 (ID + 제목 고정, 헤더 고정)
    ws.freeze_panes = "C2"

    # 자동 필터
    ws.auto_filter.ref = f"A1:{get_column_letter(len(all_headers))}{n_rows}"

    return ws


def create_ai_reference_sheet(wb, all_records):
    """AI 상세 분석 참조 시트 (스크리닝 후 활용)"""
    ws = wb.create_sheet(title="AI_참조")

    headers = ["ID", "카테고리", "MASEM점수", "감지구인", "방법론",
               "AI도구유형", "교육수준", "시대", "Claude판단", "Claude사유"]

    ref_fill = PatternFill(start_color="D6DCE4", end_color="D6DCE4", fill_type="solid")
    ref_font = Font(name="맑은 고딕", bold=True, color="333333", size=10)

    for i, h in enumerate(headers, 1):
        apply_header_style(ws, 1, i, h, fill=ref_fill, font=ref_font)

    row_idx = 2
    for category, records in all_records:
        for rec in records:
            ws.cell(row=row_idx, column=1, value=rec.get("A", "")).font = DATA_FONT
            ws.cell(row=row_idx, column=2, value=category).font = DATA_FONT
            col_map = {"K": 3, "L": 4, "M": 5, "N": 6, "O": 7, "P": 8, "Q": 9, "R": 10}
            for v4_col, xl_col in col_map.items():
                c = ws.cell(row=row_idx, column=xl_col, value=rec.get(v4_col, ""))
                c.font = DATA_FONT
                c.border = THIN_BORDER
            row_idx += 1

    # 열 너비
    widths = [12, 12, 12, 20, 15, 20, 15, 15, 12, 50]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = "C2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{row_idx - 1}"


def create_guide_sheet(wb):
    """코딩 가이드 시트"""
    ws = wb.create_sheet(title="코딩가이드")

    guide_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    guide_font = Font(name="맑은 고딕", bold=True, color="FFFFFF", size=12)

    # 제목
    c = ws.cell(row=1, column=1, value="스크리닝 코딩 가이드")
    c.fill = guide_fill
    c.font = guide_font
    ws.merge_cells("A1:D1")

    # 판단 기준
    section_font = Font(name="맑은 고딕", bold=True, size=11)
    ws.cell(row=3, column=1, value="1. 판단 입력값").font = section_font

    rows = [
        ("O (Include)", "포함", "AI 채택/수용 관련 TAM/UTAUT 구인을 측정한 실증 연구"),
        ("X (Exclude)", "제외", "아래 제외코드 중 하나에 해당하는 경우"),
        ("? (Uncertain)", "불확실", "초록만으로 판단 불가, 전문(full-text) 확인 필요"),
    ]
    for i, (code, label, desc) in enumerate(rows, 4):
        ws.cell(row=i, column=1, value=code).font = Font(name="맑은 고딕", bold=True, size=10)
        ws.cell(row=i, column=2, value=label).font = DATA_FONT
        ws.cell(row=i, column=3, value=desc).font = DATA_FONT

    # 제외코드
    ws.cell(row=8, column=1, value="2. 제외코드 (Exclude 선택 시)").font = section_font
    row = 9
    for code, desc in EXCLUSION_CODES.items():
        ws.cell(row=row, column=1, value=code).font = Font(name="맑은 고딕", bold=True, size=10)
        ws.cell(row=row, column=2, value=desc).font = DATA_FONT
        row += 1

    # 워크플로우
    ws.cell(row=row + 1, column=1, value="3. 워크플로우").font = section_font
    steps = [
        "① 각 시트(AI_Include, Conflict, Uncertain)를 열고 위에서부터 순서대로 진행",
        "② 제목과 초록을 읽고 '리뷰어 판단' 컬럼에 O/X/? 입력",
        "③ X(제외) 선택 시 반드시 '제외코드' 컬럼에서 코드 선택",
        "④ 필요 시 '메모' 컬럼에 판단 근거 기록",
        "⑤ 모든 행 완료 후 파일 저장",
        "⑥ convert_screening.py 실행하여 CSV 변환",
    ]
    for i, step in enumerate(steps, row + 2):
        ws.cell(row=i, column=1, value=step).font = DATA_FONT

    # 포함 기준
    inc_row = row + 2 + len(steps) + 1
    ws.cell(row=inc_row, column=1, value="4. 포함 기준 (Inclusion Criteria)").font = section_font
    criteria = [
        "- AI(인공지능)가 초점 기술 (ChatGPT, ITS, 생성형 AI 등)",
        "- 교육 맥락 (학생, 교수자, 교육행정가)",
        "- TAM/UTAUT 구인 측정 (PE, EE, SI, FC, BI, Trust 등)",
        "- 실증적 양적 연구 (상관계수 또는 경로계수 추출 가능)",
        "- 2015-2025년 출판",
        "- 영어 논문",
    ]
    for i, crit in enumerate(criteria, inc_row + 1):
        ws.cell(row=i, column=1, value=crit).font = DATA_FONT

    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 15
    ws.column_dimensions["C"].width = 60
    ws.column_dimensions["D"].width = 30


def main():
    # 데이터 로드
    with open(PARSED_DATA, "r") as f:
        data = json.load(f)

    include_recs = data["include"]   # 575
    conflict_recs = data["conflict"]  # 175
    uncertain_recs = data["uncertain"]  # 707

    print(f"Loaded: include={len(include_recs)}, conflict={len(conflict_recs)}, uncertain={len(uncertain_recs)}")

    # 워크북 생성
    wb = Workbook()
    wb.remove(wb.active)  # 기본 시트 제거

    # 1. 스크리닝 시트 3개
    create_screening_sheet(wb, f"AI_Include ({len(include_recs)})", include_recs, "include")
    create_screening_sheet(wb, f"Conflict ({len(conflict_recs)})", conflict_recs, "conflict")
    create_screening_sheet(wb, f"Uncertain ({len(uncertain_recs)})", uncertain_recs, "uncertain")

    # 2. AI 참조 시트
    create_ai_reference_sheet(wb, [
        ("include", include_recs),
        ("conflict", conflict_recs),
        ("uncertain", uncertain_recs),
    ])

    # 3. 코딩 가이드
    create_guide_sheet(wb)

    # 저장
    os.makedirs(os.path.dirname(OUTPUT_PATH), exist_ok=True)
    wb.save(OUTPUT_PATH)
    print(f"\nSaved: {OUTPUT_PATH}")
    print(f"Sheets: {wb.sheetnames}")

    # 통계
    total_human = sum(1 for r in include_recs if r.get("S") or r.get("V"))
    print(f"Preserved human review data: {total_human} records with at least one reviewer entry")


if __name__ == "__main__":
    main()
