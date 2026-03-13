#!/usr/bin/env python3
"""
Generate coder packages v3.0 — Complete redesign per 2026-03-13 discussion.

Creates per-coder folders:
  R1/
  ├── AI_Adoption_MASEM_Coding_v3_R1.xlsx
  ├── AI_Adoption_MASEM_Coding_Manual_v3.docx
  └── PDFs/  (only PDFs assigned to this coder)

Key design decisions:
- 222 studies (3 excluded: S035, S053, S084 — no PDF)
- 66 construct pairs pre-generated per study in CORRELATIONS
- culture_cluster removed entirely
- ai_type: generative, robotic, general only
- construct_1/construct_2 (alphabetical) instead of construct_row/col
- CONSTRUCT_MAPPING removed → study_label_1/2 in CORRELATIONS
- MODERATOR_VARIABLES removed → integrated in STUDY_METADATA
- AI_EXTRACTION_PROVENANCE removed (human coder file only)
- Flag (Y/blank) + notes on every sheet
- Dropdown validation on all categorical variables
- GUIDE quick reference sheet
- Sheet order = coding order
"""

import csv
import itertools
import os
import random
import shutil
from datetime import date
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# ── Paths ──
BASE = Path("/Volumes/External SSD/Projects/Meta-Analysis/jornal_AI-adoption_meta")
CONFIRMED = BASE / "data/02_screening/confirmed_includes.csv"
TRACKER = BASE / "data/04_extraction/pdf_download_tracker.xlsx"
PDF_DIR = BASE / "data/04_extraction/PDFs 2"
MANUAL_DOCX = BASE / "data/04_extraction/AI_Adoption_MASEM_Coding_Manual_v2.docx"
OUTPUT_DIR = BASE / "data/04_extraction/coder_packages"
MAPPING_CSV = BASE / "data/04_extraction/study_id_mapping_v3.csv"

# ── Constants ──
EXCLUDED_SIDS = {"S035", "S053", "S084"}  # No PDF available
CONSTRUCTS = ["ANX", "ATT", "AUT", "BI", "EE", "FC", "PE", "SE", "SI", "TRA", "TRU", "UB"]
CONSTRUCT_PAIRS = list(itertools.combinations(CONSTRUCTS, 2))  # 66 pairs, alphabetical
ROWS_PER_STUDY = len(CONSTRUCT_PAIRS)  # 66

AI_TYPES = ["generative", "robotic", "general"]
STUDY_DESIGNS = ["cross_sectional", "longitudinal", "experimental"]
DATA_COLLECTIONS = ["survey", "experiment", "mixed"]
FRAMEWORKS = ["TAM", "UTAUT", "UTAUT2", "TAM_AI", "UTAUT_AI", "TPB", "SCT", "other"]
SAMPLE_TYPES = ["students", "instructors", "mixed", "administrators"]
EDUCATION_LEVELS = ["K-12", "undergraduate", "graduate", "vocational", "mixed"]
SOURCE_TYPES = ["journal", "conference"]
R_SOURCES = ["direct", "beta_converted", "author_provided"]
SIGNIFICANCES = ["p<.001", "p<.01", "p<.05", "ns", "NR"]
CMB_VALUES = ["addressed", "partial", "not_addressed"]
FLAG_VALUES = ["Y"]

# Styles
NAVY = "1B3A5C"
BLUE = "2E74B5"
LIGHT_BLUE = "D6E4F0"
ALT_ROW = "F2F7FB"
ORANGE_BG = "FFF3E0"
GREEN_BG = "E8F5E9"
WHITE = "FFFFFF"

HEADER_FILL = PatternFill(start_color=NAVY, end_color=NAVY, fill_type="solid")
HEADER_FONT = Font(name="Calibri", size=10, bold=True, color=WHITE)
BODY_FONT = Font(name="Calibri", size=10)
MONO_FONT = Font(name="Consolas", size=10)
LINK_FONT = Font(name="Calibri", size=10, color="0563C1", underline="single")
TITLE_FONT = Font(name="Calibri", size=14, bold=True, color=NAVY)
SECTION_FONT = Font(name="Calibri", size=11, bold=True, color=BLUE)

random.seed(2026)


def load_studies():
    """Load confirmed includes, assign study_ids, exclude missing PDFs."""
    with open(CONFIRMED) as f:
        rows = list(csv.DictReader(f))
    rows.sort(key=lambda r: (r["year"], r["record_id"]))
    studies = []
    for i, row in enumerate(rows, 1):
        sid = f"S{i:03d}"
        if sid in EXCLUDED_SIDS:
            continue
        row["study_id"] = sid
        row["doi_clean"] = row.get("doi_url", "").split("doi.org/", 1)[-1] if "doi.org/" in row.get("doi_url", "") else row.get("doi_url", "")
        studies.append(row)
    return studies


def select_calibration(studies, n=10):
    """Select 10 calibration studies stratified by year."""
    by_year = {}
    for s in studies:
        by_year.setdefault(s["year"], []).append(s)
    allocation = {"2022": 1, "2023": 1, "2024": 3, "2025": 4, "2026": 1}
    selected = []
    for year, count in allocation.items():
        pool = by_year.get(year, [])
        if pool:
            selected.extend(random.sample(pool, min(count, len(pool))))
    while len(selected) < n:
        remaining = [s for s in by_year.get("2025", []) if s not in selected]
        if remaining:
            selected.append(random.choice(remaining))
        else:
            break
    return selected[:n]


def assign_phases(studies, cal_ids):
    """Assign studies to phases."""
    remaining = [s for s in studies if s["study_id"] not in cal_ids]
    random.shuffle(remaining)
    phase1 = remaining[:100]
    phase2 = remaining[100:]
    pair_a = phase1[:50]
    pair_b = phase1[50:100]
    n = len(phase2)
    q, r = divmod(n, 4)
    splits = []
    offset = 0
    for i in range(4):
        size = q + (1 if i < r else 0)
        splits.append(phase2[offset:offset + size])
        offset += size
    return {
        "pair_a": pair_a, "pair_b": pair_b,
        "p2_r1": splits[0], "p2_r2": splits[1],
        "p2_r3": splits[2], "p2_r4": splits[3],
    }


def style_header(ws, ncols):
    """Style header row."""
    for col in range(1, ncols + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


def add_dropdown(ws, col_letter, values, max_row):
    """Add dropdown validation to a column."""
    formula = '"' + ",".join(values) + '"'
    dv = DataValidation(type="list", formula1=formula, allow_blank=True)
    dv.error = "Please select from the dropdown list"
    dv.errorTitle = "Invalid Value"
    dv.prompt = "Select a value"
    dv.showDropDown = False  # False = show dropdown arrow
    ws.add_data_validation(dv)
    dv.add(f"{col_letter}2:{col_letter}{max_row}")


def set_col_width(ws, col, width):
    ws.column_dimensions[get_column_letter(col)].width = width


# ══════════════════════════════════════════════════════
# SHEET BUILDERS
# ══════════════════════════════════════════════════════

def build_guide_sheet(wb, coder_label):
    """GUIDE: Quick reference card."""
    ws = wb.create_sheet("GUIDE", 0)

    # Title
    ws.merge_cells("A1:F1")
    c = ws.cell(row=1, column=1, value=f"MASEM Coding Quick Guide — {coder_label}")
    c.font = Font(name="Calibri", size=16, bold=True, color=NAVY)
    c.alignment = Alignment(horizontal="center")

    row = 3
    # Coding steps
    ws.cell(row=row, column=1, value="Coding Steps").font = SECTION_FONT
    row += 1
    steps = [
        "Step 1:  STUDY_METADATA tab → fill study-level variables",
        "Step 2:  CORRELATIONS tab → extract r values for each construct pair",
        "Step 3:  Back to STUDY_METADATA → fill §5.5 (n_constructs, matrix_completeness, CMB)",
        "Step 4:  ASSIGNMENT tab → mark Status as done",
        "If excluding:  EXCLUSION_LOG tab → record reason",
        "If disagreement:  DISCREPANCY_LOG tab → record details",
    ]
    for step in steps:
        ws.cell(row=row, column=1, value=step).font = BODY_FONT
        row += 1

    row += 1
    ws.cell(row=row, column=1, value="12 Target Constructs").font = SECTION_FONT
    row += 1
    constructs_info = [
        ("PE", "Performance Expectancy", "Belief AI helps attain performance gains"),
        ("EE", "Effort Expectancy", "Perceived ease of using AI"),
        ("SI", "Social Influence", "Important others believe one should use AI"),
        ("FC", "Facilitating Conditions", "Organizational/technical infrastructure"),
        ("BI", "Behavioral Intention", "Strength of intention to adopt AI"),
        ("UB", "Use Behavior", "Actual use of AI technology"),
        ("ATT", "Attitude", "Overall evaluative judgment about AI"),
        ("SE", "Self-Efficacy", "Belief in own capability to use AI"),
        ("TRU", "AI Trust", "Willingness to be vulnerable based on AI expectations"),
        ("ANX", "AI Anxiety", "Apprehension or fear about AI"),
        ("TRA", "AI Transparency", "Perceived ability to understand AI decisions"),
        ("AUT", "Perceived AI Autonomy", "Perceived degree of AI independent operation"),
    ]
    for abbr, name, defn in constructs_info:
        ws.cell(row=row, column=1, value=abbr).font = Font(name="Consolas", size=10, bold=True, color=BLUE)
        ws.cell(row=row, column=2, value=name).font = BODY_FONT
        ws.cell(row=row, column=3, value=defn).font = Font(name="Calibri", size=9, color="666666")
        row += 1

    row += 1
    ws.cell(row=row, column=1, value="Key Formulas & Rules").font = SECTION_FONT
    row += 1
    rules = [
        "Beta-to-r:  r = β + 0.05 × λ  (λ=1 if β≥0, λ=−1 if β<0)",
        "matrix_completeness = n_correlations / (n_constructs × (n_constructs − 1) / 2)",
        "CMB: Ctrl+F → \"Harman\", \"CMV\", \"CMB\", \"common method\"",
        "Flag: Enter Y in flag column + describe in notes column",
        "Construct pairs: alphabetical order (construct_1 < construct_2)",
        "Unreported pairs: leave r_value blank (do NOT enter 0)",
    ]
    for rule in rules:
        ws.cell(row=row, column=1, value=rule).font = BODY_FONT
        row += 1

    set_col_width(ws, 1, 12)
    set_col_width(ws, 2, 25)
    set_col_width(ws, 3, 50)


def build_assignment_sheet(wb, calibration, phase1, phase2, coder_label, pair_label):
    """ASSIGNMENT: Task list with progress tracking."""
    ws = wb.create_sheet("ASSIGNMENT")

    headers = ["Study ID", "Title", "Year", "DOI", "Phase", "PDF", "Status", "Flag", "Notes"]
    for i, h in enumerate(headers, 1):
        ws.cell(row=1, column=i, value=h)
    style_header(ws, len(headers))

    row = 2
    phase_colors = {
        "Phase 0": PatternFill(start_color=ORANGE_BG, end_color=ORANGE_BG, fill_type="solid"),
        "Phase 1": PatternFill(start_color=LIGHT_BLUE, end_color=LIGHT_BLUE, fill_type="solid"),
        "Phase 2": None,
    }

    all_studies = []
    for s in calibration:
        all_studies.append((s, "Phase 0: Calibration"))
    for s in phase1:
        all_studies.append((s, f"Phase 1: {pair_label}"))
    for s in phase2:
        all_studies.append((s, "Phase 2: Single"))

    for s, phase in all_studies:
        ws.cell(row=row, column=1, value=s["study_id"]).font = Font(name="Consolas", size=10, bold=True)
        ws.cell(row=row, column=2, value=s["title"][:80]).font = BODY_FONT
        ws.cell(row=row, column=3, value=int(s["year"])).font = BODY_FONT
        # DOI as hyperlink
        doi_url = s.get("doi_url", "")
        cell = ws.cell(row=row, column=4, value=s.get("doi_clean", ""))
        if doi_url:
            cell.hyperlink = doi_url
            cell.font = LINK_FONT
        ws.cell(row=row, column=5, value=phase).font = BODY_FONT
        ws.cell(row=row, column=6, value=f"{s['study_id']}.pdf").font = Font(name="Consolas", size=9)
        ws.cell(row=row, column=7, value="").font = BODY_FONT  # Status
        # Phase color
        phase_key = phase.split(":")[0]
        fill = phase_colors.get(phase_key)
        if fill:
            for col in range(1, len(headers) + 1):
                ws.cell(row=row, column=col).fill = fill
        row += 1

    # Summary at bottom
    row += 1
    ws.cell(row=row, column=1, value="Summary").font = SECTION_FONT
    ws.cell(row=row + 1, column=1, value=f"Phase 0 (Calibration): {len(calibration)}").font = BODY_FONT
    ws.cell(row=row + 2, column=1, value=f"Phase 1 (Dual Coding): {len(phase1)}").font = BODY_FONT
    ws.cell(row=row + 3, column=1, value=f"Phase 2 (Single): {len(phase2)}").font = BODY_FONT
    ws.cell(row=row + 4, column=1, value=f"Total: {len(all_studies)}").font = Font(name="Calibri", size=10, bold=True)

    # Status dropdown
    add_dropdown(ws, "G", ["done", "in_progress", "excluded"], row - 1)
    add_dropdown(ws, "H", FLAG_VALUES, row - 1)

    set_col_width(ws, 1, 10)
    set_col_width(ws, 2, 55)
    set_col_width(ws, 3, 6)
    set_col_width(ws, 4, 30)
    set_col_width(ws, 5, 22)
    set_col_width(ws, 6, 12)
    set_col_width(ws, 7, 12)
    set_col_width(ws, 8, 6)
    set_col_width(ws, 9, 30)


def build_study_metadata_sheet(wb, coder_studies, coder_label):
    """STUDY_METADATA: One row per study, AI pre-filled + human coded."""
    ws = wb.create_sheet("STUDY_METADATA")

    headers = [
        # Identification (AI Pre-Coded)
        "study_id", "first_author", "year", "title", "doi", "source_type",
        # Study Design (Human Coded)
        "study_design", "data_collection", "theoretical_framework",
        # Sample & Context (Human Coded)
        "sample_size", "sample_type", "country", "education_level",
        # AI Technology (Human Coded)
        "ai_type", "ai_tool_name",
        # Moderators (Human Coded)
        "temporal_period", "user_role",
        # Quality Assessment (Human Coded — after §6)
        "n_constructs_measured", "n_correlations_reported", "matrix_completeness",
        "common_method_bias",
        # Source Management
        "human_coder", "coding_date", "coding_phase",
        # Flag & Notes
        "flag", "notes",
    ]
    for i, h in enumerate(headers, 1):
        ws.cell(row=1, column=i, value=h)
    style_header(ws, len(headers))

    # Pre-populate
    for idx, s in enumerate(coder_studies, 2):
        ws.cell(row=idx, column=1, value=s["study_id"]).font = Font(name="Consolas", size=10)
        ws.cell(row=idx, column=3, value=int(s["year"])).font = BODY_FONT
        ws.cell(row=idx, column=4, value=s["title"]).font = BODY_FONT
        ws.cell(row=idx, column=5, value=s.get("doi_clean", "")).font = BODY_FONT
        ws.cell(row=idx, column=22, value=coder_label).font = BODY_FONT
        # Temporal period auto-derived from year
        year = int(s["year"])
        ws.cell(row=idx, column=16, value="pre_chatgpt" if year <= 2022 else "post_chatgpt").font = BODY_FONT

    max_row = len(coder_studies) + 1

    # Dropdowns
    col_map = {h: i for i, h in enumerate(headers, 1)}
    add_dropdown(ws, get_column_letter(col_map["source_type"]), SOURCE_TYPES, max_row)
    add_dropdown(ws, get_column_letter(col_map["study_design"]), STUDY_DESIGNS, max_row)
    add_dropdown(ws, get_column_letter(col_map["data_collection"]), DATA_COLLECTIONS, max_row)
    add_dropdown(ws, get_column_letter(col_map["theoretical_framework"]), FRAMEWORKS, max_row)
    add_dropdown(ws, get_column_letter(col_map["sample_type"]), SAMPLE_TYPES, max_row)
    add_dropdown(ws, get_column_letter(col_map["education_level"]), EDUCATION_LEVELS, max_row)
    add_dropdown(ws, get_column_letter(col_map["ai_type"]), AI_TYPES, max_row)
    add_dropdown(ws, get_column_letter(col_map["common_method_bias"]), CMB_VALUES, max_row)
    add_dropdown(ws, get_column_letter(col_map["user_role"]), ["student", "instructor", "both"], max_row)
    add_dropdown(ws, get_column_letter(col_map["flag"]), FLAG_VALUES, max_row)

    # Column widths
    widths = {
        "study_id": 10, "first_author": 14, "year": 6, "title": 50, "doi": 25,
        "source_type": 12, "study_design": 16, "data_collection": 14,
        "theoretical_framework": 18, "sample_size": 12, "sample_type": 14,
        "country": 16, "education_level": 16, "ai_type": 14, "ai_tool_name": 16,
        "temporal_period": 16, "user_role": 12,
        "n_constructs_measured": 18, "n_correlations_reported": 20,
        "matrix_completeness": 18, "common_method_bias": 18,
        "human_coder": 12, "coding_date": 12, "coding_phase": 14,
        "flag": 6, "notes": 30,
    }
    for h, w in widths.items():
        if h in col_map:
            set_col_width(ws, col_map[h], w)


def build_correlations_sheet(wb, coder_studies):
    """CORRELATIONS: 66 pairs × N studies, pre-generated."""
    ws = wb.create_sheet("CORRELATIONS")

    headers = [
        "study_id",
        "construct_1", "construct_2",
        "r_value", "r_source", "original_beta",
        "p_value", "significance",
        "study_label_1", "study_label_2",
        "source_location",
        "flag", "notes",
    ]
    for i, h in enumerate(headers, 1):
        ws.cell(row=1, column=i, value=h)
    style_header(ws, len(headers))

    row = 2
    alt_fill = PatternFill(start_color=ALT_ROW, end_color=ALT_ROW, fill_type="solid")
    study_sep_fill = PatternFill(start_color=LIGHT_BLUE, end_color=LIGHT_BLUE, fill_type="solid")

    for s_idx, study in enumerate(coder_studies):
        sid = study["study_id"]
        for p_idx, (c1, c2) in enumerate(CONSTRUCT_PAIRS):
            ws.cell(row=row, column=1, value=sid).font = Font(name="Consolas", size=10)
            ws.cell(row=row, column=2, value=c1).font = Font(name="Consolas", size=10)
            ws.cell(row=row, column=3, value=c2).font = Font(name="Consolas", size=10)
            # Light shading for first row of each study block
            if p_idx == 0:
                for col in range(1, len(headers) + 1):
                    ws.cell(row=row, column=col).fill = study_sep_fill
            row += 1

    max_row = row - 1

    # Dropdowns
    add_dropdown(ws, "B", CONSTRUCTS, max_row)  # construct_1
    add_dropdown(ws, "C", CONSTRUCTS, max_row)  # construct_2
    add_dropdown(ws, "E", R_SOURCES, max_row)   # r_source
    add_dropdown(ws, "H", SIGNIFICANCES, max_row)  # significance
    add_dropdown(ws, "L", FLAG_VALUES, max_row)  # flag

    # Column widths
    set_col_width(ws, 1, 10)   # study_id
    set_col_width(ws, 2, 14)   # construct_1
    set_col_width(ws, 3, 14)   # construct_2
    set_col_width(ws, 4, 10)   # r_value
    set_col_width(ws, 5, 16)   # r_source
    set_col_width(ws, 6, 14)   # original_beta
    set_col_width(ws, 7, 10)   # p_value
    set_col_width(ws, 8, 12)   # significance
    set_col_width(ws, 9, 22)   # study_label_1
    set_col_width(ws, 10, 22)  # study_label_2
    set_col_width(ws, 11, 16)  # source_location
    set_col_width(ws, 12, 6)   # flag
    set_col_width(ws, 13, 30)  # notes

    # Group by study (outline) for easy collapse
    row = 2
    for study in coder_studies:
        start = row + 1  # second row of block
        end = row + ROWS_PER_STUDY - 1
        if end > start:
            ws.row_dimensions.group(start, end, outline_level=1, hidden=False)
        row += ROWS_PER_STUDY

    print(f"    CORRELATIONS: {max_row} rows ({len(coder_studies)} studies × {ROWS_PER_STUDY} pairs)")


def build_exclusion_log(wb):
    """EXCLUSION_LOG sheet."""
    ws = wb.create_sheet("EXCLUSION_LOG")
    headers = ["study_id", "first_author", "year", "title",
               "exclusion_stage", "exclusion_code", "detailed_rationale", "flag", "notes"]
    for i, h in enumerate(headers, 1):
        ws.cell(row=1, column=i, value=h)
    style_header(ws, len(headers))

    excl_stages = ["full_text", "data_extraction"]
    excl_codes = ["E-FT1", "E-FT2", "E-FT3", "E-FT4", "E-FT5", "E-FT6",
                  "E1", "E2", "E3", "E4", "E5", "E6", "E7", "E8", "E9", "E10", "E11", "E12"]
    add_dropdown(ws, "E", excl_stages, 200)
    add_dropdown(ws, "F", excl_codes, 200)
    add_dropdown(ws, "H", FLAG_VALUES, 200)

    set_col_width(ws, 1, 10)
    set_col_width(ws, 2, 14)
    set_col_width(ws, 3, 6)
    set_col_width(ws, 4, 50)
    set_col_width(ws, 5, 16)
    set_col_width(ws, 6, 14)
    set_col_width(ws, 7, 40)
    set_col_width(ws, 8, 6)
    set_col_width(ws, 9, 30)


def build_discrepancy_log(wb):
    """DISCREPANCY_LOG sheet."""
    ws = wb.create_sheet("DISCREPANCY_LOG")
    headers = ["study_id", "variable_name", "my_value", "other_coder_value",
               "discrepancy_type", "resolution", "resolved_by", "resolution_date",
               "flag", "notes"]
    for i, h in enumerate(headers, 1):
        ws.cell(row=1, column=i, value=h)
    style_header(ws, len(headers))

    add_dropdown(ws, "E", ["clerical", "interpretive", "substantive"], 200)
    add_dropdown(ws, "I", FLAG_VALUES, 200)

    set_col_width(ws, 1, 10)
    set_col_width(ws, 2, 18)
    set_col_width(ws, 3, 16)
    set_col_width(ws, 4, 18)
    set_col_width(ws, 5, 16)
    set_col_width(ws, 6, 30)
    set_col_width(ws, 7, 12)
    set_col_width(ws, 8, 14)
    set_col_width(ws, 9, 6)
    set_col_width(ws, 10, 30)


def build_codebook_sheet(wb):
    """CODEBOOK: Variable reference (read-only)."""
    ws = wb.create_sheet("CODEBOOK")
    headers = ["variable_name", "sheet", "type", "valid_values", "coding_rules", "example"]
    for i, h in enumerate(headers, 1):
        ws.cell(row=1, column=i, value=h)
    style_header(ws, len(headers))

    entries = [
        # STUDY_METADATA
        ["— STUDY_METADATA —", "", "", "", "", ""],
        ["study_id", "STUDY_METADATA", "string", "S001–S225", "Pre-assigned, do not change", "S001"],
        ["first_author", "STUDY_METADATA", "string", "", "Last name of first author", "Kim"],
        ["year", "STUDY_METADATA", "integer", "2022–2026", "Publication year", "2024"],
        ["title", "STUDY_METADATA", "string", "", "Full paper title", "AI Adoption..."],
        ["doi", "STUDY_METADATA", "string", "", "DOI without https://doi.org/", "10.1000/xyz"],
        ["source_type", "STUDY_METADATA", "categorical", "journal, conference", "", "journal"],
        ["study_design", "STUDY_METADATA", "categorical", ", ".join(STUDY_DESIGNS), "", "cross_sectional"],
        ["data_collection", "STUDY_METADATA", "categorical", ", ".join(DATA_COLLECTIONS), "", "survey"],
        ["theoretical_framework", "STUDY_METADATA", "categorical", ", ".join(FRAMEWORKS), "", "UTAUT2"],
        ["sample_size", "STUDY_METADATA", "integer", "≥50", "Final analytic sample N", "384"],
        ["sample_type", "STUDY_METADATA", "categorical", ", ".join(SAMPLE_TYPES), "", "students"],
        ["country", "STUDY_METADATA", "string", "", "Country of data collection", "South Korea"],
        ["education_level", "STUDY_METADATA", "categorical", ", ".join(EDUCATION_LEVELS), "", "undergraduate"],
        ["ai_type", "STUDY_METADATA", "categorical", ", ".join(AI_TYPES), "If multiple types → general", "generative"],
        ["ai_tool_name", "STUDY_METADATA", "string", "", "Specific tool if named", "ChatGPT"],
        ["temporal_period", "STUDY_METADATA", "categorical", "pre_chatgpt, post_chatgpt", "Auto-filled from year. 2022=pre, 2023+=post", "post_chatgpt"],
        ["user_role", "STUDY_METADATA", "categorical", "student, instructor, both", "", "student"],
        ["n_constructs_measured", "STUDY_METADATA", "integer", "2–12", "Count of our 12 constructs found. Fill AFTER §6.", "6"],
        ["n_correlations_reported", "STUDY_METADATA", "integer", "", "Count of rows filled in CORRELATIONS. Fill AFTER §6.", "15"],
        ["matrix_completeness", "STUDY_METADATA", "float", "0–1", "= n_correlations / (n_constructs*(n_constructs-1)/2)", "0.80"],
        ["common_method_bias", "STUDY_METADATA", "categorical", ", ".join(CMB_VALUES), "Ctrl+F: Harman, CMV, CMB, common method", "addressed"],
        ["", "", "", "", "", ""],
        # CORRELATIONS
        ["— CORRELATIONS —", "", "", "", "", ""],
        ["study_id", "CORRELATIONS", "string", "S001–S225", "Pre-filled. Each study has 66 rows (all possible pairs).", "S001"],
        ["construct_1", "CORRELATIONS", "categorical", ", ".join(CONSTRUCTS), "Alphabetically first construct. Pre-filled.", "ANX"],
        ["construct_2", "CORRELATIONS", "categorical", ", ".join(CONSTRUCTS), "Alphabetically second construct. Pre-filled.", "BI"],
        ["r_value", "CORRELATIONS", "float", "-1 to 1", "Pearson r to 2 decimals. Leave blank if not reported.", "0.52"],
        ["r_source", "CORRELATIONS", "categorical", ", ".join(R_SOURCES), "", "direct"],
        ["original_beta", "CORRELATIONS", "float", "", "Only if r_source=beta_converted", "0.40"],
        ["p_value", "CORRELATIONS", "float", "", "If reported", "0.001"],
        ["significance", "CORRELATIONS", "categorical", ", ".join(SIGNIFICANCES), "", "p<.001"],
        ["study_label_1", "CORRELATIONS", "string", "", "Original construct name for construct_1 in paper", "AI Anxiety"],
        ["study_label_2", "CORRELATIONS", "string", "", "Original construct name for construct_2 in paper", "Behavioral Intention"],
        ["source_location", "CORRELATIONS", "string", "", "Where in paper (table/page)", "Table 3"],
    ]

    for r_idx, entry in enumerate(entries, 2):
        for c_idx, val in enumerate(entry, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            if entry[0].startswith("—"):
                cell.font = SECTION_FONT
            else:
                cell.font = BODY_FONT

    set_col_width(ws, 1, 22)
    set_col_width(ws, 2, 18)
    set_col_width(ws, 3, 12)
    set_col_width(ws, 4, 35)
    set_col_width(ws, 5, 45)
    set_col_width(ws, 6, 16)


# ══════════════════════════════════════════════════════
# MAIN
# ══════════════════════════════════════════════════════

def create_coder_package(coder_label, calibration, phase1, phase2, pair_label):
    """Create complete coder package: Excel + Manual + PDFs."""
    coder_dir = OUTPUT_DIR / coder_label
    pdf_coder_dir = coder_dir / "PDFs"
    coder_dir.mkdir(parents=True, exist_ok=True)
    pdf_coder_dir.mkdir(parents=True, exist_ok=True)

    coder_studies = calibration + phase1 + phase2

    # Create workbook
    wb = openpyxl.Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    print(f"\n{coder_label}: {len(coder_studies)} studies (Cal:{len(calibration)} P1:{len(phase1)} P2:{len(phase2)})")

    build_guide_sheet(wb, coder_label)
    build_assignment_sheet(wb, calibration, phase1, phase2, coder_label, pair_label)
    build_study_metadata_sheet(wb, coder_studies, coder_label)
    build_correlations_sheet(wb, coder_studies)
    build_exclusion_log(wb)
    build_discrepancy_log(wb)
    build_codebook_sheet(wb)

    xlsx_path = coder_dir / f"AI_Adoption_MASEM_Coding_v3_{coder_label}.xlsx"
    wb.save(xlsx_path)
    print(f"  Excel: {xlsx_path.name}")

    # Copy manual
    manual_dest = coder_dir / "AI_Adoption_MASEM_Coding_Manual_v3.docx"
    shutil.copy2(MANUAL_DOCX, manual_dest)
    print(f"  Manual: {manual_dest.name}")

    # Copy assigned PDFs
    copied = 0
    for s in coder_studies:
        pdf_name = f"{s['study_id']}.pdf"
        src = PDF_DIR / pdf_name
        if src.exists():
            shutil.copy2(src, pdf_coder_dir / pdf_name)
            copied += 1
    print(f"  PDFs: {copied}/{len(coder_studies)} copied")

    return coder_studies


def main():
    studies = load_studies()
    print(f"Loaded {len(studies)} studies (after excluding {EXCLUDED_SIDS})")

    # Calibration
    calibration = select_calibration(studies, n=10)
    cal_ids = {s["study_id"] for s in calibration}
    print(f"\nPhase 0 Calibration: {len(calibration)} studies")
    for s in calibration:
        print(f"  {s['study_id']} [{s['year']}] {s['title'][:65]}...")

    # Phase assignment
    phases = assign_phases(studies, cal_ids)
    print(f"\nPhase 1: Pair A {len(phases['pair_a'])} + Pair B {len(phases['pair_b'])} = {len(phases['pair_a'])+len(phases['pair_b'])}")
    print(f"Phase 2: R1:{len(phases['p2_r1'])} R2:{len(phases['p2_r2'])} R3:{len(phases['p2_r3'])} R4:{len(phases['p2_r4'])}")

    # Clean output dir
    if OUTPUT_DIR.exists():
        shutil.rmtree(OUTPUT_DIR)

    # Generate packages
    create_coder_package("R1", calibration, phases["pair_a"], phases["p2_r1"], "Pair A (R1+R2)")
    create_coder_package("R2", calibration, phases["pair_a"], phases["p2_r2"], "Pair A (R1+R2)")
    create_coder_package("R3", calibration, phases["pair_b"], phases["p2_r3"], "Pair B (R3+R4)")
    create_coder_package("R4", calibration, phases["pair_b"], phases["p2_r4"], "Pair B (R3+R4)")

    # Save mapping CSV
    phase_map = {}
    for s in calibration:
        phase_map[s["study_id"]] = "Phase 0"
    for s in phases["pair_a"]:
        phase_map[s["study_id"]] = "Phase 1: Pair A"
    for s in phases["pair_b"]:
        phase_map[s["study_id"]] = "Phase 1: Pair B"
    for key in ["p2_r1", "p2_r2", "p2_r3", "p2_r4"]:
        coder = key.split("_")[1].upper()
        for s in phases[key]:
            phase_map[s["study_id"]] = f"Phase 2: {coder}"

    with open(MAPPING_CSV, "w", newline="") as f:
        writer = csv.writer(f)
        writer.writerow(["study_id", "record_id", "title", "year", "doi", "phase", "coders"])
        for s in studies:
            sid = s["study_id"]
            phase = phase_map.get(sid, "")
            if sid in cal_ids:
                coders = "R1, R2, R3, R4"
            elif any(s is x for x in phases["pair_a"]):
                coders = "R1, R2"
            elif any(s is x for x in phases["pair_b"]):
                coders = "R3, R4"
            else:
                for key in ["p2_r1", "p2_r2", "p2_r3", "p2_r4"]:
                    if any(s is x for x in phases[key]):
                        coders = key.split("_")[1].upper()
                        break
                else:
                    coders = ""
            writer.writerow([sid, s["record_id"], s["title"], s["year"], s.get("doi_clean", ""), phase, coders])

    print(f"\nMapping: {MAPPING_CSV.name}")
    print(f"\nDone! Packages at: {OUTPUT_DIR}")
    for d in sorted(OUTPUT_DIR.iterdir()):
        if d.is_dir():
            files = list(d.iterdir())
            pdfs = len(list((d / "PDFs").iterdir())) if (d / "PDFs").exists() else 0
            print(f"  {d.name}/ — {len(files)} items, {pdfs} PDFs")


if __name__ == "__main__":
    main()
