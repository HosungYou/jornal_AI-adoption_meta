#!/usr/bin/env python3
"""
Claude 스크리닝 결과를 v5 Excel에 반영 + BIBD 블록 할당

파이프라인:
  1,457건 (DB 검색)
    → 3-AI Screening (Codex + Gemini + Claude)
    → 2/3 majority vote
    → consensus="include"만 human full-text review 대상
    → BIBD (2 of 3 reviewers)

실행 순서:
  1. claude_screening.py → claude_screening_results.csv 생성
  2. 이 스크립트 → v5 Excel 업데이트 + BIBD 블록 할당

사용법:
    python3 update_v5_with_claude.py
    python3 update_v5_with_claude.py --stats-only  # 통계만 확인
"""

import argparse
import csv
import json
import os
import random
import zipfile
import xml.etree.ElementTree as ET
from collections import Counter
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

REPO = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
EXCEL_PATH = os.path.join(REPO, "data", "templates", "human_review_sheet_v5.xlsx")
RESULTS_CSV = os.path.join(REPO, "paper_a", "data", "02_screening", "claude_screening_results.csv")
QUEUE_CSV = os.path.join(REPO, "data", "03_screening", "human_review_queue.csv")
V4_PARSED = "/tmp/v4_parsed.json"

THIN = Border(*(Side(style="thin", color="D9D9D9"),) * 4)


def load_all_claude_decisions():
    """Claude 평가 결과 전체 로드 (v4 기존 + 새 API 결과)"""
    decisions = {}

    # 1. v4 기존 219건
    if os.path.exists(V4_PARSED):
        with open(V4_PARSED, "r") as f:
            data = json.load(f)
        cl_map = {"EXCLUDE": "exclude", "KEEP - TAM/UTAUT 구인 있음": "include",
                   "REVIEW - 판단 불가 (full-text 필요)": "uncertain"}
        for rec in data.get("include", []):
            q = rec.get("Q", "").strip()
            if q:
                decisions[rec["A"]] = {
                    "decision": cl_map.get(q, "uncertain"),
                    "code": "",
                    "reason": rec.get("R", ""),
                }

    # 2. 새 API 결과
    if os.path.exists(RESULTS_CSV):
        with open(RESULTS_CSV, "r", encoding="utf-8") as f:
            for row in csv.DictReader(f):
                decisions[row["record_id"]] = {
                    "decision": row["decision"],
                    "code": row.get("code", ""),
                    "reason": row.get("reason", ""),
                }

    return decisions


def compute_3model_consensus(codex, gemini, claude):
    """3모델 다수결 합의"""
    votes = [v for v in [codex, gemini, claude] if v in ("include", "exclude", "uncertain")]
    if len(votes) < 2:
        return votes[0] if votes else ""

    vc = Counter(votes)
    majority = vc.most_common(1)[0]
    if majority[1] >= 2:
        return majority[0]
    return "conflict"


def assign_bibd_blocks(record_ids, seed=42):
    """
    Balanced Incomplete Block Design (BIBD)
    3개 블록으로 나눠서 각 레코드가 정확히 2명에게 할당

    Block A: R1 + R2
    Block B: R2 + R3
    Block C: R1 + R3
    """
    random.seed(seed)
    ids = list(record_ids)
    random.shuffle(ids)

    n = len(ids)
    block_size = n // 3
    remainder = n % 3

    # 균등 분배
    block_a = ids[:block_size + (1 if remainder > 0 else 0)]
    start_b = len(block_a)
    block_b = ids[start_b:start_b + block_size + (1 if remainder > 1 else 0)]
    block_c = ids[start_b + len(block_b):]

    assignments = {}
    for rid in block_a:
        assignments[rid] = {"block": "A", "R1": True, "R2": True, "R3": False}
    for rid in block_b:
        assignments[rid] = {"block": "B", "R1": False, "R2": True, "R3": True}
    for rid in block_c:
        assignments[rid] = {"block": "C", "R1": True, "R2": False, "R3": True}

    return assignments


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--stats-only", action="store_true")
    args = parser.parse_args()

    # 데이터 로드
    claude_decisions = load_all_claude_decisions()
    print(f"Claude 평가 전체: {len(claude_decisions)}건")

    # CSV에서 Codex/Gemini 데이터
    csv_data = {}
    with open(QUEUE_CSV, "r", encoding="utf-8") as f:
        for row in csv.DictReader(f):
            csv_data[row["record_id"]] = row

    # 3모델 합의 계산
    consensus_dist = Counter()
    human_review_ids = []  # 인간 검토 대상

    for rid, csv_rec in csv_data.items():
        codex = csv_rec.get("screen_decision_codex", "").strip()
        gemini = csv_rec.get("screen_decision_gemini", "").strip()
        claude = claude_decisions.get(rid, {}).get("decision", "")

        consensus = compute_3model_consensus(codex, gemini, claude)
        consensus_dist[consensus] += 1

        # 인간 검토 대상: consensus="include"만 (2/3 이상 include)
        # uncertain/conflict/exclude는 AI 단계에서 제외
        if consensus == "include":
            human_review_ids.append(rid)

    print(f"\n=== 3모델 합의 분포 ===")
    for k, v in sorted(consensus_dist.items(), key=lambda x: -x[1]):
        print(f"  {k}: {v}")
    print(f"\n인간 검토 대상: {len(human_review_ids)}건")

    # BIBD 블록 할당
    bibd = assign_bibd_blocks(human_review_ids)
    block_sizes = Counter(v["block"] for v in bibd.values())
    r1_count = sum(1 for v in bibd.values() if v["R1"])
    r2_count = sum(1 for v in bibd.values() if v["R2"])
    r3_count = sum(1 for v in bibd.values() if v["R3"])

    print(f"\n=== BIBD 할당 ===")
    print(f"  Block A (R1+R2): {block_sizes['A']}건")
    print(f"  Block B (R2+R3): {block_sizes['B']}건")
    print(f"  Block C (R1+R3): {block_sizes['C']}건")
    print(f"  R1(PI): {r1_count}건")
    print(f"  R2(PhD-1): {r2_count}건")
    print(f"  R3(PhD-2): {r3_count}건")

    if args.stats_only:
        print("\n(Stats only mode)")
        return

    # ── v5 Excel 업데이트 ──
    print(f"\nUpdating Excel: {EXCEL_PATH}")
    wb = load_workbook(EXCEL_PATH)
    ws = wb["스크리닝"]

    # 기존 헤더 확인
    existing_cols = ws.max_column
    print(f"  현재 컬럼 수: {existing_cols}")

    # Claude 컬럼(7), AI합의 컬럼(8) 업데이트 + 블록/할당 컬럼 추가
    # 새 컬럼: 블록(S열=19), R1할당(T=20), R2할당(U=21), R3할당(V=22)
    new_headers = [
        ("블록", 19), ("R1\n할당", 20), ("R2\n할당", 21), ("R3\n할당", 22)
    ]

    block_fill = PatternFill("solid", fgColor="E8DAEF")
    block_font = Font("맑은 고딕", bold=True, color="6C3483", size=10)

    for label, col in new_headers:
        c = ws.cell(row=1, column=col, value=label)
        c.fill = block_fill
        c.font = block_font
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = THIN
        ws.column_dimensions[get_column_letter(col)].width = 8

    # 데이터 업데이트
    updated_claude = 0
    updated_consensus = 0

    for row in range(2, ws.max_row + 1):
        rid = ws.cell(row=row, column=1).value
        if not rid:
            continue

        # Claude 컬럼 업데이트 (새 결과로 덮어쓰기)
        cd = claude_decisions.get(rid)
        if cd:
            ws.cell(row=row, column=7).value = cd["decision"]
            ws.cell(row=row, column=7).font = Font("맑은 고딕", size=9, color="666666")
            updated_claude += 1

        # AI 합의 재계산
        codex = (ws.cell(row=row, column=5).value or "").strip()
        gemini = (ws.cell(row=row, column=6).value or "").strip()
        claude_val = (ws.cell(row=row, column=7).value or "").strip()
        new_consensus = compute_3model_consensus(codex, gemini, claude_val)
        old_consensus = ws.cell(row=row, column=8).value or ""
        if new_consensus != old_consensus:
            ws.cell(row=row, column=8).value = new_consensus
            ws.cell(row=row, column=8).font = Font("맑은 고딕", size=9, color="666666")
            updated_consensus += 1

        # 블록/할당 컬럼
        if rid in bibd:
            b = bibd[rid]
            ws.cell(row=row, column=19).value = b["block"]
            ws.cell(row=row, column=20).value = "O" if b["R1"] else ""
            ws.cell(row=row, column=21).value = "O" if b["R2"] else ""
            ws.cell(row=row, column=22).value = "O" if b["R3"] else ""
            for col in range(19, 23):
                ws.cell(row=row, column=col).font = Font("맑은 고딕", size=9)
                ws.cell(row=row, column=col).border = THIN
                ws.cell(row=row, column=col).alignment = Alignment(horizontal="center")
        else:
            # consensus != include → AI 단계에서 제외, 블록 할당 없음
            ws.cell(row=row, column=19).value = "-"
            ws.cell(row=row, column=20).value = ""
            ws.cell(row=row, column=21).value = ""
            ws.cell(row=row, column=22).value = ""
            for col in range(19, 23):
                ws.cell(row=row, column=col).font = Font("맑은 고딕", size=9, color="CCCCCC")
                ws.cell(row=row, column=col).border = THIN
                ws.cell(row=row, column=col).alignment = Alignment(horizontal="center")

    # 자동 필터 업데이트
    ws.auto_filter.ref = f"A1:V{ws.max_row}"

    wb.save(EXCEL_PATH)
    print(f"  Claude 업데이트: {updated_claude}건")
    print(f"  AI 합의 업데이트: {updated_consensus}건")
    print(f"  BIBD 블록 할당: {len(bibd)}건")
    print(f"\nDone. 파일: {EXCEL_PATH}")


if __name__ == "__main__":
    main()
