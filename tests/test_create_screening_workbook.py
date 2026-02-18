"""
Tests for scripts/screening/create_screening_workbook.py
"""

import sys
import pytest
import pandas as pd
import openpyxl
from pathlib import Path

sys.path.insert(0, str(Path(__file__).parent.parent / "scripts" / "screening"))

from create_screening_workbook import main as create_screening_main

REQUIRED_SHEETS = {"SCREENING", "CODEBOOK", "EXCLUSION_LOG", "SEARCH_LOG"}

EXPECTED_SCREENING_COLUMNS = [
    "record_id", "title", "authors", "year", "journal", "doi",
    "abstract", "keywords", "source_database",
    "screen_decision_codex", "screen_confidence_codex",
    "exclude_code_codex", "rationale_codex",
    "screen_decision_gemini", "screen_confidence_gemini",
    "exclude_code_gemini", "rationale_gemini",
    "screen_consensus",
    "human1_decision", "human2_decision",
    "adjudicated_final_decision", "exclude_code",
    "decision_rationale", "adjudicator_id",
]

SEARCH_LOG_DBS = ["Web of Science", "Scopus", "PsycINFO", "IEEE Xplore"]


def _make_input_csv(tmp_path, n=10):
    records = []
    for i in range(1, n + 1):
        db = SEARCH_LOG_DBS[i % 4]
        records.append({
            "record_id": i,
            "title": f"Study {i} on AI adoption",
            "authors": f"Author{i}, A.",
            "year": 2020 + (i % 6),
            "journal": f"Journal {i}",
            "doi": f"10.1234/test{i:04d}" if i <= 8 else "",
            "abstract": f"Abstract {i}",
            "keywords": "AI; education",
            "source_database": db,
        })
    df = pd.DataFrame(records)
    path = tmp_path / "input.csv"
    df.to_csv(path, index=False)
    return path, df


def _build_workbook(tmp_path, n=10):
    input_csv, df = _make_input_csv(tmp_path, n=n)
    xlsx_path = tmp_path / "screening_workbook.xlsx"
    create_screening_main(["--input", str(input_csv), "--output", str(xlsx_path)])
    return xlsx_path, df


# ---------------------------------------------------------------------------
# Sheet presence
# ---------------------------------------------------------------------------

def test_workbook_has_required_sheets(tmp_path):
    xlsx_path, _ = _build_workbook(tmp_path)
    wb = openpyxl.load_workbook(xlsx_path)
    sheet_names_upper = {s.upper() for s in wb.sheetnames}
    for required in REQUIRED_SHEETS:
        assert required in sheet_names_upper, f"Missing sheet: {required}"


# ---------------------------------------------------------------------------
# SCREENING sheet row count
# ---------------------------------------------------------------------------

def test_screening_sheet_row_count(tmp_path):
    n = 10
    xlsx_path, df = _build_workbook(tmp_path, n=n)
    wb = openpyxl.load_workbook(xlsx_path)
    ws = next(ws for ws in wb.worksheets if ws.title.upper() == "SCREENING")
    # Row 1 = header; rows 2..n+1 = data
    data_rows = ws.max_row - 1
    assert data_rows == n


# ---------------------------------------------------------------------------
# SCREENING sheet columns
# ---------------------------------------------------------------------------

def test_screening_sheet_columns(tmp_path):
    xlsx_path, _ = _build_workbook(tmp_path)
    wb = openpyxl.load_workbook(xlsx_path)
    ws = next(ws for ws in wb.worksheets if ws.title.upper() == "SCREENING")
    header = [cell.value for cell in ws[1] if cell.value is not None]
    for col in EXPECTED_SCREENING_COLUMNS:
        assert col in header, f"Expected column missing from SCREENING sheet: {col}"


# ---------------------------------------------------------------------------
# Data validation dropdowns on decision columns
# ---------------------------------------------------------------------------

def test_screening_dropdowns_exist(tmp_path):
    xlsx_path, _ = _build_workbook(tmp_path)
    wb = openpyxl.load_workbook(xlsx_path)
    ws = next(ws for ws in wb.worksheets if ws.title.upper() == "SCREENING")
    # Collect all data validation sqref ranges as strings
    dv_ranges = []
    for dv in ws.data_validations.dataValidation:
        dv_ranges.append(str(dv.sqref))
    # There should be at least one data validation (for decision columns)
    assert len(dv_ranges) > 0, "No data validations found on SCREENING sheet"


# ---------------------------------------------------------------------------
# SEARCH_LOG pre-filled rows
# ---------------------------------------------------------------------------

def test_search_log_prefilled(tmp_path):
    xlsx_path, _ = _build_workbook(tmp_path)
    wb = openpyxl.load_workbook(xlsx_path)
    ws = next(ws for ws in wb.worksheets if ws.title.upper() == "SEARCH_LOG")
    # Collect non-header cell values from first column
    db_values = [ws.cell(row=r, column=1).value for r in range(2, ws.max_row + 1)]
    db_values = [v for v in db_values if v]
    for db in SEARCH_LOG_DBS:
        assert db in db_values, f"Search log missing pre-filled entry for: {db}"


# ---------------------------------------------------------------------------
# CODEBOOK populated
# ---------------------------------------------------------------------------

def test_codebook_populated(tmp_path):
    xlsx_path, _ = _build_workbook(tmp_path)
    wb = openpyxl.load_workbook(xlsx_path)
    ws = next(ws for ws in wb.worksheets if ws.title.upper() == "CODEBOOK")
    # Should have at least a header row plus some content rows
    assert ws.max_row >= 2, "CODEBOOK sheet appears empty"
    # Check that screening variables are represented
    all_text = " ".join(
        str(ws.cell(row=r, column=c).value or "")
        for r in range(1, ws.max_row + 1)
        for c in range(1, ws.max_column + 1)
    )
    for variable in ["screen_decision", "exclude_code", "adjudicated_final_decision"]:
        assert variable in all_text, f"CODEBOOK missing entry for: {variable}"
