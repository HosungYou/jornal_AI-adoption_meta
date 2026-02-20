#!/usr/bin/env python3
"""
Create AI Adoption MASEM Coding Template Excel file.

This workbook operationalizes the canonical coding rules in:
docs/03_data_extraction/AI_Adoption_MASEM_Coding_Manual_v1.docx
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter
from datetime import datetime

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------
CONSTRUCTS = ["PE", "EE", "SI", "FC", "BI", "UB", "ATT", "SE", "TRU", "ANX", "TRA", "AUT"]
DATA_ROWS = 1000
HEADER_FILL = PatternFill(start_color="D6EAF8", end_color="D6EAF8", fill_type="solid")
HEADER_FONT = Font(bold=True, size=11)
HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)
THIN_BORDER = Border(
    bottom=Side(style="thin", color="B0B0B0"),
)

SHEET_TAB_COLORS = {
    "CODEBOOK": "1F77B4",           # blue
    "STUDY_METADATA": "2CA02C",     # green
    "CORRELATION_MATRIX": "D62728", # red
    "CONSTRUCT_MAPPING": "FF7F0E",  # orange
    "MODERATOR_VARIABLES": "9467BD",# purple
    "AI_EXTRACTION_PROVENANCE": "8C564B",  # brown
    "DISCREPANCY_LOG": "E377C2",    # pink
    "EXCLUSION_LOG": "7F7F7F",      # gray
    "SEARCH_LOG": "17BECF",         # cyan
}

# ---------------------------------------------------------------------------
# Sheet definitions: (col_name, width, type_str, valid_values, coding_rules, example)
# ---------------------------------------------------------------------------

STUDY_METADATA_COLS = [
    ("study_id", 10, "integer", "Unique sequential ID", "Assign sequentially starting from 1", "1"),
    ("first_author", 20, "string", "", "Last name of first author", "Smith"),
    ("year", 10, "integer", "2015-2025", "Publication year", "2023"),
    ("title", 45, "string", "", "Full title of the paper", "AI Adoption in Healthcare..."),
    ("doi", 30, "string", "", "Digital Object Identifier", "10.1000/xyz123"),
    ("source_type", 15, "list", "journal,conference", "Type of publication outlet", "journal"),
    ("search_source", 18, "list", "WoS,Scopus,PsycINFO,IEEE,ACM,ERIC,Education_Source,GoogleScholar,citation_tracking", "Database where study was found", "ERIC"),
    ("country", 18, "string", "", "Country where data was collected", "South Korea"),
    ("culture_type", 15, "list", "individualist,collectivist", "Based on Hofstede classification", "collectivist"),
    ("education_level", 18, "list", "K-12,undergraduate,graduate,mixed", "Education level of sample", "undergraduate"),
    ("user_role", 15, "list", "student,instructor,administrator,mixed", "Role of respondents in education", "student"),
    ("discipline", 18, "list", "STEM,humanities,social_science,health_science,mixed", "Academic discipline", "social_science"),
    ("ai_tool_type", 20, "list", "chatbot_LLM,ITS,LMS_AI,auto_grading,writing_assistant,adaptive_learning,general", "Type of educational AI tool", "chatbot_LLM"),
    ("ai_tool_name", 20, "string", "", "Specific AI tool if named", "ChatGPT"),
    ("institutional_type", 18, "list", "public,private,online,community_college,mixed", "Type of educational institution", "public"),
    ("sample_size", 12, "integer", "Positive integer", "Total valid sample size (N)", "384"),
    ("sample_type", 18, "list", "students,instructors,administrators,mixed", "Type of respondents", "students"),
    ("mandatory_voluntary", 18, "list", "mandatory,voluntary,mixed", "Whether AI use is mandatory or voluntary", "voluntary"),
    ("theoretical_framework", 20, "list", "TAM,UTAUT,UTAUT2,TAM_AI,UTAUT_AI,other", "Primary theoretical framework", "UTAUT2"),
    ("study_design", 18, "list", "cross_sectional,longitudinal,experimental", "Research design", "cross_sectional"),
    ("data_collection", 15, "list", "survey,interview,mixed", "Data collection method", "survey"),
    ("post_chatgpt", 12, "list", "yes,no", "Published 2023 or later (post-ChatGPT era)", "yes"),
    ("country_development", 18, "list", "developed,developing", "Country development status (World Bank)", "developing"),
    ("hofstede_individualism", 20, "float", "0-100", "Hofstede individualism score for the country", "18"),
    ("n_constructs_measured", 20, "integer", "Positive integer", "Number of our 12 constructs measured in study", "6"),
    ("measures_BI", 14, "list", "yes,no", "Whether Behavioral Intention (BI) is measured", "yes"),
    ("measures_UB", 14, "list", "yes,no", "Whether Use Behavior (UB) is measured", "no"),
    ("adoption_dv_type", 22, "list", "BI_only,UB_only,BI_and_UB,readiness_as_BI,other", "Type of adoption/use outcome variable", "BI_only"),
    ("n_correlations_reported", 22, "integer", "Positive integer", "Number of pairwise correlations reported", "15"),
    ("matrix_completeness", 20, "float", "0-1", "Proportion of possible pairs reported (out of C(k,2))", "0.75"),
    ("reporting_quality", 18, "list", "1,2,3,4,5", "1=poor to 5=excellent reporting quality", "4"),
    ("sample_adequacy", 18, "list", "adequate,marginal,inadequate", "Based on N relative to parameters", "adequate"),
    ("construct_validity", 18, "list", "good,acceptable,poor", "Overall measurement quality", "good"),
    ("common_method_bias", 20, "list", "addressed,not_addressed,partial", "Whether CMB was tested/addressed", "addressed"),
    ("overall_quality", 15, "list", "high,medium,low", "Overall study quality assessment", "high"),
    ("rob_notes", 35, "string", "", "Risk of bias notes", "Harman single factor test reported"),
    ("screen_decision_codex", 20, "list", "include,exclude,uncertain", "Codex CLI screening decision", "include"),
    ("screen_decision_gemini", 20, "list", "include,exclude,uncertain", "Gemini CLI screening decision", "include"),
    ("screen_consensus", 18, "list", "include,exclude,conflict", "Consensus result between Codex and Gemini", "conflict"),
    ("human1_decision", 16, "list", "include,exclude,uncertain", "First human screener decision", "include"),
    ("human2_decision", 16, "list", "include,exclude,uncertain", "Second human screener decision", "exclude"),
    ("adjudicated_final_decision", 24, "list", "include,exclude", "Final adjudicated decision", "include"),
    ("exclude_code", 16, "list", "E1,E2,E3,E4,E5,E6,E7,E8,E9,E10,E11,E12", "Exclusion code from coding manual", "E4"),
    ("decision_rationale", 35, "string", "", "Rationale for final inclusion/exclusion decision", "Measured AI awareness only; no adoption/use DV"),
    ("adjudicator_id", 15, "string", "", "PI/adjudicator initials for resolved conflicts", "HY"),
    ("screen_run_id", 18, "string", "", "Batch run identifier for screening pipeline", "scr_20260218_b01"),
    ("oauth_auth_method_codex", 22, "list", "oauth,api_key,unknown", "Authentication method used for Codex run", "oauth"),
    ("oauth_auth_method_gemini", 22, "list", "oauth,api_key,unknown", "Authentication method used for Gemini run", "oauth"),
    ("human_coder", 15, "string", "", "Initials of human coder", "HK"),
    ("coding_date", 14, "date", "YYYY-MM-DD", "Date of coding", "2025-01-15"),
    ("coding_notes", 30, "string", "", "Any additional coding notes", "Matrix extracted from Table 4"),
]

CORRELATION_MATRIX_COLS = [
    ("study_id", 10, "integer", "FK to STUDY_METADATA", "Must match a study_id in STUDY_METADATA", "1"),
    ("row_construct", 15, "list", ",".join(CONSTRUCTS), "Row construct of the correlation pair", "PE"),
    ("col_construct", 15, "list", ",".join(CONSTRUCTS), "Column construct (alphabetically after row)", "BI"),
    ("r", 10, "float", "-1 to 1", "Pearson correlation coefficient", "0.45"),
    ("r_source", 15, "list", "reported,beta_converted,calculated", "How the correlation was obtained", "reported"),
    ("original_beta", 14, "float", "", "Original beta if r was converted from beta", "0.38"),
    ("p_value", 10, "float", "0-1", "Reported p-value", "0.001"),
    ("n_for_pair", 12, "integer", "Positive integer", "Sample size for this specific pair", "384"),
    ("source_table", 15, "string", "", "Table/figure reference in the paper", "Table 3"),
    ("page_number", 12, "integer", "", "Page number in the paper", "12"),
    ("notes", 30, "string", "", "Any notes about this correlation", "Bootstrapped CI reported"),
]

CONSTRUCT_MAPPING_COLS = [
    ("study_id", 10, "integer", "FK to STUDY_METADATA", "Must match a study_id in STUDY_METADATA", "1"),
    ("original_name", 25, "string", "", "Construct name as used in the original paper", "Performance Expectancy"),
    ("mapped_to", 12, "list", ",".join(CONSTRUCTS), "Standard construct code mapped to", "PE"),
    ("mapping_confidence", 18, "list", "exact,high,moderate,low", "Confidence in the mapping", "exact"),
    ("mapping_rationale", 35, "string", "", "Why this mapping was chosen", "Identical UTAUT construct"),
    ("cronbach_alpha", 15, "float", "0-1", "Reported Cronbach's alpha", "0.92"),
    ("composite_reliability", 20, "float", "0-1", "Composite reliability (CR)", "0.94"),
    ("ave", 10, "float", "0-1", "Average Variance Extracted", "0.72"),
    ("num_items", 12, "integer", "Positive integer", "Number of measurement items", "4"),
    ("scale_source", 25, "string", "", "Source of the measurement scale", "Venkatesh et al. (2003)"),
    ("is_readiness_construct", 22, "list", "yes,no", "Whether original construct was labeled as 'readiness'", "no"),
    ("readiness_mapping_note", 35, "string", "", "If readiness: justification for mapping to target construct (e.g., BI or SE)", "Items measure intention; mapped to BI"),
]

MODERATOR_VARIABLES_COLS = [
    ("study_id", 10, "integer", "FK to STUDY_METADATA", "Must match a study_id in STUDY_METADATA", "1"),
    ("education_level", 18, "list", "K-12,undergraduate,graduate,mixed", "Education level of sample", "undergraduate"),
    ("user_role", 15, "list", "student,instructor,administrator,mixed", "Role in education", "student"),
    ("discipline", 18, "list", "STEM,humanities,social_science,health_science,mixed", "Academic discipline", "STEM"),
    ("ai_tool_type", 20, "list", "chatbot_LLM,ITS,LMS_AI,auto_grading,writing_assistant,adaptive_learning,general", "Type of educational AI tool", "chatbot_LLM"),
    ("institutional_type", 18, "list", "public,private,online,community_college,mixed", "Institution type", "public"),
    ("culture_dimension", 18, "float", "0-100", "Hofstede individualism score", "91"),
    ("pub_year", 10, "integer", "2015-2025", "Publication year", "2024"),
    ("post_chatgpt", 12, "list", "yes,no", "Published 2023+ (post-ChatGPT era)", "yes"),
    ("country_development", 20, "list", "developed,developing", "Country development status", "developed"),
]

AI_EXTRACTION_PROVENANCE_COLS = [
    ("study_id", 10, "integer", "FK to STUDY_METADATA", "Must match a study_id in STUDY_METADATA", "1"),
    ("variable_name", 20, "string", "", "Name of the variable being extracted", "r_PE_BI"),
    ("field_type", 18, "list", "study_metadata,correlation,construct_mapping,moderator", "Which sheet this variable belongs to", "correlation"),
    ("claude_value", 20, "string", "", "Value extracted by Claude", "0.45"),
    ("claude_confidence", 18, "float", "0-1", "Claude's confidence in extraction", "0.95"),
    ("claude_raw_output", 40, "string", "", "Raw Claude output for audit trail", "Found in Table 3: r=0.45"),
    ("claude_prompt_version", 20, "string", "", "Version of the Claude prompt used", "v2.1"),
    ("gpt4o_value", 20, "string", "", "Value extracted by GPT-4o", "0.45"),
    ("gpt4o_confidence", 18, "float", "0-1", "GPT-4o's confidence in extraction", "0.92"),
    ("gpt4o_raw_output", 40, "string", "", "Raw GPT-4o output for audit trail", "Table 3 reports r=0.45"),
    ("gpt4o_prompt_version", 20, "string", "", "Version of the GPT-4o prompt used", "v2.1"),
    ("groq_value", 20, "string", "", "Value extracted by Groq", "0.45"),
    ("groq_confidence", 18, "float", "0-1", "Groq's confidence in extraction", "0.88"),
    ("groq_raw_output", 40, "string", "", "Raw Groq output for audit trail", "Correlation is 0.45"),
    ("groq_prompt_version", 20, "string", "", "Version of the Groq prompt used", "v2.1"),
    ("ai_consensus_value", 20, "string", "", "Consensus value across AI coders", "0.45"),
    ("ai_consensus_method", 20, "list", "majority,mean,weighted", "Method used to determine consensus", "majority"),
    ("human_value", 20, "string", "", "Value coded by human coder", "0.45"),
    ("human_coder", 15, "string", "", "Initials of human coder", "HK"),
    ("ai_human_match", 18, "list", "exact,within_threshold,mismatch", "Agreement between AI consensus and human", "exact"),
]

DISCREPANCY_LOG_COLS = [
    ("study_id", 10, "integer", "", "Study identifier", "1"),
    ("pair_id", 12, "string", "", "Construct pair identifier", "PE-BI"),
    ("variable_name", 20, "string", "", "Variable where discrepancy occurred", "r"),
    ("coder1_value", 15, "string", "", "Value from first coder", "0.45"),
    ("coder2_value", 15, "string", "", "Value from second coder", "0.43"),
    ("discrepancy_type", 18, "list", "human_human,ai_human,ai_ai", "Type of discrepancy", "ai_human"),
    ("magnitude", 12, "list", "minor,major", "Severity of the discrepancy", "minor"),
    ("final_value", 15, "string", "", "Resolved final value", "0.45"),
    ("resolution_method", 20, "list", "discussion,third_reviewer,source_recheck,calculation_error", "How the discrepancy was resolved", "source_recheck"),
    ("resolved_by", 15, "string", "", "Who resolved the discrepancy", "HK"),
    ("resolution_date", 14, "date", "YYYY-MM-DD", "Date of resolution", "2025-02-01"),
]

EXCLUSION_LOG_COLS = [
    ("study_id", 10, "integer", "", "Study identifier (may not be in STUDY_METADATA)", "501"),
    ("first_author", 20, "string", "", "Last name of first author", "Jones"),
    ("year", 10, "integer", "2015-2025", "Publication year", "2022"),
    ("title", 45, "string", "", "Full title of the paper", "A Qualitative Study of AI..."),
    ("exclusion_stage", 18, "list", "title_abstract,full_text,post_coding", "Stage at which study was excluded", "full_text"),
    ("exclusion_reason", 25, "list", "no_correlation_data,no_AI_focus,qualitative_only,duplicate_sample,insufficient_constructs,non_english,sample_too_small,not_peer_reviewed,non_educational_context", "Primary reason for exclusion", "qualitative_only"),
    ("exclude_code", 12, "list", "E1,E2,E3,E4,E5,E6,E7,E8,E9,E10,E11,E12", "Canonical exclusion code from coding manual", "E4"),
    ("detailed_rationale", 40, "string", "", "Detailed explanation of exclusion", "Study uses only interviews, no quant data"),
    ("screen_decision_codex", 18, "list", "include,exclude,uncertain,NA", "Codex screening decision", "exclude"),
    ("screen_decision_gemini", 18, "list", "include,exclude,uncertain,NA", "Gemini screening decision", "exclude"),
    ("human1_decision", 18, "list", "include,exclude,uncertain,NA", "First human screener decision", "exclude"),
    ("human2_decision", 18, "list", "include,exclude,uncertain,NA", "Second human screener decision", "exclude"),
    ("screener1_decision", 18, "list", "include,exclude,uncertain", "First screener's decision", "exclude"),
    ("screener2_decision", 18, "list", "include,exclude,uncertain", "Second screener's decision", "exclude"),
    ("adjudicator_id", 15, "string", "", "PI/adjudicator initials", "HY"),
    ("final_decision", 14, "list", "include,exclude", "Final inclusion/exclusion decision", "exclude"),
]

SEARCH_LOG_COLS = [
    ("database", 18, "list", "WoS,Scopus,PsycINFO,IEEE,ACM,ERIC,Education_Source,GoogleScholar", "Database searched", "ERIC"),
    ("search_date", 14, "date", "YYYY-MM-DD", "Date of the search", "2025-01-10"),
    ("search_string", 60, "string", "", "Full search query used", '("artificial intelligence" OR "AI") AND ("adoption" OR "acceptance")'),
    ("results_count", 14, "integer", "Non-negative integer", "Number of results returned", "1247"),
    ("notes", 35, "string", "", "Notes about this search", "Limited to 2015-2025"),
]

# Collect all sheet definitions in order
ALL_SHEETS = {
    "STUDY_METADATA": STUDY_METADATA_COLS,
    "CORRELATION_MATRIX": CORRELATION_MATRIX_COLS,
    "CONSTRUCT_MAPPING": CONSTRUCT_MAPPING_COLS,
    "MODERATOR_VARIABLES": MODERATOR_VARIABLES_COLS,
    "AI_EXTRACTION_PROVENANCE": AI_EXTRACTION_PROVENANCE_COLS,
    "DISCREPANCY_LOG": DISCREPANCY_LOG_COLS,
    "EXCLUSION_LOG": EXCLUSION_LOG_COLS,
    "SEARCH_LOG": SEARCH_LOG_COLS,
}

CODEBOOK_COLS = [
    ("variable_name", 25),
    ("sheet", 28),
    ("type", 12),
    ("valid_values", 55),
    ("coding_rules", 50),
    ("example", 30),
    ("notes", 40),
]


def build_codebook_rows():
    """Generate codebook rows for every variable in every sheet."""
    rows = []
    for sheet_name, col_defs in ALL_SHEETS.items():
        for col_def in col_defs:
            name, width, dtype, valid_vals, coding_rules, example = col_def
            # Build notes based on context
            notes = ""
            if "FK" in str(valid_vals):
                notes = "Foreign key reference"
            elif dtype == "list":
                notes = "Dropdown list validation"
            elif dtype == "date":
                notes = "Date format: YYYY-MM-DD"
            elif dtype == "float" and "0-1" in str(valid_vals):
                notes = "Proportion/probability"
            elif dtype == "float" and "-1 to 1" in str(valid_vals):
                notes = "Correlation coefficient"
            elif name == "study_id" and sheet_name == "STUDY_METADATA":
                notes = "Primary key"

            # Clean up type for display
            display_type = dtype
            if dtype == "list":
                display_type = "categorical"

            rows.append([
                name,
                sheet_name,
                display_type,
                valid_vals if dtype != "list" else valid_vals,
                coding_rules,
                example,
                notes,
            ])
    return rows


def apply_header_format(ws, num_cols):
    """Apply formatting to header row."""
    for col_idx in range(1, num_cols + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER


def add_data_validations(ws, col_defs, data_rows):
    """Add dropdown data validations for list-type columns."""
    for col_idx, col_def in enumerate(col_defs, start=1):
        name, width, dtype = col_def[0], col_def[1], col_def[2]
        if dtype == "list":
            valid_values = col_def[3]
            col_letter = get_column_letter(col_idx)
            formula = f'"{valid_values}"'
            dv = DataValidation(
                type="list",
                formula1=formula,
                allow_blank=True,
                showDropDown=False,
                showErrorMessage=True,
                errorTitle="Invalid Entry",
                error=f"Please select from the dropdown list for {name}.",
                promptTitle=name,
                prompt=f"Select a value: {valid_values}",
                showInputMessage=True,
            )
            dv.sqref = f"{col_letter}2:{col_letter}{data_rows + 1}"
            ws.add_data_validation(dv)


def create_data_sheet(wb, sheet_name, col_defs, data_rows=DATA_ROWS):
    """Create a data sheet with headers, formatting, validations, and pre-allocated rows."""
    ws = wb.create_sheet(title=sheet_name)

    # Write headers
    for col_idx, col_def in enumerate(col_defs, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_def[0])

    # Set column widths
    for col_idx, col_def in enumerate(col_defs, start=1):
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = col_def[1]

    # Apply header formatting
    apply_header_format(ws, len(col_defs))

    # Add auto-filter
    last_col_letter = get_column_letter(len(col_defs))
    ws.auto_filter.ref = f"A1:{last_col_letter}{data_rows + 1}"

    # Freeze header row
    ws.freeze_panes = "A2"

    # Add data validations
    add_data_validations(ws, col_defs, data_rows)

    # Set tab color
    if sheet_name in SHEET_TAB_COLORS:
        ws.sheet_properties.tabColor = SHEET_TAB_COLORS[sheet_name]

    return ws


def create_codebook_sheet(wb):
    """Create the CODEBOOK sheet with all variable definitions pre-filled."""
    ws = wb.create_sheet(title="CODEBOOK")

    # Write headers
    for col_idx, (name, width) in enumerate(CODEBOOK_COLS, start=1):
        ws.cell(row=1, column=col_idx, value=name)
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = width

    # Apply header formatting
    apply_header_format(ws, len(CODEBOOK_COLS))

    # Write codebook rows
    codebook_rows = build_codebook_rows()
    for row_idx, row_data in enumerate(codebook_rows, start=2):
        for col_idx, value in enumerate(row_data, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    # Add alternating row shading for readability within each sheet group
    light_gray = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    current_sheet = None
    shade = False
    for row_idx, row_data in enumerate(codebook_rows, start=2):
        sheet_name = row_data[1]
        if sheet_name != current_sheet:
            current_sheet = sheet_name
            shade = not shade
        if shade:
            for col_idx in range(1, len(CODEBOOK_COLS) + 1):
                ws.cell(row=row_idx, column=col_idx).fill = light_gray

    # Auto-filter
    last_col_letter = get_column_letter(len(CODEBOOK_COLS))
    last_row = len(codebook_rows) + 1
    ws.auto_filter.ref = f"A1:{last_col_letter}{last_row}"

    # Freeze header
    ws.freeze_panes = "A2"

    # Tab color
    ws.sheet_properties.tabColor = SHEET_TAB_COLORS["CODEBOOK"]

    return ws


def main():
    wb = openpyxl.Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    # 1. CODEBOOK (first sheet)
    print("Creating CODEBOOK sheet...")
    create_codebook_sheet(wb)

    # 2-9. Data sheets
    for sheet_name, col_defs in ALL_SHEETS.items():
        print(f"Creating {sheet_name} sheet...")
        create_data_sheet(wb, sheet_name, col_defs)

    # Save
    output_path = "/Users/hosung/jornal_AI-adoption_meta/data/templates/AI_Adoption_MASEM_Coding_v1.xlsx"
    wb.save(output_path)
    print(f"\nSaved to: {output_path}")

    # Summary
    print("\n--- Template Summary ---")
    for ws in wb.worksheets:
        print(f"  {ws.title}: {ws.max_column} columns, tab color={ws.sheet_properties.tabColor}")
    codebook_rows = build_codebook_rows()
    print(f"  CODEBOOK entries: {len(codebook_rows)} variables documented")
    print(f"  Data rows pre-allocated: {DATA_ROWS} per sheet")
    print(f"  Constructs: {', '.join(CONSTRUCTS)} ({len(CONSTRUCTS)} total)")
    n_pairs = len(CONSTRUCTS) * (len(CONSTRUCTS) - 1) // 2
    print(f"  Pairwise correlations: {n_pairs}")


if __name__ == "__main__":
    main()
