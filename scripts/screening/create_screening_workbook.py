"""
create_screening_workbook.py
-----------------------------
Generates the 1st screening XLSX workbook (Screening Workbook) from a
screening master CSV.

Sheets produced
---------------
1. SCREENING      – one row per record, screening decision columns with dropdowns
2. CODEBOOK       – variable dictionary for all screening columns
3. EXCLUSION_LOG  – pre-allocated log rows with dropdowns
4. SEARCH_LOG     – pre-filled database search provenance rows

Usage
-----
    python create_screening_workbook.py \
        --input  data/02_processed/screening_master_16189.csv \
        --output data/templates/AI_Adoption_Screening_v1.xlsx
"""

import argparse
import os
import sys
from datetime import date

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# ---------------------------------------------------------------------------
# Style constants
# ---------------------------------------------------------------------------

HEADER_FILL = PatternFill(start_color="D6EAF8", end_color="D6EAF8", fill_type="solid")
HEADER_FONT = Font(bold=True, size=11)
HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)
THIN_BORDER = Border(bottom=Side(style="thin", color="B0B0B0"))

SHEET_TAB_COLORS = {
    "SCREENING": "2CA02C",
    "CODEBOOK": "1F77B4",
    "EXCLUSION_LOG": "7F7F7F",
    "SEARCH_LOG": "17BECF",
}

# ---------------------------------------------------------------------------
# Column definitions
# ---------------------------------------------------------------------------

# Source columns from CSV
SOURCE_COLS = [
    "record_id",
    "title",
    "abstract",
    "authors",
    "year",
    "journal",
    "doi",
    "source_database",
    "keywords",
    "language",
]

# Screening decision columns with their dropdown option strings
SCREENING_COLS_DROPDOWNS = {
    "screen_decision_codex":       "include,exclude,uncertain",
    "screen_confidence_codex":     None,
    "exclude_code_codex":          "E1,E2,E3,E4,E5,E6,E7,E8,E9,E10,E11,E12,NA",
    "rationale_codex":             None,
    "screen_decision_gemini":      "include,exclude,uncertain",
    "screen_confidence_gemini":    None,
    "exclude_code_gemini":         "E1,E2,E3,E4,E5,E6,E7,E8,E9,E10,E11,E12,NA",
    "rationale_gemini":            None,
    "screen_consensus":            "include,exclude,conflict",
    "human1_decision":             "include,exclude,uncertain",
    "human2_decision":             "include,exclude,uncertain",
    "adjudicated_final_decision":  "include,exclude",
    "exclude_code":                "E1,E2,E3,E4,E5,E6,E7,E8,E9,E10,E11,E12",
    "decision_rationale":          None,
    "adjudicator_id":              None,
    "screen_run_id":               None,
    "oauth_auth_method_codex":     "oauth,api_key,unknown",
    "oauth_auth_method_gemini":    "oauth,api_key,unknown",
}

# Column widths for SCREENING sheet (header -> width)
SCREENING_COL_WIDTHS = {
    "record_id":                   14,
    "title":                       45,
    "abstract":                    60,
    "authors":                     25,
    "year":                        8,
    "journal":                     30,
    "doi":                         30,
    "source_database":             18,
    "keywords":                    30,
    "language":                    12,
    "screen_decision_codex":       20,
    "screen_confidence_codex":     22,
    "exclude_code_codex":          18,
    "rationale_codex":             35,
    "screen_decision_gemini":      22,
    "screen_confidence_gemini":    24,
    "exclude_code_gemini":         20,
    "rationale_gemini":            35,
    "screen_consensus":            18,
    "human1_decision":             18,
    "human2_decision":             18,
    "adjudicated_final_decision":  24,
    "exclude_code":                14,
    "decision_rationale":          35,
    "adjudicator_id":              16,
    "screen_run_id":               16,
    "oauth_auth_method_codex":     22,
    "oauth_auth_method_gemini":    24,
}

# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _apply_header_style(cell, *, border=True):
    """Apply standard header styling to a single cell."""
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell.alignment = HEADER_ALIGNMENT
    if border:
        cell.border = THIN_BORDER


def _set_tab_color(ws, sheet_name: str):
    color = SHEET_TAB_COLORS.get(sheet_name)
    if color:
        ws.sheet_properties.tabColor = color


def _make_dropdown(ws, col_letter: str, formula: str, first_row: int, last_row: int):
    """Add a list DataValidation to a column range."""
    dv = DataValidation(
        type="list",
        formula1=f'"{formula}"',
        showDropDown=False,
        allow_blank=True,
    )
    dv.sqref = f"{col_letter}{first_row}:{col_letter}{last_row}"
    ws.add_data_validation(dv)


# ---------------------------------------------------------------------------
# Sheet builders
# ---------------------------------------------------------------------------

def build_screening_sheet(wb: Workbook, df: pd.DataFrame) -> int:
    """Build the SCREENING sheet. Returns the number of data rows written."""
    ws = wb.create_sheet("SCREENING")
    _set_tab_color(ws, "SCREENING")
    ws.sheet_properties.tabColor = SHEET_TAB_COLORS["SCREENING"]

    all_cols = SOURCE_COLS + list(SCREENING_COLS_DROPDOWNS.keys())

    # --- Header row ---
    for col_idx, col_name in enumerate(all_cols, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        _apply_header_style(cell)

    # Set column widths
    for col_idx, col_name in enumerate(all_cols, start=1):
        width = SCREENING_COL_WIDTHS.get(col_name, 16)
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # --- Write data rows ---
    n_rows = len(df)
    for row_idx, (_, row) in enumerate(df.iterrows(), start=2):
        for col_idx, col_name in enumerate(all_cols, start=1):
            if col_name in SOURCE_COLS:
                value = row.get(col_name, "")
                # Convert NaN to empty string
                if pd.isna(value):
                    value = ""
                ws.cell(row=row_idx, column=col_idx, value=value)

    # --- Dropdown validations ---
    if n_rows > 0:
        first_data = 2
        last_data = n_rows + 1
        for col_idx, col_name in enumerate(all_cols, start=1):
            formula = SCREENING_COLS_DROPDOWNS.get(col_name)
            if formula:
                col_letter = get_column_letter(col_idx)
                _make_dropdown(ws, col_letter, formula, first_data, last_data)

    # --- Auto-filter and freeze panes ---
    last_col_letter = get_column_letter(len(all_cols))
    ws.auto_filter.ref = f"A1:{last_col_letter}1"
    ws.freeze_panes = "A2"

    # Row height for header
    ws.row_dimensions[1].height = 30

    return n_rows


# ---------------------------------------------------------------------------

CODEBOOK_ROWS = [
    # variable_name, sheet, type, valid_values, coding_rules, example
    ("record_id",                  "SCREENING", "string",   "",                                              "Unique identifier from source database",       "WoS_000001"),
    ("title",                      "SCREENING", "string",   "",                                              "Full article title as indexed",               "AI adoption in SMEs"),
    ("abstract",                   "SCREENING", "string",   "",                                              "Full abstract text",                          "This study examines..."),
    ("authors",                    "SCREENING", "string",   "",                                              "Author list as indexed",                      "Smith J, Lee K"),
    ("year",                       "SCREENING", "integer",  "",                                              "Publication year (4 digits)",                 "2023"),
    ("journal",                    "SCREENING", "string",   "",                                              "Journal / venue name",                        "Journal of MIS"),
    ("doi",                        "SCREENING", "string",   "",                                              "Digital Object Identifier",                   "10.1016/j.jmis.2023.01.001"),
    ("source_database",            "SCREENING", "string",   "",                                              "Source database",                             "Web of Science"),
    ("keywords",                   "SCREENING", "string",   "",                                              "Author or database keywords",                 "AI; adoption; SME"),
    ("language",                   "SCREENING", "string",   "",                                              "Publication language",                        "English"),
    ("screen_decision_codex",      "SCREENING", "category", "include,exclude,uncertain",                     "Codex AI screening decision",                 "include"),
    ("screen_confidence_codex",    "SCREENING", "float",    "0.0–1.0",                                       "Codex confidence score (0–1)",                "0.95"),
    ("exclude_code_codex",         "SCREENING", "category", "E1–E12,NA",                                     "Exclusion reason code from Codex",            "E3"),
    ("rationale_codex",            "SCREENING", "string",   "",                                              "Free-text rationale from Codex",              "Not empirical study"),
    ("screen_decision_gemini",     "SCREENING", "category", "include,exclude,uncertain",                     "Gemini AI screening decision",                "exclude"),
    ("screen_confidence_gemini",   "SCREENING", "float",    "0.0–1.0",                                       "Gemini confidence score (0–1)",               "0.88"),
    ("exclude_code_gemini",        "SCREENING", "category", "E1–E12,NA",                                     "Exclusion reason code from Gemini",           "E3"),
    ("rationale_gemini",           "SCREENING", "string",   "",                                              "Free-text rationale from Gemini",             "Review paper, not empirical"),
    ("screen_consensus",           "SCREENING", "category", "include,exclude,conflict",                      "AI consensus: agree→include/exclude; else conflict", "conflict"),
    ("human1_decision",            "SCREENING", "category", "include,exclude,uncertain",                     "Human screener 1 decision",                   "include"),
    ("human2_decision",            "SCREENING", "category", "include,exclude,uncertain",                     "Human screener 2 decision (conflict only)",   "exclude"),
    ("adjudicated_final_decision", "SCREENING", "category", "include,exclude",                               "Final adjudicated decision",                  "include"),
    ("exclude_code",               "SCREENING", "category", "E1,E2,E3,E4,E5,E6,E7,E8,E9,E10,E11,E12",      "Final exclusion reason code (if excluded)",    "E5"),
    ("decision_rationale",         "SCREENING", "string",   "",                                              "Final decision rationale",                    "Meets all inclusion criteria"),
    ("adjudicator_id",             "SCREENING", "string",   "",                                              "ID of adjudicating reviewer",                 "HS"),
    ("screen_run_id",              "SCREENING", "string",   "",                                              "Screening batch/run identifier",              "run_20260217_v1"),
    ("oauth_auth_method_codex",    "SCREENING", "category", "oauth,api_key,unknown",                         "Auth method used for Codex API call",         "api_key"),
    ("oauth_auth_method_gemini",   "SCREENING", "category", "oauth,api_key,unknown",                         "Auth method used for Gemini API call",        "oauth"),
]

CODEBOOK_HEADERS = [
    "variable_name", "sheet", "type", "valid_values",
    "coding_rules", "example",
]

CODEBOOK_COL_WIDTHS = [25, 16, 12, 40, 55, 35]


def build_codebook_sheet(wb: Workbook):
    ws = wb.create_sheet("CODEBOOK")
    _set_tab_color(ws, "CODEBOOK")

    for col_idx, header in enumerate(CODEBOOK_HEADERS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        _apply_header_style(cell)
        ws.column_dimensions[get_column_letter(col_idx)].width = CODEBOOK_COL_WIDTHS[col_idx - 1]

    for row_idx, row_data in enumerate(CODEBOOK_ROWS, start=2):
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    ws.auto_filter.ref = f"A1:{get_column_letter(len(CODEBOOK_HEADERS))}1"
    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 28


# ---------------------------------------------------------------------------

EXCLUSION_LOG_HEADERS = [
    "record_id",
    "first_author",
    "year",
    "title",
    "exclusion_stage",
    "exclusion_reason",
    "exclude_code",
    "detailed_rationale",
    "screen_decision_codex",
    "screen_decision_gemini",
    "human1_decision",
    "human2_decision",
    "adjudicator_id",
    "final_decision",
]

EXCLUSION_LOG_DROPDOWNS = {
    "exclusion_stage":       "title_abstract,full_text,post_extraction",
    "exclude_code":          "E1,E2,E3,E4,E5,E6,E7,E8,E9,E10,E11,E12",
    "screen_decision_codex": "include,exclude,uncertain",
    "screen_decision_gemini":"include,exclude,uncertain",
    "human1_decision":       "include,exclude,uncertain",
    "human2_decision":       "include,exclude,uncertain",
    "final_decision":        "include,exclude",
}

EXCLUSION_LOG_COL_WIDTHS = {
    "record_id":              14,
    "first_author":           20,
    "year":                    8,
    "title":                  45,
    "exclusion_stage":        20,
    "exclusion_reason":       30,
    "exclude_code":           14,
    "detailed_rationale":     40,
    "screen_decision_codex":  22,
    "screen_decision_gemini": 24,
    "human1_decision":        18,
    "human2_decision":        18,
    "adjudicator_id":         16,
    "final_decision":         16,
}


def build_exclusion_log_sheet(wb: Workbook, n_records: int):
    ws = wb.create_sheet("EXCLUSION_LOG")
    _set_tab_color(ws, "EXCLUSION_LOG")

    for col_idx, header in enumerate(EXCLUSION_LOG_HEADERS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        _apply_header_style(cell)
        ws.column_dimensions[get_column_letter(col_idx)].width = \
            EXCLUSION_LOG_COL_WIDTHS.get(header, 16)

    # Pre-allocate rows (empty) matching input record count
    last_row = n_records + 1 if n_records > 0 else 2

    # Dropdowns over all pre-allocated rows
    for col_idx, header in enumerate(EXCLUSION_LOG_HEADERS, start=1):
        formula = EXCLUSION_LOG_DROPDOWNS.get(header)
        if formula:
            col_letter = get_column_letter(col_idx)
            _make_dropdown(ws, col_letter, formula, 2, last_row)

    ws.auto_filter.ref = f"A1:{get_column_letter(len(EXCLUSION_LOG_HEADERS))}1"
    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 28


# ---------------------------------------------------------------------------

SEARCH_LOG_HEADERS = [
    "database",
    "search_date",
    "search_string",
    "results_count",
    "notes",
]

SEARCH_LOG_PREFILL = [
    ("Web of Science",  str(date.today()), "", "", ""),
    ("Scopus",          str(date.today()), "", "", ""),
    ("PsycINFO",        str(date.today()), "", "", ""),
    ("IEEE Xplore",     str(date.today()), "", "", ""),
]

SEARCH_LOG_COL_WIDTHS = [20, 14, 80, 16, 40]


def build_search_log_sheet(wb: Workbook):
    ws = wb.create_sheet("SEARCH_LOG")
    _set_tab_color(ws, "SEARCH_LOG")

    for col_idx, header in enumerate(SEARCH_LOG_HEADERS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        _apply_header_style(cell)
        ws.column_dimensions[get_column_letter(col_idx)].width = \
            SEARCH_LOG_COL_WIDTHS[col_idx - 1]

    for row_idx, row_data in enumerate(SEARCH_LOG_PREFILL, start=2):
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 28


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def parse_args(argv=None):
    parser = argparse.ArgumentParser(
        description="Generate the AI Adoption Screening Workbook (XLSX) from a master CSV."
    )
    parser.add_argument(
        "--input",
        default="data/02_processed/screening_master_16189.csv",
        help="Path to the screening master CSV file.",
    )
    parser.add_argument(
        "--output",
        default="data/templates/AI_Adoption_Screening_v1.xlsx",
        help="Path for the output XLSX workbook.",
    )
    return parser.parse_args(argv)


def main(argv=None):
    args = parse_args(argv)

    input_path = args.input
    output_path = args.output

    # ------------------------------------------------------------------
    # Load CSV
    # ------------------------------------------------------------------
    if not os.path.isfile(input_path):
        print(f"ERROR: Input file not found: {input_path}", file=sys.stderr)
        sys.exit(1)

    print(f"Loading CSV: {input_path}")
    df = pd.read_csv(input_path, dtype=str, low_memory=False)

    # Ensure all expected source columns exist (add as empty if missing)
    for col in SOURCE_COLS:
        if col not in df.columns:
            print(f"  WARNING: Column '{col}' not in CSV – adding as empty.")
            df[col] = ""

    # Keep only the columns we care about (in order)
    df = df[SOURCE_COLS].copy()
    n_records = len(df)
    print(f"  Records loaded: {n_records:,}")

    # ------------------------------------------------------------------
    # Build workbook
    # ------------------------------------------------------------------
    wb = Workbook()
    # Remove default sheet
    default_sheet = wb.active
    wb.remove(default_sheet)

    print("Building SCREENING sheet …")
    n_written = build_screening_sheet(wb, df)

    print("Building CODEBOOK sheet …")
    build_codebook_sheet(wb)

    print("Building EXCLUSION_LOG sheet …")
    build_exclusion_log_sheet(wb, n_records)

    print("Building SEARCH_LOG sheet …")
    build_search_log_sheet(wb)

    # ------------------------------------------------------------------
    # Save
    # ------------------------------------------------------------------
    os.makedirs(os.path.dirname(output_path) if os.path.dirname(output_path) else ".", exist_ok=True)
    wb.save(output_path)

    # ------------------------------------------------------------------
    # Summary
    # ------------------------------------------------------------------
    print("\n" + "=" * 60)
    print("Workbook created successfully.")
    print(f"  Output : {output_path}")
    print(f"  Sheets :")
    for ws in wb.worksheets:
        if ws.title == "SCREENING":
            print(f"    {ws.title:<20}  {n_written:>7,} data rows")
        elif ws.title == "CODEBOOK":
            print(f"    {ws.title:<20}  {len(CODEBOOK_ROWS):>7,} variable rows")
        elif ws.title == "EXCLUSION_LOG":
            print(f"    {ws.title:<20}  {n_records:>7,} pre-allocated rows")
        elif ws.title == "SEARCH_LOG":
            print(f"    {ws.title:<20}  {len(SEARCH_LOG_PREFILL):>7,} pre-filled rows")
    print("=" * 60)


if __name__ == "__main__":
    main()
