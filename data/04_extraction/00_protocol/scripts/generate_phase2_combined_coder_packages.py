#!/usr/bin/env python3
"""Generate Phase 0/1/2 combined coder workbooks for Phase 2 distribution.

The original workbook design keeps Phase 0 calibration, Phase 1, and Phase 2 in
one coder workbook. This script preserves that design while writing the output to
the Phase 2 distribution folder so the frozen Phase 1 raw workbooks are not
overwritten.

Phase 0/1 values are copied from the frozen Phase 1 workbooks. Phase 2 rows use
the rotated-pair allocation:

- R1 + R4: Pair C, 57 studies
- R2 + R3: Pair D, 56 studies

PDFs are not copied. Hyperlinks are stripped from the generated workbooks.
"""

from __future__ import annotations

import shutil
from datetime import date
from pathlib import Path

import openpyxl
from openpyxl.cell.cell import MergedCell

from generate_coder_packages import (
    BASE,
    assign_phases,
    build_assignment_sheet,
    build_codebook_sheet,
    build_correlations_sheet,
    build_discrepancy_log,
    build_exclusion_log,
    build_guide_sheet,
    build_study_metadata_sheet,
    load_studies,
    select_calibration,
)


SOURCE_DIR = BASE / "data/04_extraction/01_raw_human_coder_data_freeze/phase1/coder_packages"
OUTPUT_DIR = BASE / "data/04_extraction/01_raw_human_coder_data_freeze/phase2/coder_packages"
INSTRUCTION_FILE = OUTPUT_DIR / "Phase2_Coder_Instructions_KR.md"
DATE_STAMP = date.today().strftime("%Y%m%d")


def row_values_by_header(ws):
    headers = [cell.value for cell in ws[1]]
    return headers, {name: i + 1 for i, name in enumerate(headers) if name}


def source_prior_ids(source_wb) -> set[str]:
    ws = source_wb["ASSIGNMENT"]
    headers, idx = row_values_by_header(ws)
    study_col = idx["Study ID"]
    phase_col = idx["Phase"]
    ids = set()
    for row in range(2, ws.max_row + 1):
        sid = ws.cell(row=row, column=study_col).value
        phase = ws.cell(row=row, column=phase_col).value
        if isinstance(sid, str) and isinstance(phase, str) and phase.startswith(("Phase 0", "Phase 1")):
            ids.add(sid)
    return ids


def copy_matching_rows(source_ws, target_ws, key_fields, include_ids: set[str]) -> None:
    source_headers, source_idx = row_values_by_header(source_ws)
    target_headers, target_idx = row_values_by_header(target_ws)
    common_headers = [h for h in target_headers if h in source_idx and h in target_idx]

    def key_for(ws, idx, row):
        values = []
        for field in key_fields:
            values.append(ws.cell(row=row, column=idx[field]).value)
        return tuple(values)

    source_rows = {}
    for row in range(2, source_ws.max_row + 1):
        sid = source_ws.cell(row=row, column=source_idx["study_id" if "study_id" in source_idx else "Study ID"]).value
        if sid in include_ids:
            source_rows[key_for(source_ws, source_idx, row)] = row

    for target_row in range(2, target_ws.max_row + 1):
        sid = target_ws.cell(
            row=target_row,
            column=target_idx["study_id" if "study_id" in target_idx else "Study ID"],
        ).value
        if sid not in include_ids:
            continue
        source_row = source_rows.get(key_for(target_ws, target_idx, target_row))
        if not source_row:
            continue
        for header in common_headers:
            target_ws.cell(row=target_row, column=target_idx[header]).value = source_ws.cell(
                row=source_row,
                column=source_idx[header],
            ).value


def copy_log_rows(source_ws, target_ws, include_ids: set[str]) -> None:
    source_headers, source_idx = row_values_by_header(source_ws)
    target_headers, target_idx = row_values_by_header(target_ws)
    common_headers = [h for h in target_headers if h in source_idx and h in target_idx]
    sid_header = "study_id" if "study_id" in source_idx else "Study ID"

    write_row = 2
    for source_row in range(2, source_ws.max_row + 1):
        sid = source_ws.cell(row=source_row, column=source_idx[sid_header]).value
        if sid not in include_ids:
            continue
        if not any(source_ws.cell(row=source_row, column=col).value for col in range(1, source_ws.max_column + 1)):
            continue
        for header in common_headers:
            cell = target_ws.cell(row=write_row, column=target_idx[header])
            source_cell = source_ws.cell(row=source_row, column=source_idx[header])
            cell.value = source_cell.value
        write_row += 1


def strip_hyperlinks(wb) -> None:
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell, MergedCell):
                    continue
                cell.hyperlink = None


def mark_prior_phases_done(wb) -> None:
    ws = wb["ASSIGNMENT"]
    headers, idx = row_values_by_header(ws)
    phase_col = idx["Phase"]
    status_col = idx["Status"]
    for row in range(2, ws.max_row + 1):
        phase = ws.cell(row=row, column=phase_col).value
        if isinstance(phase, str) and phase.startswith(("Phase 0", "Phase 1")):
            ws.cell(row=row, column=status_col).value = "done"


def create_combined_workbook(coder_label: str, calibration, phase1, phase2, pair_label: str, phase2_pair_label: str) -> Path:
    source_path = SOURCE_DIR / coder_label / f"AI_Adoption_MASEM_Coding_v3_{coder_label}.xlsx"
    source_wb = openpyxl.load_workbook(source_path)
    prior_ids = source_prior_ids(source_wb)

    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    coder_studies = calibration + phase1 + phase2

    build_guide_sheet(wb, f"{coder_label} Phase 0/1/2")
    build_assignment_sheet(wb, calibration, phase1, phase2, coder_label, pair_label, phase2_pair_label)
    build_study_metadata_sheet(wb, coder_studies, coder_label)
    build_correlations_sheet(wb, coder_studies)
    build_exclusion_log(wb)
    build_discrepancy_log(wb)
    build_codebook_sheet(wb)

    copy_matching_rows(source_wb["ASSIGNMENT"], wb["ASSIGNMENT"], ["Study ID"], prior_ids)
    copy_matching_rows(source_wb["STUDY_METADATA"], wb["STUDY_METADATA"], ["study_id"], prior_ids)
    copy_matching_rows(source_wb["CORRELATIONS"], wb["CORRELATIONS"], ["study_id", "construct_1", "construct_2"], prior_ids)
    copy_log_rows(source_wb["EXCLUSION_LOG"], wb["EXCLUSION_LOG"], prior_ids)
    copy_log_rows(source_wb["DISCREPANCY_LOG"], wb["DISCREPANCY_LOG"], prior_ids)
    mark_prior_phases_done(wb)
    strip_hyperlinks(wb)

    coder_dir = OUTPUT_DIR / coder_label
    coder_dir.mkdir(parents=True, exist_ok=True)
    output_path = coder_dir / f"AI_Adoption_MASEM_Coding_v3_{coder_label}_Phase0_1_2_{DATE_STAMP}.xlsx"
    wb.save(output_path)
    source_wb.close()
    return output_path


def write_korean_instruction() -> None:
    text = """# AI Adoption MASEM Phase 2 코딩 안내

안녕하세요. Phase 1 코딩과 pairwise comparison workbook 생성이 완료되어 Phase 2 코딩을 시작합니다.

이번에 제공하는 Excel 파일은 기존 구조를 유지하여 Phase 0, Phase 1, Phase 2가 한 파일에 함께 들어 있습니다. Phase 0과 Phase 1은 이전 코딩값을 보존한 참조/기록 영역입니다. 새로 작성해야 하는 부분은 Phase 2 행입니다.

## 배정

- Pair C: R1 + R4, 57개 study
- Pair D: R2 + R3, 56개 study

각 코더는 본인 라벨의 Excel 파일만 사용해 독립적으로 코딩해 주세요. 다른 코더의 값이나 LLM output은 참고하지 않습니다. PDF는 각자 접근 가능한 방식으로 확인하면 되고, Excel 파일에는 PDF hyperlink를 넣지 않았습니다.

## 작업 순서

1. 본인 Excel 파일을 엽니다.
2. ASSIGNMENT 탭에서 `Phase 2` 행만 새로 코딩합니다.
3. 코딩이 끝난 study는 ASSIGNMENT 탭의 Status를 `done`으로 표시합니다.
4. 제외가 필요한 study는 EXCLUSION_LOG에 근거를 적고 Status를 `excluded`로 표시합니다.
5. 판단이 애매한 항목은 flag와 notes에 짧게 남깁니다.
6. 완료 파일명에는 본인 라벨과 날짜를 유지해 제출합니다.

## 제출 파일명 예시

`AI_Adoption_MASEM_Coding_v3_R1_Phase0_1_2_completed_YYYYMMDD.xlsx`

제출 후에는 Phase 2 raw human coder data로 freeze하고, 이후 pairwise disagreement 분석, source-document adjudication, adjudicated human reference standard freeze 순서로 진행합니다. LLM comparison과 MASEM substitution analysis는 reference freeze 이후에만 진행합니다.
"""
    INSTRUCTION_FILE.write_text(text, encoding="utf-8")


def main() -> None:
    if OUTPUT_DIR.exists():
        shutil.rmtree(OUTPUT_DIR)
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    studies = load_studies()
    calibration = select_calibration(studies, n=10)
    phases = assign_phases(studies, {s["study_id"] for s in calibration})

    outputs = [
        create_combined_workbook("R1", calibration, phases["pair_a"], phases["pair_c"], "Pair A (R1+R2)", "Pair C (R1+R4)"),
        create_combined_workbook("R2", calibration, phases["pair_a"], phases["pair_d"], "Pair A (R1+R2)", "Pair D (R2+R3)"),
        create_combined_workbook("R3", calibration, phases["pair_b"], phases["pair_d"], "Pair B (R3+R4)", "Pair D (R2+R3)"),
        create_combined_workbook("R4", calibration, phases["pair_b"], phases["pair_c"], "Pair B (R3+R4)", "Pair C (R1+R4)"),
    ]
    write_korean_instruction()

    print("Phase 0/1/2 combined coder packages generated:")
    for path in outputs:
        print(f"- {path}")
    print(f"- {INSTRUCTION_FILE}")


if __name__ == "__main__":
    main()
