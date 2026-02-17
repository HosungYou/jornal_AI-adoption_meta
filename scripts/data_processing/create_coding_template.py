#!/usr/bin/env python3
"""
Create Excel coding template for manual data extraction.
Uses openpyxl for data validation and formatting.
"""

import pandas as pd
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
import logging

logger = logging.getLogger(__name__)


class CodingTemplateGenerator:
    """Generates structured Excel templates for data coding."""

    def __init__(self):
        """Initialize template generator."""
        # Define construct list for dropdowns
        self.constructs = [
            'PE', 'EE', 'SI', 'FC', 'BI', 'UB',
            'ATT', 'SE', 'TRU', 'ANX', 'TRA', 'AUT'
        ]

        self.ai_types = [
            'Generative AI',
            'Predictive AI',
            'Decision Support AI',
            'Conversational AI',
            'General AI'
        ]

        self.regions = [
            'North America',
            'Europe',
            'Asia',
            'Middle East',
            'Latin America',
            'Africa',
            'Oceania',
            'Multi-region'
        ]

        self.study_designs = [
            'Cross-sectional survey',
            'Longitudinal survey',
            'Experimental',
            'Quasi-experimental',
            'Mixed methods'
        ]

        self.industries = [
            'Education',
            'Healthcare',
            'Business/Finance',
            'Manufacturing',
            'Technology',
            'Government',
            'Mixed/General'
        ]

    def create_sheet1_metadata(self, wb: Workbook):
        """Create Sheet 1: Study Metadata."""
        ws = wb.create_sheet("Study_Metadata", 0)

        # Headers
        headers = [
            'Study_ID',
            'Authors',
            'Year',
            'Title',
            'Journal',
            'DOI',
            'Sample_Size',
            'Response_Rate',
            'Country',
            'Region',
            'Industry_Sector',
            'AI_Type',
            'Study_Design',
            'Data_Collection_Method',
            'Mandatory_Voluntary',
            'Notes',
            'Coder_Initials'
        ]

        ws.append(headers)

        # Format header row
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)

        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        # Set column widths
        column_widths = {
            'A': 15,  # Study_ID
            'B': 25,  # Authors
            'C': 8,   # Year
            'D': 35,  # Title
            'E': 20,  # Journal
            'F': 20,  # DOI
            'G': 12,  # Sample_Size
            'H': 12,  # Response_Rate
            'I': 15,  # Country
            'J': 18,  # Region
            'K': 18,  # Industry_Sector
            'L': 20,  # AI_Type
            'M': 22,  # Study_Design
            'N': 25,  # Data_Collection_Method
            'O': 18,  # Mandatory_Voluntary
            'P': 30,  # Notes
            'Q': 12   # Coder_Initials
        }

        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width

        # Add data validation for categorical fields
        # Region dropdown (J column)
        region_dv = DataValidation(type="list", formula1=f'"{",".join(self.regions)}"')
        ws.add_data_validation(region_dv)
        region_dv.add(f'J2:J1000')

        # AI Type dropdown (L column)
        ai_type_dv = DataValidation(type="list", formula1=f'"{",".join(self.ai_types)}"')
        ws.add_data_validation(ai_type_dv)
        ai_type_dv.add(f'L2:L1000')

        # Study Design dropdown (M column)
        design_dv = DataValidation(type="list", formula1=f'"{",".join(self.study_designs)}"')
        ws.add_data_validation(design_dv)
        design_dv.add(f'M2:M1000')

        # Industry Sector dropdown (K column)
        industry_dv = DataValidation(type="list", formula1=f'"{",".join(self.industries)}"')
        ws.add_data_validation(industry_dv)
        industry_dv.add(f'K2:K1000')

        # Mandatory/Voluntary dropdown (O column)
        mandatory_dv = DataValidation(type="list", formula1='"Mandatory,Voluntary,Mixed,Not specified"')
        ws.add_data_validation(mandatory_dv)
        mandatory_dv.add(f'O2:O1000')

        logger.info("Created Sheet 1: Study_Metadata")

    def create_sheet2_correlations(self, wb: Workbook):
        """Create Sheet 2: Correlation Matrix."""
        ws = wb.create_sheet("Correlation_Matrix", 1)

        # Headers
        headers = [
            'Study_ID',
            'Construct_1',
            'Construct_2',
            'Pearson_r',
            'Sample_Size_n',
            'Significance_Level',
            'Table_Number',
            'Page_Number',
            'Source_Type',
            'Notes',
            'Verified'
        ]

        ws.append(headers)

        # Format header row
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)

        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        # Set column widths
        column_widths = {
            'A': 15,  # Study_ID
            'B': 15,  # Construct_1
            'C': 15,  # Construct_2
            'D': 12,  # Pearson_r
            'E': 15,  # Sample_Size_n
            'F': 18,  # Significance_Level
            'G': 12,  # Table_Number
            'H': 12,  # Page_Number
            'I': 15,  # Source_Type
            'J': 30,  # Notes
            'K': 10   # Verified
        }

        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width

        # Add data validation
        # Construct dropdowns (B and C columns)
        construct_dv = DataValidation(type="list", formula1=f'"{",".join(self.constructs)}"')
        ws.add_data_validation(construct_dv)
        construct_dv.add(f'B2:B1000')
        construct_dv.add(f'C2:C1000')

        # Source Type dropdown (I column)
        source_dv = DataValidation(type="list", formula1='"Table,Text,Appendix,Supplementary"')
        ws.add_data_validation(source_dv)
        source_dv.add(f'I2:I1000')

        # Verified dropdown (K column)
        verified_dv = DataValidation(type="list", formula1='"Yes,No,Uncertain"')
        ws.add_data_validation(verified_dv)
        verified_dv.add(f'K2:K1000')

        logger.info("Created Sheet 2: Correlation_Matrix")

    def create_sheet3_construct_mapping(self, wb: Workbook):
        """Create Sheet 3: Construct Mapping."""
        ws = wb.create_sheet("Construct_Mapping", 2)

        # Headers
        headers = [
            'Study_ID',
            'Original_Construct_Name',
            'Original_Definition',
            'Mapped_Construct',
            'Mapping_Confidence',
            'Number_of_Items',
            'Reliability_Alpha',
            'Source_Scale',
            'Notes',
            'Verified'
        ]

        ws.append(headers)

        # Format header row
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)

        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        # Set column widths
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['C'].width = 35
        ws.column_dimensions['D'].width = 18
        ws.column_dimensions['E'].width = 18
        ws.column_dimensions['F'].width = 15
        ws.column_dimensions['G'].width = 15
        ws.column_dimensions['H'].width = 20
        ws.column_dimensions['I'].width = 30
        ws.column_dimensions['J'].width = 10

        # Data validation
        construct_dv = DataValidation(type="list", formula1=f'"{",".join(self.constructs)}"')
        ws.add_data_validation(construct_dv)
        construct_dv.add(f'D2:D1000')

        confidence_dv = DataValidation(type="list", formula1='"Exact,High,Moderate,Low"')
        ws.add_data_validation(confidence_dv)
        confidence_dv.add(f'E2:E1000')

        verified_dv = DataValidation(type="list", formula1='"Yes,No"')
        ws.add_data_validation(verified_dv)
        verified_dv.add(f'J2:J1000')

        logger.info("Created Sheet 3: Construct_Mapping")

    def create_sheet4_moderators(self, wb: Workbook):
        """Create Sheet 4: Moderator Variables."""
        ws = wb.create_sheet("Moderator_Variables", 3)

        # Headers
        headers = [
            'Study_ID',
            'Age_Mean',
            'Age_SD',
            'Gender_Pct_Female',
            'Education_Level',
            'Experience_Mean_Years',
            'Hofstede_Individualism',
            'Notes'
        ]

        ws.append(headers)

        # Format header
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)

        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        # Set column widths
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 12
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 18
        ws.column_dimensions['E'].width = 20
        ws.column_dimensions['F'].width = 20
        ws.column_dimensions['G'].width = 22
        ws.column_dimensions['H'].width = 30

        logger.info("Created Sheet 4: Moderator_Variables")

    def create_sheet5_provenance(self, wb: Workbook):
        """Create Sheet 5: AI Extraction Provenance."""
        ws = wb.create_sheet("AI_Extraction_Provenance", 4)

        # Headers
        headers = [
            'Study_ID',
            'Field_Name',
            'AI_Extracted_Value',
            'AI_Model',
            'Confidence_Score',
            'Extraction_Date',
            'Human_Verified_Value',
            'Agreement',
            'Notes'
        ]

        ws.append(headers)

        # Format header
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)

        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        # Set column widths
        for col, width in zip(['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I'],
                             [15, 20, 25, 20, 15, 15, 25, 12, 30]):
            ws.column_dimensions[col].width = width

        # Data validation for Agreement
        agreement_dv = DataValidation(type="list", formula1='"Match,Mismatch,Partial"')
        ws.add_data_validation(agreement_dv)
        agreement_dv.add(f'H2:H1000')

        logger.info("Created Sheet 5: AI_Extraction_Provenance")

    def create_sheet6_quality(self, wb: Workbook):
        """Create Sheet 6: Quality Assessment."""
        ws = wb.create_sheet("Quality_Assessment", 5)

        # Headers
        headers = [
            'Study_ID',
            'Sample_Representativeness',
            'Response_Rate_Adequate',
            'Measures_Validated',
            'Common_Method_Bias_Addressed',
            'Statistical_Power',
            'Overall_Quality_Score'
        ]

        ws.append(headers)

        # Format header
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)

        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        # Set column widths
        for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G']:
            ws.column_dimensions[col].width = 25

        # Data validation (Yes/No/Unclear for quality items)
        quality_dv = DataValidation(type="list", formula1='"Yes,No,Unclear"')
        ws.add_data_validation(quality_dv)
        for col in ['B', 'C', 'D', 'E', 'F']:
            quality_dv.add(f'{col}2:{col}1000')

        # Overall quality score (1-5)
        score_dv = DataValidation(type="whole", operator="between", formula1="1", formula2="5")
        ws.add_data_validation(score_dv)
        score_dv.add(f'G2:G1000')

        logger.info("Created Sheet 6: Quality_Assessment")

    def create_template(self, output_path: Path):
        """
        Create complete coding template.

        Args:
            output_path: Path to save Excel template
        """
        logger.info("Creating coding template...")

        # Create workbook
        wb = Workbook()

        # Remove default sheet
        if 'Sheet' in wb.sheetnames:
            wb.remove(wb['Sheet'])

        # Create all sheets
        self.create_sheet1_metadata(wb)
        self.create_sheet2_correlations(wb)
        self.create_sheet3_construct_mapping(wb)
        self.create_sheet4_moderators(wb)
        self.create_sheet5_provenance(wb)
        self.create_sheet6_quality(wb)

        # Save workbook
        wb.save(output_path)
        logger.info(f"Template saved to {output_path}")


def main():
    """Main entry point."""
    # Setup logging
    logging.basicConfig(
        level=logging.INFO,
        format='%(asctime)s - %(levelname)s - %(message)s'
    )

    # Create output directory
    output_dir = Path('data/templates')
    output_dir.mkdir(parents=True, exist_ok=True)

    # Generate template
    generator = CodingTemplateGenerator()
    output_path = output_dir / 'AI_Adoption_MASEM_Coding_v1.xlsx'

    generator.create_template(output_path)

    logger.info("Template generation complete!")


if __name__ == "__main__":
    main()
