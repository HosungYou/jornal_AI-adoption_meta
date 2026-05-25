#!/usr/bin/env python3
"""Create combined Phase 1+2 human-coder disagreement outputs."""

from __future__ import annotations

import csv
from collections import Counter, defaultdict
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


REPO = Path(__file__).resolve().parents[4]
TODAY = date(2026, 5, 25)
STAMP = "20260525"

PHASE1_BASE = REPO / "data/04_extraction/01_raw_human_coder_data_freeze/phase1/coder_packages"
PHASE2_BASE = REPO / "data/04_extraction/01_raw_human_coder_data_freeze/phase2/freeze_candidates"
OUT_BASE = REPO / "data/04_extraction/02_pre_adjudication_disagreement/combined"
DERIVED = OUT_BASE / "derived"


@dataclass(frozen=True)
class PairSpec:
    phase_block: str
    pair: str
    coder_a: str
    coder_b: str
    path_a: Path
    path_b: Path
    phase_label_contains: str


PAIRS = [
    PairSpec(
        "phase1",
        "Pair A",
        "R1",
        "R2",
        PHASE1_BASE / "R1/AI_Adoption_MASEM_Coding_v3_R1.xlsx",
        PHASE1_BASE / "R2/AI_Adoption_MASEM_Coding_v3_R2.xlsx",
        "Phase 1: Pair A",
    ),
    PairSpec(
        "phase1",
        "Pair B",
        "R3",
        "R4",
        PHASE1_BASE / "R3/AI_Adoption_MASEM_Coding_v3_R3.xlsx",
        PHASE1_BASE / "R4/AI_Adoption_MASEM_Coding_v3_R4.xlsx",
        "Phase 1: Pair B",
    ),
    PairSpec(
        "phase2",
        "Pair C",
        "R1",
        "R4",
        PHASE2_BASE / "R1/AI_Adoption_MASEM_Coding_v3_R1_Phase0_1_2_freeze_candidate_20260525.xlsx",
        PHASE2_BASE / "R4/AI_Adoption_MASEM_Coding_v3_R4_Phase0_1_2_freeze_candidate_20260525.xlsx",
        "Phase 2: Pair C",
    ),
    PairSpec(
        "phase2",
        "Pair D",
        "R2",
        "R3",
        PHASE2_BASE / "R2/AI_Adoption_MASEM_Coding_v3_R2_Phase0_1_2_freeze_candidate_20260525.xlsx",
        PHASE2_BASE / "R3/AI_Adoption_MASEM_Coding_v3_R3_Phase0_1_2_latest_consensus_freeze_candidate_20260525.xlsx",
        "Phase 2: Pair D",
    ),
]

METADATA_EXCLUDE_FIELDS = {"study_id", "human_coder", "coding_date", "coding_phase"}


def clean(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def numeric(value: Any) -> float | None:
    if value in (None, ""):
        return None
    try:
        return float(value)
    except Exception:
        return None


def read_assignments(wb, phase_label_contains: str) -> dict[str, dict[str, str]]:
    ws = wb["ASSIGNMENT"]
    out: dict[str, dict[str, str]] = {}
    for row in ws.iter_rows(min_row=2, max_col=9, values_only=True):
        sid, title, year, doi, phase, pdf, status, flag, notes = row
        if not sid or not phase or phase_label_contains not in str(phase):
            continue
        out[str(sid)] = {
            "study_id": clean(sid),
            "title": clean(title),
            "year": clean(year),
            "doi": clean(doi),
            "phase": clean(phase),
            "pdf": clean(pdf),
            "assignment_status": clean(status),
            "assignment_flag": clean(flag),
            "assignment_notes": clean(notes),
        }
    return out


def read_exclusion_ids(wb) -> set[str]:
    if "EXCLUSION_LOG" not in wb.sheetnames:
        return set()
    ids = set()
    for row in wb["EXCLUSION_LOG"].iter_rows(min_row=2, max_col=1, values_only=True):
        if row[0]:
            ids.add(str(row[0]))
    return ids


def read_metadata(wb, assigned_ids: set[str]) -> dict[tuple[str, str], str]:
    ws = wb["STUDY_METADATA"]
    headers = [clean(ws.cell(1, c).value) for c in range(1, ws.max_column + 1)]
    out: dict[tuple[str, str], str] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        sid = row[0]
        if not sid or str(sid) not in assigned_ids:
            continue
        for idx, field in enumerate(headers):
            if not field or field in METADATA_EXCLUDE_FIELDS:
                continue
            out[(str(sid), field)] = clean(row[idx] if idx < len(row) else None)
    return out


def read_correlations(wb, assigned_ids: set[str]) -> dict[tuple[str, str, str], dict[str, str]]:
    ws = wb["CORRELATIONS"]
    out: dict[tuple[str, str, str], dict[str, str]] = {}
    for row in ws.iter_rows(min_row=2, max_col=13, values_only=True):
        sid, c1, c2 = row[0], row[1], row[2]
        if not sid or str(sid) not in assigned_ids or not c1 or not c2:
            continue
        r_value, r_source, beta = row[3], row[4], row[5]
        if r_value is None and beta is None:
            continue
        key = (str(sid), clean(c1), clean(c2))
        out[key] = {
            "value": clean(r_value if r_value is not None else beta),
            "r_value": clean(r_value),
            "r_source": clean(r_source),
            "original_beta": clean(beta),
            "p_value": clean(row[6]),
            "significance": clean(row[7]),
            "study_label_1": clean(row[8]),
            "study_label_2": clean(row[9]),
            "source_location": clean(row[10]),
            "flag": clean(row[11]),
            "notes": clean(row[12]),
        }
    return out


def computed_status(assignment_status: str, has_values: bool, is_excluded: bool) -> str:
    if assignment_status:
        return assignment_status
    if has_values:
        return "done_inferred"
    if is_excluded:
        return "excluded_inferred"
    return "blank"


def build_outputs() -> tuple[list[dict[str, str]], list[dict[str, str]], list[dict[str, str]], list[dict[str, str]]]:
    value_rows: list[dict[str, str]] = []
    diff_rows: list[dict[str, str]] = []
    study_rows: list[dict[str, str]] = []
    manifest_rows: list[dict[str, str]] = []

    for spec in PAIRS:
        if not spec.path_a.exists():
            raise FileNotFoundError(spec.path_a)
        if not spec.path_b.exists():
            raise FileNotFoundError(spec.path_b)

        wb_a = load_workbook(spec.path_a, read_only=False, data_only=True)
        wb_b = load_workbook(spec.path_b, read_only=False, data_only=True)
        assignments_a = read_assignments(wb_a, spec.phase_label_contains)
        assignments_b = read_assignments(wb_b, spec.phase_label_contains)
        assigned_ids = set(assignments_a) | set(assignments_b)
        metadata_a = read_metadata(wb_a, set(assignments_a))
        metadata_b = read_metadata(wb_b, set(assignments_b))
        correlations_a = read_correlations(wb_a, set(assignments_a))
        correlations_b = read_correlations(wb_b, set(assignments_b))
        exclusions_a = read_exclusion_ids(wb_a)
        exclusions_b = read_exclusion_ids(wb_b)

        for coder, assignments, metadata, correlations, exclusions, path in [
            (spec.coder_a, assignments_a, metadata_a, correlations_a, exclusions_a, spec.path_a),
            (spec.coder_b, assignments_b, metadata_b, correlations_b, exclusions_b, spec.path_b),
        ]:
            nonempty_studies = {sid for sid, _field in metadata if metadata[(sid, _field)]}
            nonempty_studies |= {sid for sid, _c1, _c2 in correlations}
            manifest_rows.append(
                {
                    "phase_block": spec.phase_block,
                    "pair": spec.pair,
                    "coder": coder,
                    "workbook": str(path.relative_to(REPO)),
                    "assigned_studies": str(len(assignments)),
                    "nonempty_studies": str(len(nonempty_studies & set(assignments))),
                    "metadata_values": str(sum(1 for v in metadata.values() if v)),
                    "correlation_values": str(len(correlations)),
                    "exclusion_log_ids_in_assignment": str(len(set(assignments) & exclusions)),
                }
            )
            for (sid, field), value in sorted(metadata.items()):
                if not value:
                    continue
                meta = assignments.get(sid, {})
                value_rows.append(
                    {
                        "phase_block": spec.phase_block,
                        "pair": spec.pair,
                        "study_id": sid,
                        "title": meta.get("title", ""),
                        "year": meta.get("year", ""),
                        "doi": meta.get("doi", ""),
                        "pdf": meta.get("pdf", ""),
                        "coder": coder,
                        "field_family": "metadata",
                        "field_key": field,
                        "construct_1": "",
                        "construct_2": "",
                        "value": value,
                        "value_source": "",
                        "source_location": "",
                        "flag": "",
                        "notes": "",
                    }
                )
            for (sid, c1, c2), item in sorted(correlations.items()):
                meta = assignments.get(sid, {})
                value_rows.append(
                    {
                        "phase_block": spec.phase_block,
                        "pair": spec.pair,
                        "study_id": sid,
                        "title": meta.get("title", ""),
                        "year": meta.get("year", ""),
                        "doi": meta.get("doi", ""),
                        "pdf": meta.get("pdf", ""),
                        "coder": coder,
                        "field_family": "correlation",
                        "field_key": f"{c1}<->{c2}",
                        "construct_1": c1,
                        "construct_2": c2,
                        "value": item["value"],
                        "value_source": item["r_source"] or ("original_beta" if item["original_beta"] else ""),
                        "source_location": item["source_location"],
                        "flag": item["flag"],
                        "notes": item["notes"],
                    }
                )

        study_counters: dict[str, Counter] = defaultdict(Counter)

        for sid, field in sorted(set(metadata_a) | set(metadata_b)):
            left = metadata_a.get((sid, field), "")
            right = metadata_b.get((sid, field), "")
            if left == right:
                continue
            meta = assignments_a.get(sid) or assignments_b.get(sid) or {}
            diff_rows.append(
                {
                    "phase_block": spec.phase_block,
                    "pair": spec.pair,
                    "study_id": sid,
                    "title": meta.get("title", ""),
                    "year": meta.get("year", ""),
                    "doi": meta.get("doi", ""),
                    "pdf": meta.get("pdf", ""),
                    "field_family": "metadata",
                    "field_key": field,
                    "construct_1": "",
                    "construct_2": "",
                    "coder_a": spec.coder_a,
                    "coder_a_value": left,
                    "coder_a_source": "",
                    "coder_a_notes": "",
                    "coder_b": spec.coder_b,
                    "coder_b_value": right,
                    "coder_b_source": "",
                    "coder_b_notes": "",
                    "abs_difference": "",
                    "mismatch_type": "metadata_diff",
                }
            )
            study_counters[sid]["metadata_diff"] += 1

        for key in sorted(set(correlations_a) | set(correlations_b)):
            sid, c1, c2 = key
            left = correlations_a.get(key)
            right = correlations_b.get(key)
            meta = assignments_a.get(sid) or assignments_b.get(sid) or {}
            if left is None or right is None:
                mismatch = "one_coder_only"
                diff = ""
            else:
                ln = numeric(left["value"])
                rn = numeric(right["value"])
                diff = "" if ln is None or rn is None else str(abs(ln - rn))
                mismatch = (
                    "numeric_or_source_diff"
                    if left["value"] != right["value"] or left["r_source"] != right["r_source"]
                    else "same"
                )
            if mismatch == "same":
                continue
            diff_rows.append(
                {
                    "phase_block": spec.phase_block,
                    "pair": spec.pair,
                    "study_id": sid,
                    "title": meta.get("title", ""),
                    "year": meta.get("year", ""),
                    "doi": meta.get("doi", ""),
                    "pdf": meta.get("pdf", ""),
                    "field_family": "correlation",
                    "field_key": f"{c1}<->{c2}",
                    "construct_1": c1,
                    "construct_2": c2,
                    "coder_a": spec.coder_a,
                    "coder_a_value": "" if left is None else left["value"],
                    "coder_a_source": "" if left is None else left["r_source"],
                    "coder_a_notes": "" if left is None else left["notes"],
                    "coder_b": spec.coder_b,
                    "coder_b_value": "" if right is None else right["value"],
                    "coder_b_source": "" if right is None else right["r_source"],
                    "coder_b_notes": "" if right is None else right["notes"],
                    "abs_difference": diff,
                    "mismatch_type": mismatch,
                }
            )
            study_counters[sid][mismatch] += 1

        for sid in sorted(assigned_ids):
            meta = assignments_a.get(sid) or assignments_b.get(sid) or {}
            a_has_values = any(k[0] == sid for k in metadata_a) or any(k[0] == sid for k in correlations_a)
            b_has_values = any(k[0] == sid for k in metadata_b) or any(k[0] == sid for k in correlations_b)
            status_a = computed_status(assignments_a.get(sid, {}).get("assignment_status", ""), a_has_values, sid in exclusions_a)
            status_b = computed_status(assignments_b.get(sid, {}).get("assignment_status", ""), b_has_values, sid in exclusions_b)
            counter = study_counters.get(sid, Counter())
            include = bool(counter) or status_a not in {"done", "done_inferred"} or status_b not in {"done", "done_inferred"}
            if not include:
                continue
            notes = []
            if counter["metadata_diff"]:
                notes.append("metadata differences")
            if counter["one_coder_only"]:
                notes.append("one-coder-only correlation values")
            if counter["numeric_or_source_diff"]:
                notes.append("numeric/source correlation differences")
            if "review_source" in {status_a, status_b}:
                notes.append("source review")
            if status_a.startswith("excluded") or status_b.startswith("excluded"):
                notes.append("exclusion/source-check")
            study_rows.append(
                {
                    "phase_block": spec.phase_block,
                    "pair": spec.pair,
                    "study_id": sid,
                    "title": meta.get("title", ""),
                    "year": meta.get("year", ""),
                    "doi": meta.get("doi", ""),
                    "pdf": meta.get("pdf", ""),
                    "coder_a": spec.coder_a,
                    "coder_a_status": status_a,
                    "coder_a_notes": assignments_a.get(sid, {}).get("assignment_notes", ""),
                    "coder_b": spec.coder_b,
                    "coder_b_status": status_b,
                    "coder_b_notes": assignments_b.get(sid, {}).get("assignment_notes", ""),
                    "n_difference_rows": str(sum(counter.values())),
                    "n_metadata_diff": str(counter["metadata_diff"]),
                    "n_one_coder_only": str(counter["one_coder_only"]),
                    "n_numeric_or_source_diff": str(counter["numeric_or_source_diff"]),
                    "review_note": "; ".join(notes),
                }
            )

        wb_a.close()
        wb_b.close()

    return manifest_rows, value_rows, diff_rows, study_rows


def write_csv(path: Path, rows: list[dict[str, str]], fieldnames: list[str]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames, lineterminator="\n")
        writer.writeheader()
        writer.writerows(rows)


def main() -> None:
    manifest_rows, value_rows, diff_rows, study_rows = build_outputs()

    write_csv(
        DERIVED / f"combined_coding_manifest_{STAMP}.csv",
        manifest_rows,
        [
            "phase_block",
            "pair",
            "coder",
            "workbook",
            "assigned_studies",
            "nonempty_studies",
            "metadata_values",
            "correlation_values",
            "exclusion_log_ids_in_assignment",
        ],
    )
    write_csv(
        DERIVED / f"combined_coder_values_long_{STAMP}.csv",
        value_rows,
        [
            "phase_block",
            "pair",
            "study_id",
            "title",
            "year",
            "doi",
            "pdf",
            "coder",
            "field_family",
            "field_key",
            "construct_1",
            "construct_2",
            "value",
            "value_source",
            "source_location",
            "flag",
            "notes",
        ],
    )
    write_csv(
        DERIVED / f"combined_pairwise_disagreement_long_{STAMP}.csv",
        diff_rows,
        [
            "phase_block",
            "pair",
            "study_id",
            "title",
            "year",
            "doi",
            "pdf",
            "field_family",
            "field_key",
            "construct_1",
            "construct_2",
            "coder_a",
            "coder_a_value",
            "coder_a_source",
            "coder_a_notes",
            "coder_b",
            "coder_b_value",
            "coder_b_source",
            "coder_b_notes",
            "abs_difference",
            "mismatch_type",
        ],
    )
    summary = Counter((row["phase_block"], row["pair"], row["field_family"], row["mismatch_type"]) for row in diff_rows)
    summary_rows = [
        {
            "phase_block": phase,
            "pair": pair,
            "field_family": family,
            "mismatch_type": mismatch,
            "n": str(n),
        }
        for (phase, pair, family, mismatch), n in sorted(summary.items())
    ]
    write_csv(
        DERIVED / f"combined_pairwise_disagreement_summary_{STAMP}.csv",
        summary_rows,
        ["phase_block", "pair", "field_family", "mismatch_type", "n"],
    )
    write_csv(
        DERIVED / f"combined_study_review_queue_{STAMP}.csv",
        study_rows,
        [
            "phase_block",
            "pair",
            "study_id",
            "title",
            "year",
            "doi",
            "pdf",
            "coder_a",
            "coder_a_status",
            "coder_a_notes",
            "coder_b",
            "coder_b_status",
            "coder_b_notes",
            "n_difference_rows",
            "n_metadata_diff",
            "n_one_coder_only",
            "n_numeric_or_source_diff",
            "review_note",
        ],
    )
    priority_rows = [
        row
        for row in study_rows
        if row["n_one_coder_only"] != "0"
        or row["n_numeric_or_source_diff"] != "0"
        or "source review" in row["review_note"]
        or "exclusion/source-check" in row["review_note"]
    ]
    write_csv(
        DERIVED / f"combined_correlation_review_queue_{STAMP}.csv",
        priority_rows,
        [
            "phase_block",
            "pair",
            "study_id",
            "title",
            "year",
            "doi",
            "pdf",
            "coder_a",
            "coder_a_status",
            "coder_a_notes",
            "coder_b",
            "coder_b_status",
            "coder_b_notes",
            "n_difference_rows",
            "n_metadata_diff",
            "n_one_coder_only",
            "n_numeric_or_source_diff",
            "review_note",
        ],
    )

    OUT_BASE.mkdir(parents=True, exist_ok=True)
    (OUT_BASE / "README.md").write_text(
        "\n".join(
            [
                "# Combined Phase 1+2 Pre-Adjudication Disagreement",
                "",
                f"Generated: {TODAY.isoformat()}",
                "",
                "These outputs combine Phase 1 Pair A/B and Phase 2 Pair C/D human-coder values for the full 213-study Paper B validation corpus.",
                "",
                "- Phase 1: Pair A (R1+R2) and Pair B (R3+R4), 100 studies.",
                "- Phase 2: Pair C (R1+R4) and Pair D (R2+R3), 113 studies.",
                "- Phase 0 calibration rows and historical `Phase 2: Single` rows in the Phase 1 workbooks are excluded.",
                "",
                "## Derived Files",
                "",
                f"- `derived/combined_coding_manifest_{STAMP}.csv`: coder workbook coverage by phase and pair.",
                f"- `derived/combined_coder_values_long_{STAMP}.csv`: all nonempty metadata and correlation coding values used for pre-adjudication review.",
                f"- `derived/combined_pairwise_disagreement_long_{STAMP}.csv`: metadata and correlation disagreement rows.",
                f"- `derived/combined_pairwise_disagreement_summary_{STAMP}.csv`: counts by phase, pair, field family, and mismatch type.",
                f"- `derived/combined_study_review_queue_{STAMP}.csv`: all study-level review rows for meetings, including metadata-only differences.",
                f"- `derived/combined_correlation_review_queue_{STAMP}.csv`: meeting-first queue for correlation/status/source-review issues, excluding metadata-only differences.",
                "",
                "These are pre-adjudication artifacts. They identify where human coders differ before source-document adjudication. They are not the source-anchored adjudicated human reference standard.",
            ]
        ),
        encoding="utf-8",
    )

    print("Prepared combined Phase 1+2 coding values and disagreement outputs.")
    print(f"manifest_rows={len(manifest_rows)}")
    print(f"value_rows={len(value_rows)}")
    print(f"diff_rows={len(diff_rows)}")
    print(f"study_queue_rows={len(study_rows)}")
    print(f"priority_queue_rows={len(priority_rows)}")


if __name__ == "__main__":
    main()
